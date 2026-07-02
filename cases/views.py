"""
Cases App — Views
==================
Client-facing (public) and Admin/Staff (auth-required) views.

Public views:
    - ``client_dashboard``  — grid of CaseCategory cards.
    - ``create_case``       — form to submit a new case.
    - ``case_submitted``    — confirmation page.

Staff views (login + role_access required):
    - ``case_list``         — filterable list / kanban of all cases.
    - ``case_detail``       — split-panel: chat thread + RCA form.
    - ``case_update_rca``   — HTMX partial to save RCA fields.
    - ``case_send_reply``   — HTMX partial to send a staff reply.
    - ``chat_thread``       — HTMX partial to refresh chat messages.
"""
import logging
import re
from functools import wraps

logger = logging.getLogger(__name__)

from ipware import get_client_ip as _get_client_ip
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db.models import Q, F, Value, BooleanField, Case as DBCase, When, Exists, OuterRef
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.core.cache import cache

from core.models import CompanyUnit, Employee, User
from .forms import CaseCreateForm, CaseRCAForm, StaffReplyForm
from .models import (
    Attachment, CaseCategory, CaseRecord, ChangeRequestApproval,
    ChangeRequestDocument, Message, CaseComment, CaseAuditLog,
    RCATemplate, DocumentTemplate, DocumentTemplateField, FormSubmission,
)

from licensing.decorators import feature_required


# =====================================================================
# Access Control Decorators
# =====================================================================

def staff_required(view_func):
    """
    Decorator to ensure user is authenticated AND has staff-level
    role_access (SuperAdmin, Manager, SupportDesk, or Auditor).
    Auditor has Read-Only access (GET requests only).
    """
    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        if not hasattr(request.user, "role_access"):
            return HttpResponseForbidden("Access denied.")
        allowed_roles = {
            User.RoleAccess.SUPERADMIN,
            User.RoleAccess.MANAGER,
            User.RoleAccess.SUPPORTDESK,
            User.RoleAccess.AUDITOR,
        }
        if request.user.role_access not in allowed_roles:
            return HttpResponseForbidden("Insufficient role access.")
            
        # Enforce read-only for Auditor
        if request.user.role_access == User.RoleAccess.AUDITOR and request.method not in ("GET", "HEAD", "OPTIONS"):
            return HttpResponseForbidden("Access denied. Auditors have read-only access.")
            
        return view_func(request, *args, **kwargs)
    return _wrapped


def _case_is_closed(case, request):
    """Return True and add an error message if the case is closed."""
    if case.status == CaseRecord.Status.CLOSED:
        messages.error(request, "Cannot modify a closed ticket.")
        return True
    return False


def _check_confidential_access(case, user):
    """Return True if the user is allowed to access this case's details."""
    if not getattr(case.category, "is_confidential", False):
        return True
    return user.role_access == "SuperAdmin" or user.can_handle_confidential


def _annotate_confidential_access(qs, user):
    """Annotate a CaseRecord queryset with ``has_confidential_access_flag``."""
    if user.role_access == "SuperAdmin" or user.can_handle_confidential:
        return qs.annotate(
            has_confidential_access_flag=Value(True, output_field=BooleanField()),
        )
    return qs.annotate(
        has_confidential_access_flag=DBCase(
            When(category__is_confidential=False, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        )
    )


def manager_or_admin_required(view_func):
    """Decorator for Manager/SuperAdmin only actions."""
    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        allowed = {User.RoleAccess.SUPERADMIN, User.RoleAccess.MANAGER}
        if request.user.role_access not in allowed:
            return HttpResponseForbidden("Manager or SuperAdmin access required.")
        return view_func(request, *args, **kwargs)
    return _wrapped


# =====================================================================
# Enterprise E-Forms UI Views (Phase 2)
# =====================================================================

@login_required
def my_eforms_list(request):
    """
    Client portal view for users to see their own initiated FormSubmissions.
    """
    submissions = FormSubmission.objects.filter(initiator_user=request.user).order_by('-created_at')
    
    from django.core.paginator import Paginator
    paginator = Paginator(submissions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, "cases/eforms/my_eforms.html", {
        "submissions": page_obj
    })

@login_required
def eform_template_list(request):
    """
    Shows a grid of available E-Form Templates for the user to start a new request.
    """
    templates = DocumentTemplate.objects.filter(is_standalone=True).order_by('title')
    return render(request, "cases/eforms/template_list.html", {
        "templates": templates
    })

@login_required
def eform_initiate(request, template_id):
    """
    Creates a new draft FormSubmission for the selected template.
    """
    template = get_object_or_404(DocumentTemplate, id=template_id, is_standalone=True)
    
    # Determine the first stage
    stages = template.form_stages if isinstance(template.form_stages, list) else []
    first_stage = stages[0] if stages else ""

    import secrets
    submission = FormSubmission.objects.create(
        template=template,
        initiator_user=request.user,
        status=FormSubmission.Status.DRAFT,
        current_stage=first_stage,
        guest_token=secrets.token_urlsafe(32),
        data_payload={}
    )
    return redirect('cases:eform_detail', submission_id=submission.id)

@login_required
def eform_detail(request, submission_id):
    """
    The dynamic form renderer. Shows the fields to fill based on the current stage.
    """
    submission = get_object_or_404(FormSubmission, id=submission_id)
    
    # Security: Ensure user is initiator (later we will add approver logic)
    if submission.initiator_user != request.user and not request.user.is_staff:
        return HttpResponseForbidden("You do not have access to this form.")

    template = submission.template
    fields = template.fields.all().order_by('order')

    if request.method == 'POST':
        # Handle form save
        payload = submission.data_payload or {}
        for field in fields:
            # Only save fields that belong to the current stage or have no assigned stage
            # In a real app, we also enforce validation based on field_type.
            if field.assigned_stage == submission.current_stage or not field.assigned_stage:
                field_key = f"field_{field.id}"
                if field_key in request.POST:
                    payload[field_key] = request.POST.get(field_key)
        
        submission.data_payload = payload
        submission.save()
        
        action = request.POST.get('action')
        if action == 'submit_stage':
            # Logic to move to the next stage
            stages = template.form_stages if isinstance(template.form_stages, list) else []
            try:
                current_idx = stages.index(submission.current_stage)
                if current_idx + 1 < len(stages):
                    submission.current_stage = stages[current_idx + 1]
                    submission.status = FormSubmission.Status.IN_PROGRESS
                    submission.save()
                    messages.success(request, f"Form advanced to stage: {submission.current_stage}")
                else:
                    submission.current_stage = ""
                    submission.status = FormSubmission.Status.PENDING_ESIGN
                    
                    # Phase 3: Trigger PDF Generation
                    from cases.utils_pdf import generate_form_submission_pdf
                    from django.core.files.base import ContentFile
                    pdf_bytes = generate_form_submission_pdf(submission)
                    
                    if pdf_bytes:
                        filename = f"eform_{submission.id.hex[:8]}.pdf"
                        submission.generated_pdf.save(filename, ContentFile(pdf_bytes), save=False)
                        
                        # Hand-off to E-Sign Engine
                        from esign.models import SignatureDocument, Signer
                        sig_doc = SignatureDocument.objects.create(
                            title=f"Signature Request: {template.title}",
                            created_by=submission.initiator_user,
                            status='draft'
                        )
                        sig_doc.original_pdf.save(filename, ContentFile(pdf_bytes), save=True)
                        submission.signature_document = sig_doc
                        
                        # Loop through template fields to find explicit E-Sign Role mappings
                        for f in template.fields.all():
                            if getattr(f, 'is_signer', False):
                                val = str(submission.data_payload.get(f"field_{f.id}", "")).strip()
                                if val:
                                    # If it looks like an email, use it as email
                                    if "@" in val and "." in val:
                                        Signer.objects.create(
                                            document=sig_doc,
                                            order=f.order,
                                            role='signer',
                                            external_name=f.label,
                                            external_email=val
                                        )
                                    else:
                                        # Use it as a name
                                        Signer.objects.create(
                                            document=sig_doc,
                                            order=f.order,
                                            role='signer',
                                            external_name=val,
                                            external_email=""
                                        )
                        
                        # Note: we no longer use signers_map logic since each field is a unique signer
                        
                    submission.save()
                    messages.success(request, "Form fully submitted and is pending E-Sign.")
            except ValueError:
                pass
            
            return redirect('cases:my_eforms')
        
        messages.success(request, "Draft saved successfully.")
        return redirect('cases:eform_detail', submission_id=submission.id)

    payload = submission.data_payload or {}
    for field in fields:
        field.current_value = payload.get(f"field_{field.id}", "")

    # Provide guest link for UI
    guest_link = request.build_absolute_uri(reverse('cases:eform_public_detail', kwargs={'token': submission.guest_token})) if submission.guest_token else None

    return render(request, "cases/eforms/detail.html", {
        "submission": submission,
        "template": template,
        "fields": fields,
        "guest_link": guest_link
    })

def eform_public_detail(request, token):
    """
    Vendor / Public View.
    Allows external parties to fill out the form for the stage they are currently on.
    No login required.
    """
    submission = get_object_or_404(FormSubmission, guest_token=token)
    
    if submission.status in ['completed', 'cancelled', 'pending_esign']:
        return HttpResponse("This form is no longer accepting input.", status=403)

    template = submission.template
    fields = template.fields.all().order_by('order')

    if request.method == 'POST':
        payload = submission.data_payload or {}
        for field in fields:
            # Vendor can only edit fields explicitly assigned to the current stage
            if field.assigned_stage == submission.current_stage:
                field_key = f"field_{field.id}"
                if field_key in request.POST:
                    payload[field_key] = request.POST.get(field_key)
        
        submission.data_payload = payload
        submission.save()
        
        action = request.POST.get('action')
        if action == 'submit_stage':
            stages = template.form_stages if isinstance(template.form_stages, list) else []
            try:
                current_idx = stages.index(submission.current_stage)
                if current_idx + 1 < len(stages):
                    submission.current_stage = stages[current_idx + 1]
                    submission.status = FormSubmission.Status.IN_PROGRESS
                    submission.save()
                else:
                    submission.current_stage = ""
                    submission.status = FormSubmission.Status.PENDING_ESIGN
                    
                    # PDF generation
                    from cases.utils_pdf import generate_form_submission_pdf
                    from django.core.files.base import ContentFile
                    pdf_bytes = generate_form_submission_pdf(submission)
                    if pdf_bytes:
                        filename = f"eform_{submission.id.hex[:8]}.pdf"
                        submission.generated_pdf.save(filename, ContentFile(pdf_bytes), save=False)
                        
                        from esign.models import SignatureDocument
                        sig_doc = SignatureDocument.objects.create(
                            title=f"Signature Request: {template.title}",
                            created_by=submission.initiator_user,
                            status='draft'
                        )
                        sig_doc.original_pdf.save(filename, ContentFile(pdf_bytes), save=True)
                        submission.signature_document = sig_doc
                    submission.save()
                    
                return HttpResponse("Thank you. The form has been submitted to the next stage.", status=200)
            except ValueError:
                pass
        
        messages.success(request, "Draft saved.")
        return redirect('cases:eform_public_detail', token=token)

    payload = submission.data_payload or {}
    for field in fields:
        field.current_value = payload.get(f"field_{field.id}", "")

    return render(request, "cases/eforms/detail.html", {
        "submission": submission,
        "template": template,
        "fields": fields,
        "is_public": True
    })

@staff_required
def form_workflows_list(request):
    """
    Staff dashboard to monitor all active and completed FormSubmissions across the enterprise.
    """
    submissions = FormSubmission.objects.all().order_by('-created_at')
    
    query = request.GET.get('q', '').strip()
    if query:
        submissions = submissions.filter(template__title__icontains=query)
    
    from django.core.paginator import Paginator
    paginator = Paginator(submissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, "cases/eforms/form_workflows.html", {
        "submissions": page_obj,
        "search_query": query
    })


# =====================================================================
# Staff — Analytics Dashboard
# =====================================================================

@staff_required
@feature_required('analytics')
def dashboard(request):
    """
    Analytics dashboard for staff — summary cards, charts, tables.
    Supports date-range, assigned_to, category, source, and priority filters.
    """
    from django.db.models import Count, Avg, Sum, ExpressionWrapper, DurationField
    from django.db.models.functions import TruncDate, TruncMonth, Coalesce
    from django.utils.timezone import localtime, make_aware
    from datetime import datetime, timedelta

    now = timezone.now()
    today = now.date()

    # ── Filter params ──────────────────────────────────────────────
    date_from_str = request.GET.get("date_from", "")
    date_to_str = request.GET.get("date_to", "")
    filter_assigned = request.GET.get("assigned_to", "")
    filter_category = request.GET.get("category", "")
    filter_source = request.GET.get("source", "")
    filter_priority = request.GET.get("priority", "")
    filter_unit = request.GET.get("company_unit", "")

    # Default: first day of current month → today
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except ValueError:
            date_from = today.replace(day=1)
    else:
        date_from = today.replace(day=1)

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            date_to = today
    else:
        date_to = today

    dt_from = make_aware(datetime.combine(date_from, datetime.min.time()))
    dt_to = make_aware(datetime.combine(date_to, datetime.max.time()))

    # ── Base queryset (non-spam, non-sub-ticket) ───────────────────
    qs = CaseRecord.objects.filter(
        is_spam=False,
        master_ticket__isnull=True,
        created_at__range=(dt_from, dt_to),
    ).select_related("category", "assigned_to", "requester", "requester__unit")

    # Apply filters
    if filter_assigned:
        if filter_assigned == "unassigned":
            qs = qs.filter(assigned_to__isnull=True)
        else:
            qs = qs.filter(assigned_to_id=filter_assigned)
    if filter_category:
        qs = qs.filter(category__slug=filter_category)
    if filter_source:
        qs = qs.filter(source=filter_source)
    if filter_priority:
        qs = qs.filter(priority=filter_priority)
    if filter_unit:
        qs = qs.filter(requester__unit_id=filter_unit)

    # ── Also build an "all time active" queryset (not filtered by date) for summary cards ──
    active_qs = CaseRecord.objects.filter(
        is_spam=False,
        master_ticket__isnull=True,
        status__in=[CaseRecord.Status.OPEN, CaseRecord.Status.INVESTIGATING, CaseRecord.Status.PENDING_INFO],
    )
    if filter_assigned:
        if filter_assigned == "unassigned":
            active_qs = active_qs.filter(assigned_to__isnull=True)
        else:
            active_qs = active_qs.filter(assigned_to_id=filter_assigned)
    if filter_category:
        active_qs = active_qs.filter(category__slug=filter_category)
    if filter_source:
        active_qs = active_qs.filter(source=filter_source)
    if filter_priority:
        active_qs = active_qs.filter(priority=filter_priority)
    if filter_unit:
        active_qs = active_qs.filter(requester__unit_id=filter_unit)

    # ── SUMMARY CARDS ──────────────────────────────────────────────
    total_in_range = qs.count()
    total_active = active_qs.count()
    total_unassigned = active_qs.filter(assigned_to__isnull=True).count()

    resolved_in_range = qs.filter(status__in=[CaseRecord.Status.RESOLVED, CaseRecord.Status.CLOSED]).count()
    resolved_today = CaseRecord.objects.filter(
        is_spam=False, master_ticket__isnull=True,
        status__in=[CaseRecord.Status.RESOLVED, CaseRecord.Status.CLOSED],
        updated_at__date=today,
    ).count()

    unread_count = active_qs.filter(has_unread_messages=True).count()

    # SLA breached: resolution_due_at < now AND status not closed/resolved
    sla_breached = active_qs.filter(
        resolution_due_at__lt=now,
    ).exclude(status__in=[CaseRecord.Status.RESOLVED, CaseRecord.Status.CLOSED]).count()

    # SLA at risk: due within 24h
    sla_at_risk = active_qs.filter(
        resolution_due_at__gt=now,
        resolution_due_at__lte=now + timedelta(hours=24),
    ).exclude(status__in=[CaseRecord.Status.RESOLVED, CaseRecord.Status.CLOSED]).count()

    # Avg resolution time (for closed tickets in range)
    closed_in_range = qs.filter(status=CaseRecord.Status.CLOSED, resolution_due_at__isnull=False)
    avg_resolution = None
    if closed_in_range.exists():
        from django.db.models import ExpressionWrapper, DurationField
        avg_dur = closed_in_range.annotate(
            dur=ExpressionWrapper(F("updated_at") - F("created_at"), output_field=DurationField())
        ).aggregate(avg=Avg("dur"))["avg"]
        if avg_dur:
            total_hours = avg_dur.total_seconds() / 3600
            if total_hours >= 24:
                avg_resolution = f"{total_hours / 24:.1f} days"
            else:
                avg_resolution = f"{total_hours:.1f} hrs"

    # ── CHART DATA: Ticket Trend (auto-granularity) ──────────────
    day_span = (date_to - date_from).days + 1
    trend_granularity = request.GET.get("granularity", "")
    if not trend_granularity:
        # Auto: daily ≤ 31 days, weekly ≤ 90, monthly beyond
        if day_span <= 31:
            trend_granularity = "daily"
        elif day_span <= 90:
            trend_granularity = "weekly"
        else:
            trend_granularity = "monthly"

    if trend_granularity == "monthly":
        from django.db.models.functions import TruncMonth as TruncPeriod
        fmt = "%b %Y"
    elif trend_granularity == "weekly":
        from django.db.models.functions import TruncWeek as TruncPeriod
        fmt = "%d %b"
    else:
        TruncPeriod = TruncDate
        fmt = "%d %b"

    trend_data = (
        qs.annotate(period=TruncPeriod("created_at"))
        .values("period")
        .annotate(created=Count("id"))
        .order_by("period")
    )
    resolved_trend = (
        qs.filter(status__in=[CaseRecord.Status.RESOLVED, CaseRecord.Status.CLOSED])
        .annotate(period=TruncPeriod("updated_at"))
        .values("period")
        .annotate(resolved=Count("id"))
        .order_by("period")
    )
    created_map = {}
    for row in trend_data:
        key = row["period"].date() if hasattr(row["period"], "date") else row["period"]
        created_map[key] = row["created"]

    resolved_map = {}
    for r in resolved_trend:
        key = r["period"].date() if hasattr(r["period"], "date") else r["period"]
        resolved_map[key] = r["resolved"]

    # Merge all dates from both datasets so no data points are lost
    all_dates = sorted(set(created_map.keys()) | set(resolved_map.keys()))

    trend_labels = []
    trend_created = []
    trend_resolved = []
    for d in all_dates:
        trend_labels.append(d.strftime(fmt))
        trend_created.append(created_map.get(d, 0))
        trend_resolved.append(resolved_map.get(d, 0))

    # ── CHART DATA: Status Distribution ────────────────────────────
    status_dist = qs.values("status").annotate(count=Count("id")).order_by("status")
    status_labels = []
    status_counts = []
    status_colors = {
        "Open": "#6366f1",
        "Investigating": "#f59e0b",
        "PendingInfo": "#0ea5e9",
        "Resolved": "#10b981",
        "Closed": "#64748b",
    }
    status_bg = []
    for s in status_dist:
        status_labels.append(s["status"])
        status_counts.append(s["count"])
        status_bg.append(status_colors.get(s["status"], "#94a3b8"))

    # ── CHART DATA: Source Channel ─────────────────────────────────
    source_dist = qs.values("source").annotate(count=Count("id")).order_by("-count")
    source_labels = []
    source_counts = []
    source_display = {"EvolutionAPI_WA": "WhatsApp", "Email": "Email", "WebForm": "Web Form"}
    for s in source_dist:
        source_labels.append(source_display.get(s["source"], s["source"]))
        source_counts.append(s["count"])

    # ── CHART DATA: Priority Distribution ──────────────────────────
    priority_dist = qs.values("priority").annotate(count=Count("id")).order_by("-count")
    priority_labels = []
    priority_counts = []
    priority_colors_map = {"Low": "#10b981", "Medium": "#f59e0b", "High": "#ef4444", "Critical": "#7c3aed"}
    priority_bg = []
    for p in priority_dist:
        priority_labels.append(p["priority"])
        priority_counts.append(p["count"])
        priority_bg.append(priority_colors_map.get(p["priority"], "#94a3b8"))

    # ── CHART DATA: Top Categories ─────────────────────────────────
    cat_dist = (
        qs.filter(category__isnull=False)
        .values("category__name")
        .annotate(count=Count("id"))
        .order_by("-count")[:10]
    )
    cat_labels = [c["category__name"] for c in cat_dist]
    cat_counts = [c["count"] for c in cat_dist]

    # ── CHART DATA: Agent Workload ─────────────────────────────────
    agent_dist = (
        active_qs.filter(assigned_to__isnull=False)
        .values("assigned_to__username")
        .annotate(count=Count("id"))
        .order_by("-count")[:10]
    )
    agent_labels = [a["assigned_to__username"] for a in agent_dist]
    agent_counts = [a["count"] for a in agent_dist]

    # ── CHART DATA: By Company Unit ────────────────────────────────
    unit_dist = (
        qs.filter(requester__unit__isnull=False)
        .values("requester__unit__name")
        .annotate(count=Count("id"))
        .order_by("-count")[:10]
    )
    unit_labels = [u["requester__unit__name"] for u in unit_dist]
    unit_counts = [u["count"] for u in unit_dist]

    # ── SLA Performance (closed tickets with SLA data) ─────────────
    closed_with_sla = qs.filter(
        status=CaseRecord.Status.CLOSED,
        resolution_due_at__isnull=False,
    )
    sla_total = closed_with_sla.count()
    sla_met = closed_with_sla.filter(updated_at__lte=F("resolution_due_at")).count()
    sla_pct = round((sla_met / sla_total * 100), 1) if sla_total > 0 else 0

    # ── TABLES: Recent Tickets ─────────────────────────────────────
    recent_tickets = qs.order_by("-created_at")[:8]

    # ── TABLES: SLA At Risk ────────────────────────────────────────
    sla_at_risk_tickets = active_qs.filter(
        resolution_due_at__gt=now,
        resolution_due_at__lte=now + timedelta(hours=24),
    ).order_by("resolution_due_at")[:5]

    sla_breached_tickets = active_qs.filter(
        resolution_due_at__lt=now,
    ).order_by("resolution_due_at")[:5]

    # ── TABLES: Top Requesters ─────────────────────────────────────
    top_requesters = (
        qs.filter(requester__isnull=False)
        .values("requester__full_name", "requester__unit__name")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )

    # ── CHART DATA: First Response Time (FRT) Trend ────────────────
    from django.db.models import Min, Subquery, OuterRef
    frt_subquery = Message.objects.filter(
        case=OuterRef("pk"),
        direction=Message.Direction.OUTBOUND,
    ).order_by("sent_at").values("sent_at")[:1]

    frt_qs = (
        qs.annotate(first_reply_at=Subquery(frt_subquery))
        .filter(first_reply_at__isnull=False)
        .annotate(
            frt_seconds=ExpressionWrapper(
                F("first_reply_at") - F("created_at"),
                output_field=DurationField(),
            )
        )
        .annotate(period=TruncPeriod("created_at"))
        .values("period")
        .annotate(avg_frt=Avg("frt_seconds"))
        .order_by("period")
    )
    frt_labels = []
    frt_values = []
    for row in frt_qs:
        p = row["period"]
        d = p.date() if hasattr(p, "date") else p
        frt_labels.append(d.strftime(fmt))
        frt_hours = row["avg_frt"].total_seconds() / 3600 if row["avg_frt"] else 0
        frt_values.append(round(frt_hours, 1))

    # Overall avg FRT
    overall_frt_qs = (
        qs.annotate(first_reply_at=Subquery(frt_subquery))
        .filter(first_reply_at__isnull=False)
        .annotate(
            frt_seconds=ExpressionWrapper(
                F("first_reply_at") - F("created_at"),
                output_field=DurationField(),
            )
        )
        .aggregate(avg=Avg("frt_seconds"))
    )
    avg_frt = None
    if overall_frt_qs["avg"]:
        frt_h = overall_frt_qs["avg"].total_seconds() / 3600
        avg_frt = f"{frt_h:.1f} hrs" if frt_h < 24 else f"{frt_h / 24:.1f} days"

    # ── CHART DATA: Ticket Aging (active tickets) ────────────────
    aging_buckets = [
        ("< 1 day", 0, 1),
        ("1-3 days", 1, 3),
        ("3-7 days", 3, 7),
        ("7-14 days", 7, 14),
        ("> 14 days", 14, 9999),
    ]
    aging_labels = []
    aging_counts = []
    for label, d_min, d_max in aging_buckets:
        cutoff_max = now - timedelta(days=d_min)
        cutoff_min = now - timedelta(days=d_max)
        cnt = active_qs.filter(created_at__lte=cutoff_max, created_at__gt=cutoff_min).count()
        aging_labels.append(label)
        aging_counts.append(cnt)

    # ── CHART DATA: Peak Hours Heatmap ───────────────────────────
    from django.db.models.functions import ExtractHour, ExtractIsoWeekDay
    peak_data = (
        qs.annotate(
            hour=ExtractHour("created_at"),
            weekday=ExtractIsoWeekDay("created_at"),  # 1=Mon … 7=Sun
        )
        .values("weekday", "hour")
        .annotate(count=Count("id"))
        .order_by("weekday", "hour")
    )
    # Build 7x24 matrix [weekday][hour]
    heatmap_matrix = [[0] * 24 for _ in range(7)]
    for row in peak_data:
        heatmap_matrix[row["weekday"] - 1][row["hour"]] = row["count"]

    # ── CHART DATA: Resolution Rate Trend ────────────────────────
    resrate_labels = []
    resrate_values = []
    for d in all_dates:
        c = created_map.get(d, 0)
        r = resolved_map.get(d, 0)
        resrate_labels.append(d.strftime(fmt))
        pct = round((r / c * 100), 1) if c > 0 else 0
        resrate_values.append(pct)

    # ── CHART DATA: Reopen Rate ──────────────────────────────────
    # Tickets that were resolved/closed but now active again
    reopened_count = CaseRecord.objects.filter(
        is_spam=False,
        master_ticket__isnull=True,
        created_at__range=(dt_from, dt_to),
        status__in=[CaseRecord.Status.OPEN, CaseRecord.Status.INVESTIGATING, CaseRecord.Status.PENDING_INFO],
        edit_permission_status__isnull=False,  # had amendment workflow = was closed before
    ).count()
    reopen_pct = round((reopened_count / total_in_range * 100), 1) if total_in_range > 0 else 0

    # ── TABLE: Stale Tickets (no message activity > 3 days) ──────
    stale_cutoff = now - timedelta(days=3)
    stale_tickets = (
        active_qs.annotate(
            last_msg_at=Subquery(
                Message.objects.filter(case=OuterRef("pk"))
                .order_by("-sent_at")
                .values("sent_at")[:1]
            )
        )
        .filter(
            Q(last_msg_at__lt=stale_cutoff) | Q(last_msg_at__isnull=True, created_at__lt=stale_cutoff)
        )
        .select_related("category", "assigned_to", "requester")
        .order_by("created_at")[:8]
    )

    # ── WORD CLOUD: Tags ────────────────────────────────────────────
    import re
    from collections import Counter

    STOPWORDS = {
        "yang", "dan", "di", "ke", "dari", "ini", "itu", "untuk", "dengan",
        "pada", "adalah", "akan", "sudah", "tidak", "bisa", "ada", "juga",
        "atau", "oleh", "karena", "saat", "telah", "belum", "user", "the",
        "and", "for", "that", "this", "with", "was", "has", "been", "but",
        "are", "were", "had", "have", "can", "will", "its", "may", "our",
        "all", "one", "two", "new", "used", "more", "than", "each", "into",
        "done", "pagi", "selamat", "malam", "siang", "sore", "tim", "gws", "citis",
    }

    tag_counter = Counter()
    for row in qs.filter(tags__isnull=False).exclude(tags="").values_list("tags", flat=True):
        for tag in row.split(","):
            t = tag.strip().lower()
            if t and len(t) > 1:
                tag_counter[t] += 1

    tag_cloud = [[word, count] for word, count in tag_counter.most_common(60)]

    # ── WORD CLOUD: RCA & Solving Steps ──────────────────────────────
    rca_counter = Counter()
    rca_fields = qs.filter(
        Q(root_cause_analysis__isnull=False) | Q(solving_steps__isnull=False)
    ).values_list("root_cause_analysis", "solving_steps")

    word_pattern = re.compile(r'[a-zA-Z\u00C0-\u024F]+')
    for rca, steps in rca_fields:
        for text in (rca or "", steps or ""):
            for word in word_pattern.findall(text.lower()):
                if len(word) > 2 and word not in STOPWORDS:
                    rca_counter[word] += 1

    rca_cloud = [[word, count] for word, count in rca_counter.most_common(80)]

    # ── Filter options for dropdowns ───────────────────────────────
    staff_users = User.objects.filter(
        role_access__in=[User.RoleAccess.SUPERADMIN, User.RoleAccess.MANAGER, User.RoleAccess.SUPPORTDESK]
    ).order_by("username")
    categories = CaseCategory.objects.filter(parent__isnull=True).order_by("name")
    all_company_units = CompanyUnit.objects.order_by("code")

    # ── MAP DATA: Company Units with geolocation + ticket counts ───
    import json
    map_case_filter = Q(
        employees__cases__is_spam=False,
        employees__cases__master_ticket__isnull=True,
        employees__cases__created_at__range=(dt_from, dt_to),
    )
    if filter_assigned:
        if filter_assigned == "unassigned":
            map_case_filter &= Q(employees__cases__assigned_to__isnull=True)
        else:
            map_case_filter &= Q(employees__cases__assigned_to_id=filter_assigned)
    if filter_category:
        map_case_filter &= Q(employees__cases__category__slug=filter_category)
    if filter_source:
        map_case_filter &= Q(employees__cases__source=filter_source)
    if filter_priority:
        map_case_filter &= Q(employees__cases__priority=filter_priority)

    map_units_qs = (
        CompanyUnit.objects
        .filter(latitude__isnull=False, longitude__isnull=False)
        .annotate(ticket_count=Count('employees__cases', filter=map_case_filter, distinct=True))
        .order_by('-ticket_count')
    )
    # When a specific unit is selected, show only that unit on the map
    if filter_unit:
        map_units_qs = map_units_qs.filter(id=filter_unit)
    map_units_list = list(map_units_qs)
    map_units_json = json.dumps([
        {
            'id': str(u.id),
            'name': u.name,
            'code': u.code,
            'city': u.city,
            'province': u.province,
            'lat': float(u.latitude),
            'lng': float(u.longitude),
            'ticket_count': u.ticket_count,
        }
        for u in map_units_list
    ])
    map_total_units = len(map_units_list)
    map_covered_tickets = sum(u.ticket_count for u in map_units_list)

    context = {
        # Summary cards
        "total_in_range": total_in_range,
        "total_active": total_active,
        "total_unassigned": total_unassigned,
        "resolved_in_range": resolved_in_range,
        "resolved_today": resolved_today,
        "unread_count": unread_count,
        "sla_breached": sla_breached,
        "sla_at_risk": sla_at_risk,
        "avg_resolution": avg_resolution,
        # Charts (JSON for Chart.js)
        "trend_labels": json.dumps(trend_labels),
        "trend_created": json.dumps(trend_created),
        "trend_resolved": json.dumps(trend_resolved),
        "status_labels": json.dumps(status_labels),
        "status_counts": json.dumps(status_counts),
        "status_bg": json.dumps(status_bg),
        "source_labels": json.dumps(source_labels),
        "source_counts": json.dumps(source_counts),
        "priority_labels": json.dumps(priority_labels),
        "priority_counts": json.dumps(priority_counts),
        "priority_bg": json.dumps(priority_bg),
        "cat_labels": json.dumps(cat_labels),
        "cat_counts": json.dumps(cat_counts),
        "agent_labels": json.dumps(agent_labels),
        "agent_counts": json.dumps(agent_counts),
        "unit_labels": json.dumps(unit_labels),
        "unit_counts": json.dumps(unit_counts),
        "sla_pct": sla_pct,
        "sla_met": sla_met,
        "sla_total": sla_total,
        # Tables
        "recent_tickets": recent_tickets,
        "sla_at_risk_tickets": sla_at_risk_tickets,
        "sla_breached_tickets": sla_breached_tickets,
        "top_requesters": top_requesters,
        # Filters
        "staff_users": staff_users,
        "categories": categories,
        "filter_date_from": date_from.strftime("%Y-%m-%d"),
        "filter_date_to": date_to.strftime("%Y-%m-%d"),
        "filter_assigned": filter_assigned,
        "filter_category": filter_category,
        "filter_source": filter_source,
        "filter_priority": filter_priority,
        "filter_unit": filter_unit,
        "all_company_units": all_company_units,
        "date_range_label": f"{date_from.strftime('%d %b %Y')} — {date_to.strftime('%d %b %Y')}",
        "trend_granularity": trend_granularity,
        # Word clouds
        "tag_cloud": json.dumps(tag_cloud),
        "rca_cloud": json.dumps(rca_cloud),
        # New analytics
        "frt_labels": json.dumps(frt_labels),
        "frt_values": json.dumps(frt_values),
        "avg_frt": avg_frt,
        "aging_labels": json.dumps(aging_labels),
        "aging_counts": json.dumps(aging_counts),
        "heatmap_matrix": json.dumps(heatmap_matrix),
        "resrate_labels": json.dumps(resrate_labels),
        "resrate_values": json.dumps(resrate_values),
        "reopened_count": reopened_count,
        "reopen_pct": reopen_pct,
        "stale_tickets": stale_tickets,
        # Regional map
        "map_units_json": map_units_json,
        "map_units": map_units_list,
        "map_total_units": map_total_units,
        "map_covered_tickets": map_covered_tickets,
    }
    return render(request, "admin/dashboard.html", context)


# =====================================================================
# Public — Client Portal
# =====================================================================

def client_dashboard(request):
    """
    Display the service catalogue as a responsive grid of CaseCategory cards,
    plus any DynamicForms configured to appear on the portal.

    Only root categories (parent=None) are displayed.  Categories that have
    children act as "Main Categories" — clicking them shows sub-categories.
    Categories without children link directly to the ticket form.
    """
    from core.models import DynamicForm
    categories = (
        CaseCategory.objects
        .filter(parent__isnull=True)
        .exclude(slug__in=["whatsapp-general", "email-general"])
        .prefetch_related("children")
    )
    portal_forms = DynamicForm.objects.filter(is_published=True, show_on_portal=True)
    return render(request, "client/dashboard.html", {"categories": categories, "portal_forms": portal_forms})


@login_required
@require_POST
def create_category(request):
    """
    AJAX endpoint to create a new CaseCategory.
    Only accessible by SuperAdmin users.
    """
    if request.user.role_access != User.RoleAccess.SUPERADMIN:
        return JsonResponse({"error": "Access denied."}, status=403)

    name = request.POST.get("name", "").strip()
    description = request.POST.get("description", "").strip()
    icon = request.POST.get("icon", "").strip()
    prefix_code = request.POST.get("prefix_code", "RQ").strip()[:2]
    parent_id = request.POST.get("parent", "").strip()
    is_confidential = request.POST.get("is_confidential") == "on"
    is_attachment_mandatory = request.POST.get("is_attachment_mandatory") == "on"
    template_subject = request.POST.get("template_subject", "").strip()
    template_text = request.POST.get("template_text", "").strip()

    if not name:
        return JsonResponse({"error": "Category name is required."}, status=400)

    parent = None
    if parent_id:
        parent = CaseCategory.objects.filter(id=parent_id).first()

    try:
        category = CaseCategory(
            name=name,
            description=description,
            icon=icon or "📝",
            prefix_code=prefix_code or "RQ",
            parent=parent,
            is_confidential=is_confidential,
            is_attachment_mandatory=is_attachment_mandatory,
            template_subject=template_subject,
            template_text=template_text,
            created_by=request.user,
            updated_by=request.user,
        )
        category.save()
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

    messages.success(request, f'Category "{name}" created successfully.')
    return JsonResponse({"success": True, "id": str(category.id), "name": category.name})


@login_required
@require_POST
def update_category(request, category_id):
    """
    AJAX endpoint to update a CaseCategory.
    Only accessible by SuperAdmin users.
    """
    if request.user.role_access != User.RoleAccess.SUPERADMIN:
        return JsonResponse({"error": "Access denied."}, status=403)

    category = get_object_or_404(CaseCategory, id=category_id)

    name = request.POST.get("name", "").strip()
    if not name:
        return JsonResponse({"error": "Category name is required."}, status=400)

    category.name = name
    category.description = request.POST.get("description", "").strip()
    category.icon = request.POST.get("icon", "").strip() or "📝"
    category.prefix_code = request.POST.get("prefix_code", "RQ").strip()[:2] or "RQ"
    category.is_confidential = request.POST.get("is_confidential") == "on"
    category.is_attachment_mandatory = request.POST.get("is_attachment_mandatory") == "on"
    category.template_subject = request.POST.get("template_subject", "").strip()
    category.template_text = request.POST.get("template_text", "").strip()
    category.updated_by = request.user

    # Update parent
    parent_id = request.POST.get("parent", "").strip()
    if parent_id:
        if str(category.id) == parent_id:
            return JsonResponse({"error": "A category cannot be its own parent."}, status=400)
        new_parent = CaseCategory.objects.filter(id=parent_id).first()
        category.parent = new_parent
    else:
        category.parent = None

    # Allow changing slug if name changed
    new_slug = request.POST.get("slug", "").strip()
    if new_slug and new_slug != category.slug:
        if CaseCategory.objects.filter(slug=new_slug).exclude(id=category.id).exists():
            return JsonResponse({"error": f'Slug "{new_slug}" is already in use.'}, status=400)
        category.slug = new_slug

    try:
        category.save()
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

    messages.success(request, f'Category "{name}" updated successfully.')
    return JsonResponse({"success": True})


@login_required
@require_POST
def delete_category(request, category_id):
    """
    AJAX endpoint to delete a CaseCategory.
    Only accessible by SuperAdmin users.
    Prevents deletion if category has associated tickets.
    """
    if request.user.role_access != User.RoleAccess.SUPERADMIN:
        return JsonResponse({"error": "Access denied."}, status=403)

    category = get_object_or_404(CaseCategory, id=category_id)

    # Check for associated tickets
    ticket_count = CaseRecord.objects.filter(category=category).count()
    if ticket_count > 0:
        return JsonResponse({
            "error": f'Cannot delete "{category.name}": {ticket_count} ticket(s) are using this category. '
                     f'Please reassign them first.'
        }, status=400)

    # Check for children
    children_count = category.children.count()
    if children_count > 0:
        return JsonResponse({
            "error": f'Cannot delete "{category.name}": it has {children_count} sub-category(ies). '
                     f'Please delete or move them first.'
        }, status=400)

    name = category.name
    category.delete()

    messages.success(request, f'Category "{name}" deleted successfully.')
    return JsonResponse({"success": True})


def category_children(request, slug):
    """
    Display sub-categories of a category at any level.
    If the category has no children (leaf), redirect to the ticket form.
    """
    parent = get_object_or_404(CaseCategory, slug=slug)
    children = parent.children.prefetch_related("children").all()
    if not children.exists():
        return redirect("cases:create_case_category", slug=slug)

    # Build breadcrumb trail: current → parent → grandparent → ...
    breadcrumbs = []
    node = parent.parent
    while node:
        breadcrumbs.insert(0, node)
        node = node.parent

    # All root categories for parent dropdown in edit modal
    all_root_categories = (
        CaseCategory.objects
        .filter(parent__isnull=True)
        .exclude(slug__in=["whatsapp-general", "email-general"])
    )

    return render(request, "client/category_children.html", {
        "parent": parent,
        "children": children,
        "breadcrumbs": breadcrumbs,
        "all_root_categories": all_root_categories,
    })


def create_case(request, slug=None):
    """
    Public case submission form.

    If ``slug`` is provided, the category is pre-selected.
    """
    import json
    from core.models import JobRole as JobRoleModel, SiteConfig
    initial = {}
    selected_category = None

    _site_cfg = SiteConfig.get_solo()
    _job_role_mode = _site_cfg.job_role_mode

    if slug:
        selected_category = get_object_or_404(CaseCategory, slug=slug)
        # If this is a parent category with children, redirect to sub-category selection
        if selected_category.children.exists():
            return redirect("cases:category_children", slug=slug)
        initial["category"] = selected_category

    # Only show leaf categories (no children) in the form dropdown.
    # Exclude any category that has at least one child.
    parent_ids = CaseCategory.objects.filter(
        parent__isnull=False
    ).values_list("parent_id", flat=True)
    categories_qs = (
        CaseCategory.objects
        .exclude(slug__in=["whatsapp-general", "email-general"])
        .exclude(id__in=parent_ids)
    )
    category_templates_json = json.dumps({
        str(c.id): c.template_text for c in categories_qs if c.template_text
    })
    category_templates_subject_json = json.dumps({
        str(c.id): c.template_subject for c in categories_qs if c.template_subject
    })
    category_cr_enabled_json = json.dumps([
        str(c.id) for c in categories_qs if c.enable_change_request
    ])

    # Build document templates per category for the dynamic CR modal
    _cat_ids_with_cr = [c.id for c in categories_qs if c.enable_change_request]
    _cr_templates = (
        DocumentTemplate.objects
        .filter(categories__id__in=_cat_ids_with_cr)
        .prefetch_related("fields", "categories")
        .distinct()
    )
    _cat_templates_map = {}
    for tpl in _cr_templates:
        tpl_data = {
            "id": str(tpl.id),
            "title": tpl.title,
            "fields": [
                {
                    "order": f.order,
                    "label": f.label,
                    "placeholder": f.placeholder,
                    "default_content": f.default_content,
                    "required": f.is_required,
                }
                for f in tpl.ordered_fields()
            ],
        }
        for cat in tpl.categories.all():
            _cat_templates_map.setdefault(str(cat.id), []).append(tpl_data)
    category_document_templates_json = json.dumps(_cat_templates_map)

    available_approvers = User.objects.filter(
        is_active=True,
    ).order_by("first_name", "last_name")
    available_approvers_json = json.dumps([
        {'id': str(u.id), 'name': u.get_full_name() or u.username, 'email': u.email or ''}
        for u in available_approvers
    ])

    if request.method == "POST":
        # Rate Limiting Check
        client_ip, _ = _get_client_ip(request)
        client_ip = client_ip or "unknown_ip"
        cache_key = f"create_case_rate_limit_{client_ip}"
        attempts = cache.get(cache_key, 0)
        
        if attempts >= 5:
            messages.error(request, "Too many requests. Please wait 10 minutes before submitting a new ticket.")
            return render(request, "client/create_case.html", {
                "categories": categories_qs,
                "company_units": CompanyUnit.objects.all(),
            }, status=429)

        form = CaseCreateForm(request.POST, request.FILES, job_role_mode=_job_role_mode)
        if form.is_valid():
            # Validate attachment sizes (10 MB limit)
            uploaded_files = request.FILES.getlist("attachments")
            file_errors = form.validate_attachments(uploaded_files)
            
            # Check if attachment is mandatory for this category
            category = form.cleaned_data.get("category")
            if category and category.is_attachment_mandatory and not uploaded_files:
                file_errors.append("Attachment is mandatory for this category. Please upload at least one file.")

            if file_errors:
                for err in file_errors:
                    form.add_error(None, err)
                return render(request, "client/create_case.html", {
                    "form": form,
                    "job_role_mode": _job_role_mode,
                    "selected_category": selected_category,
                    "categories": categories_qs,
                    "company_units": CompanyUnit.objects.all(),
                    "category_templates_json": category_templates_json,
                    "category_templates_subject_json": category_templates_subject_json,
                    "category_cr_enabled_json": category_cr_enabled_json,
                    "category_document_templates_json": category_document_templates_json,
                    "available_approvers_json": available_approvers_json,
                })

            email = form.cleaned_data["requester_email"]
            phone = form.cleaned_data["requester_phone"]
            company_unit = form.cleaned_data["company_unit"]

            cr_request_change_check = request.POST.get("cr_request_change", "").strip()
            cr_chronology_check = request.POST.get("cr_chronology", "").strip()
            # Also detect content in dynamic template fields
            cr_dynamic_fields_check = any(
                v.strip() for k, v in request.POST.items() if k.startswith("cr_field_")
            )
            cr_has_content_early = bool(cr_request_change_check) or bool(cr_chronology_check) or cr_dynamic_fields_check

            # Link ticket to logged-in portal user if they toggled tracking OR if CR content is present
            if request.user.is_authenticated and (
                request.POST.get("track_via_chat") == "true" or cr_has_content_early
            ):
                email = request.user.email
                phone = ""

            # Auto-link an existing Employee by email OR phone, else create one
            employee = Employee.find_by_contact(email=email, phone=phone)
            created = employee is None
            if created:
                employee = Employee.objects.create(
                    full_name=form.cleaned_data["requester_name"],
                    email=email or None,
                    phone_number=phone or None,
                    job_role=form.cleaned_data["job_role"],
                    unit=company_unit,
                )

            # Update existing employee details if they changed
            if not created:
                updated = False
                if form.cleaned_data["requester_name"] and employee.full_name != form.cleaned_data["requester_name"]:
                    employee.full_name = form.cleaned_data["requester_name"]
                    updated = True
                if company_unit and employee.unit != company_unit:
                    employee.unit = company_unit
                    updated = True
                if form.cleaned_data["job_role"] and employee.job_role != form.cleaned_data["job_role"]:
                    employee.job_role = form.cleaned_data["job_role"]
                    updated = True
                if updated:
                    employee.save(update_fields=["full_name", "unit", "job_role"])

            category = form.cleaned_data["category"]
            case = CaseRecord.objects.create(
                requester=employee,
                requester_email=email,
                requester_phone=phone,
                requester_name=form.cleaned_data["requester_name"],
                requester_job_role=form.cleaned_data["job_role"],
                requester_unit_name=company_unit.name,
                category=category,
                subject=form.cleaned_data["subject"],
                problem_description=form.cleaned_data["problem_description"],
                link=form.cleaned_data.get("link", ""),
                source=CaseRecord.Source.WEBFORM,
                status=CaseRecord.Status.OPEN,
                has_unread_messages=True,
            )

            # Create the initial message from the problem description
            msg = Message.objects.create(
                case=case,
                sender_employee=employee,
                body=form.cleaned_data["problem_description"],
                direction=Message.Direction.INBOUND,
                channel=Message.Channel.WEB,
            )

            # Handle multiple attachments
            for uploaded_file in request.FILES.getlist("attachments"):
                Attachment.objects.create(
                    message=msg,
                    file=uploaded_file,
                    original_filename=uploaded_file.name,
                    mime_type=getattr(uploaded_file, "content_type", ""),
                    file_size=uploaded_file.size,
                )

            # Inline change request — process BEFORE dispatching desk notification
            # so we can delay notification until CR is fully approved.
            cr_request_change = request.POST.get("cr_request_change", "").strip()
            cr_chronology = request.POST.get("cr_chronology", "").strip()

            def _quill_has_content(html):
                text = re.sub(r'<[^>]+>', '', html or '').strip()
                return bool(text) or '<img' in (html or '')

            cr_is_pending_approval = False

            # Dynamic template fields (new flow)
            cr_template_id = request.POST.get("cr_template_id", "").strip()
            cr_template = None
            cr_field_values = []
            cr_has_dynamic_content = False
            if cr_template_id:
                cr_template = DocumentTemplate.objects.filter(id=cr_template_id).prefetch_related("fields").first()
                if cr_template:
                    for field in cr_template.ordered_fields():
                        val = request.POST.get(f"cr_field_{field.order}", "").strip()
                        cr_field_values.append({"label": field.label, "value": val})
                    cr_has_dynamic_content = any(_quill_has_content(fv["value"]) for fv in cr_field_values)

            cr_use_dynamic = cr_has_dynamic_content and cr_template is not None
            cr_use_legacy = (not cr_use_dynamic) and _quill_has_content(cr_request_change) and _quill_has_content(cr_chronology)

            if category.enable_change_request and (cr_use_dynamic or cr_use_legacy):
                from django.utils import timezone as _tz
                approver_ids = request.POST.getlist("cr_approvers")
                has_approvers = bool(approver_ids)
                cr_doc = ChangeRequestDocument.objects.create(
                    case=case,
                    document_template=cr_template if cr_use_dynamic else None,
                    field_values=cr_field_values if cr_use_dynamic else [],
                    request_change=cr_request_change if cr_use_legacy else "",
                    chronology=cr_chronology if cr_use_legacy else "",
                    status=(
                        ChangeRequestDocument.Status.PENDING_APPROVAL
                        if has_approvers else ChangeRequestDocument.Status.APPROVED
                    ),
                    submitted_at=_tz.now(),
                )
                if has_approvers:
                    for order, uid in enumerate(approver_ids, start=1):
                        approver = User.objects.filter(id=uid).first()
                        if approver:
                            ChangeRequestApproval.objects.create(
                                document=cr_doc,
                                approver=approver,
                                order=order,
                            )
                    case.status = CaseRecord.Status.PENDING_APPROVAL
                    case.save(update_fields=["status"])
                    # Add the submitter as follower so CR workflow history is accessible in My Tickets
                    if request.user.is_authenticated:
                        case.followers.add(request.user)
                    _notify_change_request_approvers(cr_doc)
                    # Desk notification is deferred — _finalize_approved_change_request
                    # will dispatch it once all approvers have approved.
                    cr_is_pending_approval = True
                else:
                    # No approvers — auto-approved, still add submitter as follower
                    if request.user.is_authenticated:
                        case.followers.add(request.user)
                    _finalize_approved_change_request(cr_doc)

            # Notify desk agents only when the ticket is immediately active.
            # Tickets pending CR approval are hidden from the desk until approved.
            if not cr_is_pending_approval:
                from gateways.tasks import _dispatch_new_ticket_notifs
                _dispatch_new_ticket_notifs(str(case.id))

            # Create Audit Log entry
            CaseAuditLog.objects.create(
                case=case,
                created_by=User.objects.filter(is_superuser=True).first() if not request.user.is_authenticated else request.user,
                action=CaseAuditLog.ActionText.CREATED,
                new_value="Case created via Public Portal"
            )

            messages.success(request, f"Your ticket ({case.case_number}) has been created successfully.")

            # Increment Rate Limit Counter
            cache.set(cache_key, attempts + 1, timeout=600)  # 10 minutes timeout

            return redirect("cases:case_submitted", case_id=case.id)
    else:
        if request.user.is_authenticated and request.user.email:
            try:
                emp = Employee.objects.select_related("unit").get(email=request.user.email)
                initial.setdefault("requester_name", emp.full_name)
                initial.setdefault("requester_email", emp.email)
                if emp.unit:
                    initial.setdefault("company_unit", emp.unit)
                if emp.job_role:
                    initial.setdefault("job_role", emp.job_role)
            except Employee.DoesNotExist:
                initial.setdefault("requester_email", request.user.email)
        form = CaseCreateForm(initial=initial, job_role_mode=_job_role_mode)

    return render(request, "client/create_case.html", {
        "form": form,
        "job_role_mode": _job_role_mode,
        "selected_category": selected_category,
        "categories": categories_qs,
        "company_units": CompanyUnit.objects.all(),
        "category_templates_json": category_templates_json,
        "category_templates_subject_json": category_templates_subject_json,
        "category_cr_enabled_json": category_cr_enabled_json,
        "category_document_templates_json": category_document_templates_json,
        "available_approvers_json": available_approvers_json,
    })


def case_submitted(request, case_id):
    """Confirmation page after a case has been submitted."""
    case = get_object_or_404(CaseRecord, id=case_id)

    # Build approval links for pending CR tickets so the submitter can share them.
    pending_approvals = []
    if case.status == CaseRecord.Status.PENDING_APPROVAL:
        active_doc = (
            case.change_requests
            .filter(status=ChangeRequestDocument.Status.PENDING_APPROVAL)
            .order_by("-revision_number")
            .first()
        )
        if active_doc:
            for appr in active_doc.approvals.filter(
                status=ChangeRequestApproval.Status.PENDING
            ).order_by("order").select_related("approver"):
                url = request.build_absolute_uri(
                    reverse("cases:change_request_approve", args=[appr.token])
                )
                pending_approvals.append({
                    "order": appr.order,
                    "name": appr.approver.get_full_name() or appr.approver.username,
                    "email": appr.approver.email or "",
                    "url": url,
                })

    return render(request, "client/case_submitted.html", {
        "case": case,
        "pending_approvals": pending_approvals,
    })

def send_case_email(request, case_id):
    """View to trigger an email to the requester with case details."""
    from django.contrib import messages
    from gateways.tasks import send_case_acknowledgment_task
    
    case = get_object_or_404(CaseRecord, id=case_id)
    
    if request.method == "POST":
        try:
            send_case_acknowledgment_task.delay(str(case.id))
            messages.success(request, f"Ticket information has been successfully sent to {case.requester_email}.")
        except Exception as e:
            messages.error(request, f"Failed to send ticket information email. The mail service might be temporarily down. Please take a screenshot of this page for your records. (Error: {e})")
            
    return redirect("cases:case_submitted", case_id=case.id)


# =====================================================================
# Dynamic Form Public Renderer
# =====================================================================

def public_form_view(request, slug):
    """
    Renders a published DynamicForm in a beautiful client-facing view.
    Handles the submission logic and transforms raw POST data into the JSON answers format.
    """
    from core.models import DynamicForm, FormSubmission, SiteConfig
    from django.contrib import messages
    from django.http import Http404
    from django.core.files.storage import FileSystemStorage
    import os
    from django.conf import settings
    from django.apps import apps

    form_obj = get_object_or_404(DynamicForm, slug=slug)

    is_preview = request.GET.get('preview') == '1'

    # Check if form is published (skip check in preview mode for staff)
    if not form_obj.is_published and not request.user.is_staff:
        raise Http404("This form is not currently available.")

    # Check if form requires login
    if form_obj.requires_login and not request.user.is_authenticated:
        messages.warning(request, "You must be logged in to view and submit this form.")
        # Assuming you have a 'accounts:login' or using reverse
        return redirect(f"{settings.LOGIN_URL}?next={request.path}")

    fields = form_obj.fields.all()

    client_ip, _ = _get_client_ip(request)
    client_ip = client_ip or "unknown_ip"
    cache_key = f"public_form_rate_limit_{client_ip}_{slug}"

    if request.method == "POST":
        # Rate Limiting Check
        attempts = cache.get(cache_key, 0)
        if attempts >= 5:
            messages.error(request, "Too many requests. Please wait 10 minutes before submitting this form again.")
            return render(request, "client/public_form.html", {
                "dynamic_form": form_obj,
                "fields": fields,
                "is_preview": is_preview,
                "company_units": apps.get_model('core', 'CompanyUnit').objects.all()
            }, status=429)

        # Build the JSON response dict
        answers = {}
        errors = {}

        site_config = SiteConfig.get_solo()
        max_size_bytes = site_config.max_upload_size_mb * 1024 * 1024
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'form_uploads'), base_url=f"{settings.MEDIA_URL}form_uploads/")

        for field in fields:
            if 'attachment' in field.field_type:
                files = request.FILES.getlist(f"field_{field.id}")
                if field.is_required and not files:
                    errors[str(field.id)] = "This field is required."
                    answers[str(field.id)] = None
                    continue

                saved_files = []
                for f in files:
                    if f.size > max_size_bytes:
                        errors[str(field.id)] = f"File {f.name} exceeds {site_config.max_upload_size_mb}MB limit."
                        break
                    
                    filename = fs.save(os.path.basename(f.name), f)
                    saved_files.append(fs.url(filename))
                
                if field.field_type == 'attachment':
                    answers[str(field.id)] = saved_files[0] if saved_files else None
                else:
                    answers[str(field.id)] = saved_files

            elif field.field_type == 'checkbox':
                # Checkbox fields use getlist because they can have multiple values
                val = request.POST.getlist(f"field_{field.id}")
                if field.is_required and not val:
                    errors[str(field.id)] = "This field is required."
                answers[str(field.id)] = val
            else:
                val = request.POST.get(f"field_{field.id}", "").strip()
                if field.is_required and not val:
                    errors[str(field.id)] = "This field is required."
                answers[str(field.id)] = val

        if form_obj.collect_user:
            answers['sys_user_login'] = request.POST.get('sys_user_login', '')
        if form_obj.collect_company:
            company_unit_val = request.POST.get('sys_company_unit', '')
            if not company_unit_val:
                errors['sys_company_unit'] = "Company Unit is required."
            answers['sys_company_unit'] = company_unit_val

        if errors:
            messages.error(request, "Please fill out all required fields marked with *")
            return render(request, "client/public_form.html", {
                "dynamic_form": form_obj,
                "fields": fields,
                "answers": answers,
                "errors": errors,
                "company_units": apps.get_model('core', 'CompanyUnit').objects.all()
            })
        
        # Save submission
        FormSubmission.objects.create(
            form=form_obj,
            submitted_by=request.user if request.user.is_authenticated else None,
            answers=answers
        )

        # Increment Rate Limit Counter
        if not is_preview:
            cache.set(cache_key, attempts + 1, timeout=600)  # 10 minutes timeout

        messages.success(request, form_obj.success_message)
        # Using a simple success flag in context or simply rendering directly
        return render(request, "client/public_form.html", {
            "dynamic_form": form_obj,
            "success": True,
            "company_units": apps.get_model('core', 'CompanyUnit').objects.all()
        })

    return render(request, "client/public_form.html", {
        "dynamic_form": form_obj,
        "fields": fields,
        "is_preview": is_preview,
        "company_units": apps.get_model('core', 'CompanyUnit').objects.all()
    })


# =====================================================================
# Auth Required Desk (Admin/Staff) Views below
# =====================================================================

@staff_required
def case_list(request):
    """
    Filterable list of all cases for the support desk.
    Supports status/source/category filtering via GET params.
    """
    cases = CaseRecord.objects.select_related(
        "requester", "requester__unit", "category", "assigned_to"
    ).all()

    # --- Filtering ---
    folder = request.GET.get("folder", "inbox")
    status_filter = request.GET.get("status")
    source_filter = request.GET.get("source")
    category_filter = request.GET.get("category")
    assigned_to_filter = request.GET.get("assigned_to")
    type_filter = request.GET.get("type")
    tags_filter = request.GET.get("tags", "").strip()
    followers_filter = request.GET.get("followers")
    read_filter = request.GET.get("read_status", "")
    unit_filter = request.GET.get("company_unit")

    search_query = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    # Folder specific base filtering
    if folder == "spam":
        cases = cases.filter(is_spam=True)
    elif folder == "archive":
        cases = cases.filter(is_archived=True)
    else:  # inbox
        cases = cases.filter(is_spam=False, is_archived=False, master_ticket__isnull=True)
        # Tickets in the CR approval workflow are hidden from the desk inbox
        # until fully approved; desk agents can still find them via status filter.
        if not status_filter:
            cases = cases.exclude(status__in=[
                CaseRecord.Status.PENDING_APPROVAL,
                CaseRecord.Status.REVISION_REQUIRED,
            ])

    if status_filter:
        status_list = [s.strip() for s in status_filter.split(',') if s.strip()]
        cases = cases.filter(status__in=status_list)
    if source_filter:
        cases = cases.filter(source=source_filter)
    if category_filter:
        cases = cases.filter(category__slug=category_filter)
    if assigned_to_filter:
        if assigned_to_filter == "unassigned":
            cases = cases.filter(assigned_to__isnull=True)
        else:
            cases = cases.filter(assigned_to_id=assigned_to_filter)
    if type_filter:
        cases = cases.filter(case_type=type_filter)
    if unit_filter:
        cases = cases.filter(requester__unit_id=unit_filter)
    if tags_filter:
        cases = cases.filter(tags__icontains=tags_filter)
    if followers_filter:
        cases = cases.filter(followers__id=followers_filter)
    if read_filter == "unread":
        cases = cases.filter(has_unread_messages=True)
    elif read_filter == "unopened":
        cases = cases.filter(last_viewed_at__isnull=True)
    elif read_filter == "read":
        cases = cases.filter(has_unread_messages=False, last_viewed_at__isnull=False)

    if search_query:
        # Support full case number format (e.g. "RQ-E8973E6D"):
        # strip any PREFIX- prefix and search the UUID hex portion directly.
        dash_idx = search_query.find('-')
        uuid_part = search_query[dash_idx + 1:].lower() if dash_idx != -1 else None
        q = (
            Q(id__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(requester_name__icontains=search_query) |
            Q(requester__full_name__icontains=search_query) |
            Q(requester_email__icontains=search_query) |
            Q(requester__email__icontains=search_query) |
            Q(requester__phone_number__icontains=search_query) |
            Q(requester_unit_name__icontains=search_query) |
            Q(requester__unit__name__icontains=search_query)
        )
        if uuid_part:
            q |= Q(id__istartswith=uuid_part)
        cases = cases.filter(q)
    if date_from:
        cases = cases.filter(created_at__date__gte=date_from)
    if date_to:
        cases = cases.filter(created_at__date__lte=date_to)

    # --- Sorting ---
    sort_field = request.GET.get("sort", "created_at")
    sort_order = request.GET.get("order", "desc")
    ALLOWED_SORT_FIELDS = {
        "case_number": "id",
        "subject": "subject",
        "status": "status",
        "source": "source",
        "requester": "requester__full_name",
        "category": "category__name",
        "assigned_to": "assigned_to__username",
        "created_at": "created_at",
        "last_viewed_at": "last_viewed_at",
    }
    # Annotate confidential access for the current user
    cases = _annotate_confidential_access(cases, request.user)

    db_field = ALLOWED_SORT_FIELDS.get(sort_field, "created_at")
    if sort_order == "asc":
        cases = cases.order_by(F(db_field).asc(nulls_last=True))
    else:
        cases = cases.order_by(F(db_field).desc(nulls_last=True))

    from django.core.paginator import Paginator
    paginator = Paginator(cases, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate unread counts
    unread_inbox = CaseRecord.objects.filter(is_spam=False, is_archived=False, master_ticket__isnull=True, has_unread_messages=True).count()
    unread_archive = CaseRecord.objects.filter(is_archived=True, has_unread_messages=True).count()
    unread_spam = CaseRecord.objects.filter(is_spam=True, has_unread_messages=True).count()

    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    assignees = User.objects.filter(is_staff=True, is_active=True).exclude(
        role_access__in=[User.RoleAccess.AUDITOR, User.RoleAccess.PORTALUSER]
    ).order_by("first_name", "username")
    all_followers = User.objects.filter(is_active=True).order_by("first_name", "username")

    return render(request, "admin/case_list.html", {
        "cases": page_obj,
        "statuses": CaseRecord.Status.choices,
        "sources": CaseRecord.Source.choices,
        "categories": CaseCategory.objects.all(),
        "types": CaseRecord.Type.choices,
        "assignees": assignees,
        "all_followers": all_followers,
        "current_filters": {
            "folder": folder,
            "status": status_filter or "",
            "source": source_filter or "",
            "category": category_filter or "",
            "assigned_to": assigned_to_filter or "",
            "type": type_filter or "",
            "tags": tags_filter or "",
            "followers": followers_filter or "",
            "read_status": read_filter or "",
            "q": search_query,
            "date_from": date_from or "",
            "date_to": date_to or "",
        },
        "current_sort": sort_field,
        "current_order": sort_order,
        "unread_inbox": unread_inbox,
        "unread_archive": unread_archive,
        "unread_spam": unread_spam,
    })


# ---------------------------------------------------------------
# Auto-Refresh Partial Views
# ---------------------------------------------------------------

@staff_required
def case_list_partial(request):
    """
    Returns only the table rows HTML fragment for the auto-refresh mechanism
    on the case list page. Accepts the same GET params as case_list.
    """
    cases = CaseRecord.objects.select_related(
        "requester", "requester__unit", "category", "assigned_to"
    ).all()

    folder = request.GET.get("folder", "inbox")
    status_filter = request.GET.get("status")
    source_filter = request.GET.get("source")
    category_filter = request.GET.get("category")
    assigned_to_filter = request.GET.get("assigned_to")
    type_filter = request.GET.get("type")
    tags_filter = request.GET.get("tags", "").strip()
    followers_filter = request.GET.get("followers")
    read_filter = request.GET.get("read_status", "")
    unit_filter = request.GET.get("company_unit")
    search_query = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if folder == "spam":
        cases = cases.filter(is_spam=True)
    elif folder == "archive":
        cases = cases.filter(is_archived=True)
    else:
        cases = cases.filter(is_spam=False, is_archived=False, master_ticket__isnull=True)

    if status_filter:
        status_list = [s.strip() for s in status_filter.split(',') if s.strip()]
        cases = cases.filter(status__in=status_list)
    if source_filter:
        cases = cases.filter(source=source_filter)
    if category_filter:
        cases = cases.filter(category__slug=category_filter)
    if assigned_to_filter:
        if assigned_to_filter == "unassigned":
            cases = cases.filter(assigned_to__isnull=True)
        else:
            cases = cases.filter(assigned_to_id=assigned_to_filter)
    if type_filter:
        cases = cases.filter(case_type=type_filter)
    if unit_filter:
        cases = cases.filter(requester__unit_id=unit_filter)
    if tags_filter:
        cases = cases.filter(tags__icontains=tags_filter)
    if followers_filter:
        cases = cases.filter(followers__id=followers_filter)
    if read_filter == "unread":
        cases = cases.filter(has_unread_messages=True)
    elif read_filter == "unopened":
        cases = cases.filter(last_viewed_at__isnull=True)
    elif read_filter == "read":
        cases = cases.filter(has_unread_messages=False, last_viewed_at__isnull=False)
    if search_query:
        cases = cases.filter(
            Q(id__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(requester_name__icontains=search_query) |
            Q(requester__full_name__icontains=search_query) |
            Q(requester_email__icontains=search_query) |
            Q(requester__email__icontains=search_query) |
            Q(requester__phone_number__icontains=search_query) |
            Q(requester_unit_name__icontains=search_query) |
            Q(requester__unit__name__icontains=search_query)
        )
    if date_from:
        cases = cases.filter(created_at__date__gte=date_from)
    if date_to:
        cases = cases.filter(created_at__date__lte=date_to)

    sort_field = request.GET.get("sort", "created_at")
    sort_order = request.GET.get("order", "desc")
    ALLOWED_SORT_FIELDS = {
        "case_number": "id", "subject": "subject", "status": "status",
        "source": "source", "requester": "requester__full_name",
        "category": "category__name", "assigned_to": "assigned_to__username",
        "created_at": "created_at", "last_viewed_at": "last_viewed_at",
    }
    # Annotate confidential access for the current user
    cases = _annotate_confidential_access(cases, request.user)

    db_field = ALLOWED_SORT_FIELDS.get(sort_field, "created_at")
    if sort_order == "asc":
        cases = cases.order_by(F(db_field).asc(nulls_last=True))
    else:
        cases = cases.order_by(F(db_field).desc(nulls_last=True))

    from django.core.paginator import Paginator
    paginator = Paginator(cases, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    unread_inbox = CaseRecord.objects.filter(is_spam=False, is_archived=False, master_ticket__isnull=True, has_unread_messages=True).count()
    unread_archive = CaseRecord.objects.filter(is_archived=True, has_unread_messages=True).count()
    unread_spam = CaseRecord.objects.filter(is_spam=True, has_unread_messages=True).count()

    return render(request, "partials/case_table_rows.html", {
        "cases": page_obj,
        "unread_inbox": unread_inbox,
        "unread_archive": unread_archive,
        "unread_spam": unread_spam,
    })


@staff_required
def case_kanban_partial(request):
    """
    Returns only the kanban columns HTML fragment for the auto-refresh mechanism.
    """
    cases = CaseRecord.objects.select_related(
        "requester", "requester__unit", "category", "assigned_to"
    ).all()

    folder = request.GET.get("folder", "inbox")
    source_filter = request.GET.get("source")
    category_filter = request.GET.get("category")
    priority_filter = request.GET.get("priority")
    status_filter = request.GET.get("status")
    assigned_to_filter = request.GET.get("assigned_to")
    type_filter = request.GET.get("type")
    tags_filter = request.GET.get("tags", "").strip()
    followers_filter = request.GET.get("followers")
    read_filter = request.GET.get("read_status", "")
    unit_filter = request.GET.get("company_unit")
    search_query = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if folder == "spam":
        cases = cases.filter(is_spam=True)
    elif folder == "archive":
        cases = cases.filter(is_archived=True)
    else:
        cases = cases.filter(is_spam=False, is_archived=False, master_ticket__isnull=True)

    if source_filter:
        cases = cases.filter(source=source_filter)
    if category_filter:
        cases = cases.filter(category__slug=category_filter)
    if priority_filter:
        cases = cases.filter(priority=priority_filter)
    if assigned_to_filter:
        if assigned_to_filter == "unassigned":
            cases = cases.filter(assigned_to__isnull=True)
        else:
            cases = cases.filter(assigned_to_id=assigned_to_filter)
    if type_filter:
        cases = cases.filter(case_type=type_filter)
    if unit_filter:
        cases = cases.filter(requester__unit_id=unit_filter)
    if tags_filter:
        cases = cases.filter(tags__icontains=tags_filter)
    if followers_filter:
        cases = cases.filter(followers__id=followers_filter)
    if read_filter == "unread":
        cases = cases.filter(has_unread_messages=True)
    elif read_filter == "unopened":
        cases = cases.filter(last_viewed_at__isnull=True)
    elif read_filter == "read":
        cases = cases.filter(has_unread_messages=False, last_viewed_at__isnull=False)
    if search_query:
        cases = cases.filter(
            Q(id__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(requester_name__icontains=search_query) |
            Q(requester__full_name__icontains=search_query) |
            Q(requester_email__icontains=search_query) |
            Q(requester__email__icontains=search_query) |
            Q(requester__phone_number__icontains=search_query) |
            Q(requester_unit_name__icontains=search_query) |
            Q(requester__unit__name__icontains=search_query)
        )
    if date_from:
        cases = cases.filter(created_at__date__gte=date_from)
    if date_to:
        cases = cases.filter(created_at__date__lte=date_to)

    # Annotate confidential access for the current user
    cases = _annotate_confidential_access(cases, request.user)

    status_columns = [
        ("Open", "⏳", "Open", "#6366f1"),
        ("Investigating", "🔍", "Investigating", "#f59e0b"),
        ("PendingInfo", "⏸️", "Pending Info", "#0ea5e9"),
        ("Resolved", "✅", "Resolved", "#10b981"),
        ("Closed", "🔒", "Closed", "#64748b"),
    ]
    if status_filter:
        status_list = [s.strip() for s in status_filter.split(',') if s.strip()]
        status_columns = [col for col in status_columns if col[0] in status_list]

    kanban_data = []
    for status_val, icon, label, color in status_columns:
        column_cases = [c for c in cases if c.status == status_val]
        kanban_data.append({
            "status": status_val, "icon": icon, "label": label,
            "color": color, "cases": column_cases, "count": len(column_cases),
        })

    return render(request, "partials/case_kanban_columns.html", {
        "kanban_data": kanban_data,
    })


@staff_required
@require_POST
def case_unmerge(request, case_id):
    """
    Unmerges a sub-ticket from its master ticket.
    """
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    if case.master_ticket:
        master_num = case.master_ticket.case_number
        case.master_ticket = None
        case.save(update_fields=["master_ticket"])
        messages.success(request, f"Ticket {case.case_number} has been unmerged from {master_num}.")
    return redirect("desk:case_detail", case_id=case.id)

@staff_required
def case_bulk_action(request):
    """
    Handles bulk actions applied to multiple cases: Merge, Archive, or Spam.
    """
    if request.method == "POST":
        action = request.POST.get("action")
        case_ids = request.POST.getlist("case_ids")
        master_id = request.POST.get("master_id")

        if not action or not case_ids:
            return redirect("desk:case_list")

        cases_qs = CaseRecord.objects.filter(id__in=case_ids)

        if action == "spam":
            cases_qs.update(is_spam=True, is_archived=False)
        elif action == "archive":
            cases_qs.update(is_archived=True, is_spam=False)
        elif action == "unspam":
            cases_qs.update(is_spam=False)
        elif action == "unarchive":
            cases_qs.update(is_archived=False)
        elif action == "merge" and master_id:
            try:
                master = CaseRecord.objects.get(id=master_id)
                # Exclude the master itself from being set as its own sub-ticket
                # Also ensure only Open tickets can be merged into a master ticket
                sub_tickets = cases_qs.exclude(id=master_id).filter(status=CaseRecord.Status.OPEN)
                sub_tickets.update(master_ticket=master, status=master.status)
            except CaseRecord.DoesNotExist:
                pass

    return redirect(request.META.get("HTTP_REFERER", "desk:case_list"))


@staff_required
def case_detail(request, case_id):
    """
    Split-panel case detail view:
    - Left panel:  real-time chat/thread with the employee.
    - Right panel: RCA + solving steps form.
    """
    case = get_object_or_404(
        CaseRecord.objects.select_related("requester", "category", "assigned_to"),
        id=case_id,
    )
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")

    unread_msg_ids = []
    if case.has_unread_messages:
        unread_msg_ids = list(case.messages.filter(direction="IN", is_read=False).values_list("id", flat=True))
        # Collect WhatsApp external IDs for read receipts before marking as read
        wa_external_ids = list(
            case.messages.filter(
                id__in=unread_msg_ids,
                channel="WhatsApp",
                external_id__gt="",
            ).values_list("external_id", flat=True)
        )
        case.has_unread_messages = False
        case.save(update_fields=["has_unread_messages"])
        case.messages.filter(id__in=unread_msg_ids).update(is_read=True)
        # Send read receipts (blue checkmarks) to WhatsApp asynchronously
        if wa_external_ids:
            from gateways.tasks import mark_wa_messages_read_task
            try:
                mark_wa_messages_read_task.delay(str(case.id), wa_external_ids)
            except Exception as e:
                logger.warning("Failed to queue mark_wa_messages_read for case %s: %s", case.id, e)

    # Record when staff last viewed this ticket
    from django.utils import timezone
    case.last_viewed_at = timezone.now()
    case.save(update_fields=["last_viewed_at"])

    messages_qs = case.messages.select_related(
        "sender_employee", "sender_staff",
        "quoted_message__sender_employee", "quoted_message__sender_staff",
    ).prefetch_related("attachments", "reactions").all()
    rca_form = CaseRCAForm(instance=case)
    reply_form = StaffReplyForm()

    # Capture the referring URL (the dashboard/list with filters)
    back_url = request.GET.get("back") or request.META.get("HTTP_REFERER", "/desk/cases/")
    # If the referer is just the current page itself, fallback to the list
    if f"/desk/cases/{case_id}" in back_url:
        back_url = "/desk/cases/"

    # Sub-ticket / Master-ticket contextual tabs
    master = case.master_ticket if case.master_ticket else case
    related_cases = [master] + list(master.sub_tickets.exclude(id=master.id).all())

    # Prev / Next ticket navigation (ordered by created_at DESC — same as list)
    prev_case = (
        CaseRecord.objects.filter(created_at__gt=case.created_at)
        .order_by("created_at")
        .values_list("id", flat=True)
        .first()
    )
    next_case = (
        CaseRecord.objects.filter(created_at__lt=case.created_at)
        .order_by("-created_at")
        .values_list("id", flat=True)
        .first()
    )

    # (Confidential access is now category-based; no per-ticket user list needed)

    return render(request, "admin/case_detail.html", {
        "case": case,
        "chat_messages": messages_qs,
        "unread_msg_ids": unread_msg_ids,
        "rca_form": rca_form,
        "reply_form": reply_form,
        "back_url": back_url,
        "related_cases": related_cases,
        "master_case": master,
        "company_units": CompanyUnit.objects.all(),
        "all_employees": Employee.objects.select_related("unit").all(),
        "all_categories": CaseCategory.objects.select_related("parent").exclude(
            id__in=CaseCategory.objects.filter(children__isnull=False).values("id")
        ),
        "rca_templates": RCATemplate.objects.filter(
            Q(category=case.category) | Q(category__isnull=True)
        ).select_related("category"),
        "prev_case_id": prev_case,
        "next_case_id": next_case,
    })


@staff_required
def case_update_requester(request, case_id):
    """
    HTMX endpoint / Standard POST to update the Requester (Employee) information.
    Used mainly to fix typos for tickets incoming from WhatsApp / Email.
    """
    if request.method == "POST":
        case = get_object_or_404(CaseRecord, id=case_id)
        if not _check_confidential_access(case, request.user):
            return HttpResponseForbidden("You do not have access to this confidential ticket.")
        if _case_is_closed(case, request):
            return redirect("desk:case_detail", case_id=case_id)
        if case.requester:
            full_name = request.POST.get("full_name")
            email = request.POST.get("email")
            phone_number = request.POST.get("phone_number")
            job_role = request.POST.get("job_role")
            unit_id = request.POST.get("unit_id")

            # Track changes for audit log
            changes = []
            if full_name and full_name != case.requester.full_name:
                changes.append(("full_name", case.requester.full_name, full_name))
                case.requester.full_name = full_name
            if email and email != (case.requester.email or ""):
                changes.append(("email", case.requester.email or "", email))
                case.requester.email = email
            if phone_number and phone_number != (case.requester.phone_number or ""):
                changes.append(("phone_number", case.requester.phone_number or "", phone_number))
                case.requester.phone_number = phone_number
            if job_role and job_role != (case.requester.job_role or ""):
                changes.append(("job_role", case.requester.job_role or "", job_role))
                case.requester.job_role = job_role
            if unit_id:
                try:
                    new_unit = CompanyUnit.objects.get(id=unit_id)
                    old_unit_name = case.requester.unit.name if case.requester.unit else ""
                    if new_unit.name != old_unit_name:
                        changes.append(("unit", old_unit_name, new_unit.name))
                    case.requester.unit = new_unit
                    case.requester_unit_name = new_unit.name
                    case.save(update_fields=["requester_unit_name"])
                except CompanyUnit.DoesNotExist:
                    pass

            case.requester.save()

            # Write audit logs for each changed field
            for field_name, old_val, new_val in changes:
                CaseAuditLog.objects.create(
                    case=case,
                    action=CaseAuditLog.ActionText.UPDATED,
                    field_name=f"requester_{field_name}",
                    old_value=old_val,
                    new_value=new_val,
                    created_by=request.user,
                )

    return redirect("desk:case_detail", case_id=case_id)


@staff_required
def case_update_subject(request, case_id):
    """HTMX/POST endpoint — update the case subject (title)."""
    if request.method == "POST":
        case = get_object_or_404(CaseRecord, id=case_id)
        if not _check_confidential_access(case, request.user):
            return HttpResponseForbidden("You do not have access to this confidential ticket.")
        if _case_is_closed(case, request):
            return redirect("desk:case_detail", case_id=case_id)
        new_subject = request.POST.get("subject", "").strip()
        if new_subject and new_subject != case.subject:
            old_subject = case.subject
            case.subject = new_subject
            case.save(update_fields=["subject"])
            CaseAuditLog.objects.create(
                case=case,
                action=CaseAuditLog.ActionText.UPDATED,
                field_name="subject",
                old_value=old_subject,
                new_value=new_subject,
                created_by=request.user,
            )
    return redirect("desk:case_detail", case_id=case_id)


@staff_required
def case_update_category(request, case_id):
    """HTMX/POST endpoint — update the case category."""
    if request.method == "POST":
        case = get_object_or_404(CaseRecord, id=case_id)
        if not _check_confidential_access(case, request.user):
            return HttpResponseForbidden("You do not have access to this confidential ticket.")
        if _case_is_closed(case, request):
            return redirect("desk:case_detail", case_id=case_id)
        new_category_id = request.POST.get("category_id", "").strip()
        if new_category_id:
            new_category = get_object_or_404(CaseCategory, id=new_category_id)
            if new_category != case.category:
                old_category_name = case.category.name
                case.category = new_category
                case.save(update_fields=["category"])
                CaseAuditLog.objects.create(
                    case=case,
                    action=CaseAuditLog.ActionText.UPDATED,
                    field_name="category",
                    old_value=old_category_name,
                    new_value=new_category.name,
                    created_by=request.user,
                )
    return redirect("desk:case_detail", case_id=case_id)


@staff_required
def case_change_requester(request, case_id):
    """
    POST endpoint to change (replace) the Requester on a CaseRecord.
    Swaps the linked Employee and syncs denormalized fields.
    """
    if request.method == "POST":
        case = get_object_or_404(CaseRecord, id=case_id)
        if not _check_confidential_access(case, request.user):
            return HttpResponseForbidden("You do not have access to this confidential ticket.")
        if _case_is_closed(case, request):
            return redirect("desk:case_detail", case_id=case_id)
        new_employee_id = request.POST.get("new_employee_id")
        if new_employee_id:
            try:
                new_employee = Employee.objects.select_related("unit").get(id=new_employee_id)
                old_requester_name = case.requester.full_name if case.requester else "N/A"
                case.requester = new_employee
                case.requester_name = new_employee.full_name
                case.requester_email = new_employee.email or ""
                case.requester_job_role = new_employee.job_role or ""
                case.requester_unit_name = new_employee.unit.name if new_employee.unit else ""
                case.save(update_fields=[
                    "requester", "requester_name", "requester_email",
                    "requester_job_role", "requester_unit_name",
                ])

                CaseAuditLog.objects.create(
                    case=case,
                    action=CaseAuditLog.ActionText.UPDATED,
                    field_name="requester",
                    old_value=old_requester_name,
                    new_value=new_employee.full_name,
                    created_by=request.user,
                )
            except Employee.DoesNotExist:
                pass
    return redirect("desk:case_detail", case_id=case_id)


@staff_required
def case_forward_escalation(request, case_id):
    """
    POST endpoint to Forward or Escalate a case to an external Email or WhatsApp number.
    Triggers gateways.tasks.escalate_case_task.
    """
    if request.method == "POST":
        case = get_object_or_404(CaseRecord, id=case_id)
        if not _check_confidential_access(case, request.user):
            return HttpResponseForbidden("You do not have access to this confidential ticket.")
        if _case_is_closed(case, request):
            return redirect("desk:case_detail", case_id=case_id)
        forward_to = request.POST.get("forward_to", "").strip()
        channel = request.POST.get("channel", "EMAIL").upper()
        custom_message = request.POST.get("custom_message", "").strip()
        selected_message_ids = request.POST.get("selected_message_ids", "").strip()

        if channel == "WHATSAPP":
            from core.models import SiteConfig
            if not SiteConfig.get_solo().wa_outbound_enabled:
                from django.contrib import messages
                messages.error(request, "Outbound WhatsApp messages are currently disabled.")
                return redirect("desk:case_detail", case_id=case_id)

        if forward_to and channel in ["EMAIL", "WHATSAPP"]:
            from cases.models import Message
            # Use body prefix "*** TICKET ESCALATED VIA" to identify it as escalation later
            msg_channel = Message.Channel.EMAIL if channel == 'EMAIL' else Message.Channel.WHATSAPP

            # Append Initials for Escalate
            staff_initials = getattr(request.user, "initials", "sys")
            full_body = f"*** TICKET ESCALATED VIA {channel} TO: {forward_to} ***\n\nInternal Agent Note:\n{custom_message}\n\n-{staff_initials}"

            msg = Message.objects.create(
                case=case,
                sender_staff=request.user,
                body=full_body,
                direction=Message.Direction.OUTBOUND,
                channel=msg_channel,
                delivery_status=Message.DeliveryStatus.PENDING,
            )

            # Fire celery task to send via API
            from gateways.tasks import escalate_case_task
            import random

            try:
                # Use apply_async with countdown to simulate human delay (2-6 seconds)
                delay_secs = random.randint(2, 6)
                escalate_case_task.apply_async(
                    args=[str(case.id), forward_to, channel, custom_message, str(msg.id)],
                    kwargs={"selected_message_ids": selected_message_ids},
                    countdown=delay_secs
                )
                from django.contrib import messages
                messages.success(request, f"Ticket successfully escalated to {forward_to} via {channel}.")
            except Exception as e:
                msg.delivery_status = Message.DeliveryStatus.FAILED
                msg.delivery_error = f"System Error: {str(e)}"
                msg.save(update_fields=["delivery_status", "delivery_error"])
                from django.contrib import messages
                messages.error(request, f"Failed to dispatch escalation task: {str(e)}")

    return redirect("desk:case_detail", case_id=case_id)


@staff_required
def case_update_rca(request, case_id):
    """
    HTMX endpoint — update root cause analysis, solving steps, status,
    and SLA fields.
    """
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")

    if case.is_locked:
        return HttpResponseForbidden(
            f"Ticket {case.case_number} is locked — approval workflow is in progress. "
            "No field changes are allowed until approval completes."
        )

    if request.method == "POST":
        original_status = case.status
        form = CaseRCAForm(request.POST, instance=case)
        if form.is_valid():
            desired_status = form.cleaned_data.get("status")
            
            # SLA validation if attempting to Close for the first time
            if desired_status == CaseRecord.Status.CLOSED and original_status != CaseRecord.Status.CLOSED:
                if not form.cleaned_data.get("root_cause_analysis") or not form.cleaned_data.get("solving_steps"):
                    # Revert instance mutation so the template doesn't think it's closed yet
                    case.status = original_status
                    return render(request, "partials/rca_form.html", {
                        "case": case,
                        "rca_form": form,
                        "error_message": "SLA Details (Root Cause Analysis & Solving Steps) are required to close the ticket.",
                    })
                    
                # Save changes but revert status until confirmed in the modal
                case = form.save(commit=False)
                case.status = original_status
                case.updated_by = request.user
                
                # Check what changed before saving to create audit logs
                if form.changed_data:
                    for field in form.changed_data:
                        if field == 'status': 
                            continue # we reverted status above
                        old_val = form.initial.get(field)
                        new_val = form.cleaned_data.get(field)
                        
                        # Followers is M2M, handled after save typically, but we log the attempt
                        if field == 'followers':
                            old_val = ", ".join([u.username for u in case.followers.all()])
                            new_val = ", ".join([u.username for u in new_val]) if new_val else ""

                        CaseAuditLog.objects.create(
                            case=case,
                            action=CaseAuditLog.ActionText.UPDATED,
                            field_name=field,
                            old_value=str(old_val) if old_val else "",
                            new_value=str(new_val) if new_val else "",
                            created_by=request.user,
                        )

                case.save()
                if case.sub_tickets.exists():
                    case.sub_tickets.update(status=case.status, updated_by=request.user)
                form.save_m2m() # Important for followers
                # Check if we should notify about assignment change during RCA save

                
                if 'assigned_to' in form.changed_data and case.assigned_to:
                    from gateways.tasks import send_assignment_email_task
                    case_url = request.build_absolute_uri(reverse("desk:case_detail", kwargs={"case_id": case.id}))
                    send_assignment_email_task.delay(str(case.id), str(request.user.id), case_url)
                
                from django.template.loader import render_to_string
                form_response = render(request, "partials/rca_form.html", {
                    "case": case,
                    "rca_form": CaseRCAForm(instance=case),
                    "success_message": "Details saved. Please confirm ticket closure below.",
                })
                
                modal_html = render_to_string("partials/closure_modal.html", {"case": case}, request=request)
                oob_modal = f'<div id="htmx-modal" hx-swap-oob="innerHTML">{modal_html}</div>'
                
                return HttpResponse(form_response.content.decode() + oob_modal)

            # Normal save for non-closing updates OR amending an already closed ticket
            case = form.save(commit=False)
            case.updated_by = request.user
            
            # --- Amendment Logic ---
            if original_status == CaseRecord.Status.CLOSED and case.edit_permission_status == CaseRecord.EditPermissionStatus.APPROVED:
                if form.changed_data:
                    case.amendment_count += 1
                case.edit_permission_status = CaseRecord.EditPermissionStatus.NONE
            # -----------------------
            
            # Check what changed before saving to create audit logs
            if form.changed_data:
                for field in form.changed_data:
                    old_val = form.initial.get(field)
                    new_val = form.cleaned_data.get(field)
                    
                    action_type = CaseAuditLog.ActionText.UPDATED
                    if field == 'status':
                        action_type = CaseAuditLog.ActionText.STATUS_CHANGE
                    elif field in ['assigned_to', 'followers']:
                        action_type = CaseAuditLog.ActionText.ASSIGNED
                    
                    if field == 'followers':
                        old_val = ", ".join([u.username for u in case.followers.all()])
                        new_val = ", ".join([u.username for u in new_val]) if new_val else ""
                    elif field == 'assigned_to':
                        old_val = str(old_val) if old_val else "Unassigned"
                        new_val = str(new_val) if new_val else "Unassigned"

                    CaseAuditLog.objects.create(
                        case=case,
                        action=action_type,
                        field_name=field,
                        old_value=str(old_val) if old_val is not None else "",
                        new_value=str(new_val) if new_val is not None else "",
                        created_by=request.user,
                    )

            if case.status == CaseRecord.Status.INVESTIGATING:
                case.hold_wa_session = True
            elif case.status in [CaseRecord.Status.RESOLVED, CaseRecord.Status.CLOSED]:
                case.hold_wa_session = False

            case.save()
            if case.sub_tickets.exists():
                case.sub_tickets.update(status=case.status, updated_by=request.user)
            form.save_m2m() # Important for followers

            if 'assigned_to' in form.changed_data and case.assigned_to:
                from gateways.tasks import send_assignment_email_task
                case_url = request.build_absolute_uri(reverse("desk:case_detail", kwargs={"case_id": case.id}))
                send_assignment_email_task.delay(str(case.id), str(request.user.id), case_url)

            redirect_url = request.POST.get("back_url", "/desk/cases/")
            if not url_has_allowed_host_and_scheme(
                redirect_url,
                allowed_hosts={request.get_host()},
                require_https=request.is_secure(),
            ) or f"/desk/cases/{case_id}" in redirect_url:
                redirect_url = "/desk/cases/"

            response = HttpResponse()
            response["HX-Redirect"] = redirect_url
            return response
        else:
            # Revert instance mutation if validation fails so the template logic maintains the true DB status
            case.status = original_status
            return render(request, "partials/rca_form.html", {
                "case": case,
                "rca_form": form,
                "error_message": "Please fix the errors below.",
                "back_url": request.POST.get("back_url", "/desk/cases/"),
            })

    return HttpResponse(status=405)


@staff_required
def case_send_reply(request, case_id):
    """
    HTMX endpoint — staff sends a reply message within the case thread.
    """
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    if _case_is_closed(case, request):
        return redirect("desk:case_detail", case_id=case_id)

    if request.method == "POST":
        form = StaffReplyForm(request.POST, request.FILES)
        if form.is_valid():
            # Must have at least body or attachment
            if not form.cleaned_data.get("body") and not form.cleaned_data.get("attachment"):
                messages = case.messages.select_related(
                    "sender_employee", "sender_staff"
                ).prefetch_related("attachments").all()
                return render(request, "partials/chat_thread.html", {
                    "case": case,
                    "chat_messages": messages,
                })

            # Determine the channel based on the original case source
            reply_channel = Message.Channel.WEB
            
            send_as_email = request.POST.get("send_as_email") == "true"
            if send_as_email and case.source in (
                CaseRecord.Source.WEBFORM,
                CaseRecord.Source.WEBCHAT,
            ):
                reply_channel = Message.Channel.EMAIL
            elif case.source == CaseRecord.Source.EMAIL:
                reply_channel = Message.Channel.EMAIL
            elif case.source == CaseRecord.Source.EVOLUTION_WA:
                reply_channel = Message.Channel.WHATSAPP  # Prep for future WA outbound support

            # Handle quote reply
            quoted_message = None
            quoted_id = request.POST.get("quoted_message_id", "").strip()
            if quoted_id:
                quoted_message = Message.objects.filter(id=quoted_id, case=case).first()

            msg = Message.objects.create(
                case=case,
                sender_staff=request.user,
                body=form.cleaned_data.get("body") or "",
                cc_emails=form.cleaned_data.get("cc_emails", ""),
                direction=Message.Direction.OUTBOUND,
                channel=reply_channel,
                delivery_status=Message.DeliveryStatus.PENDING if reply_channel != Message.Channel.WEB else Message.DeliveryStatus.SUCCESS,
                quoted_message=quoted_message,
            )

            # Handle optional attachment
            uploaded_file = form.cleaned_data.get("attachment")
            if uploaded_file:
                import uuid
                import os
                import mimetypes
                
                original_name = uploaded_file.name
                ext = os.path.splitext(original_name)[1].lower()
                dangerous_exts = {".php", ".php3", ".php4", ".php5", ".phtml", ".html", ".htm", ".exe", ".sh", ".bash", ".pl", ".py"}
                
                mime = getattr(uploaded_file, "content_type", "")
                if ext in dangerous_exts or not ext:
                    ext = mimetypes.guess_extension(mime) or ".bin"
                    
                safe_name = f"{uuid.uuid4().hex}{ext}"
                uploaded_file.name = safe_name

                Attachment.objects.create(
                    message=msg,
                    file=uploaded_file,
                    original_filename=original_name,
                    mime_type=mime,
                    file_size=uploaded_file.size,
                )

            # Update case timestamp
            case.updated_by = request.user
            case.save(update_fields=["updated_at", "updated_by"])

            # Trigger Outbound Async Tasks based on Channel
            try:
                if reply_channel == Message.Channel.EMAIL:
                    from gateways.tasks import send_outbound_email_task
                    send_outbound_email_task.delay(str(msg.id))
                elif reply_channel == Message.Channel.WHATSAPP:
                    from gateways.tasks import send_outbound_whatsapp_task
                    send_outbound_whatsapp_task.delay(str(msg.id))

                    # Reset the 60-minute session countdown for the employee
                    from gateways.tasks import check_wa_session_warning_task, check_wa_session_timeout_task
                    check_wa_session_warning_task.apply_async((str(case.id),), countdown=2700)   # 45 min: confirmation
                    check_wa_session_timeout_task.apply_async((str(case.id),), countdown=3600)   # 60 min: expiry
                elif reply_channel == Message.Channel.WEB:
                    # Push staff reply to portal user's WebSocket in real time
                    try:
                        from asgiref.sync import async_to_sync
                        from channels.layers import get_channel_layer
                        sender_name = request.user.get_full_name() or str(request.user)
                        async_to_sync(get_channel_layer().group_send)(
                            f"chat_{case.id}",
                            {
                                "type": "chat.message",
                                "message_id": str(msg.id),
                                "body": msg.body,
                                "direction": msg.direction,
                                "sender_name": sender_name,
                                "sent_at": msg.sent_at.isoformat(),
                                "is_system": False,
                            },
                        )
                    except Exception:
                        pass  # channel layer unavailable — portal user falls back to polling
            except Exception as e:
                msg.delivery_status = Message.DeliveryStatus.FAILED
                msg.delivery_error = f"Celery connection failed: {str(e)}"
                msg.save(update_fields=["delivery_status", "delivery_error"])

            # Return the refreshed chat thread partial
            messages = case.messages.select_related(
                "sender_employee", "sender_staff"
            ).prefetch_related("attachments").all()
            return render(request, "partials/chat_thread.html", {
                "case": case,
                "chat_messages": messages,
            })

    return HttpResponse(status=405)


@staff_required
def chat_thread(request, case_id):
    """
    HTMX endpoint — returns the chat thread partial for polling refresh.
    """
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")

    unread_msg_ids = []
    if case.has_unread_messages:
        unread_msg_ids = list(case.messages.filter(direction="IN", is_read=False).values_list("id", flat=True))
        wa_external_ids = list(
            case.messages.filter(
                id__in=unread_msg_ids, channel="WhatsApp", external_id__gt=""
            ).values_list("external_id", flat=True)
        )
        case.has_unread_messages = False
        case.save(update_fields=["has_unread_messages"])
        case.messages.filter(id__in=unread_msg_ids).update(is_read=True)
        if wa_external_ids:
            from gateways.tasks import mark_wa_messages_read_task
            try:
                mark_wa_messages_read_task.delay(str(case.id), wa_external_ids)
            except Exception as e:
                logger.warning("Failed to queue mark_wa_messages_read for case %s: %s", case.id, e)

    messages = case.messages.select_related(
        "sender_employee", "sender_staff",
        "quoted_message__sender_employee", "quoted_message__sender_staff",
    ).prefetch_related("attachments", "reactions").all()

    return render(request, "partials/chat_thread.html", {
        "case": case,
        "chat_messages": messages,
        "unread_msg_ids": unread_msg_ids,
    })


# =====================================================================
# Message Actions (Delete, Edit, React)
# =====================================================================

@staff_required
@require_POST
def message_delete(request, case_id, message_id):
    """Soft-delete an outbound WhatsApp message and revoke it on WhatsApp."""
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    if _case_is_closed(case, request):
        return redirect("desk:case_detail", case_id=case_id)
    msg = get_object_or_404(Message, id=message_id, case=case, direction="OUT")

    if not msg.is_deleted:
        msg.is_deleted = True
        msg.original_body = msg.body
        msg.body = ""
        msg.save(update_fields=["is_deleted", "original_body", "body"])

        if msg.channel == "WhatsApp" and msg.external_id:
            from gateways.tasks import delete_whatsapp_message_task
            try:
                delete_whatsapp_message_task.delay(str(msg.id))
            except Exception as e:
                logger.warning("Failed to queue delete_whatsapp_message for msg %s: %s", msg.id, e)

    messages = case.messages.select_related(
        "sender_employee", "sender_staff"
    ).prefetch_related("attachments", "reactions").all()
    return render(request, "partials/chat_thread.html", {
        "case": case,
        "chat_messages": messages,
        "unread_msg_ids": [],
    })


@staff_required
@require_POST
def message_edit(request, case_id, message_id):
    """Edit an outbound WhatsApp message text."""
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    if _case_is_closed(case, request):
        return redirect("desk:case_detail", case_id=case_id)
    msg = get_object_or_404(Message, id=message_id, case=case, direction="OUT")

    new_body = request.POST.get("body", "").strip()
    if not new_body or msg.is_deleted:
        messages = case.messages.select_related(
            "sender_employee", "sender_staff"
        ).prefetch_related("attachments", "reactions").all()
        return render(request, "partials/chat_thread.html", {
            "case": case, "chat_messages": messages, "unread_msg_ids": [],
        })

    if not msg.original_body:
        msg.original_body = msg.body
    msg.body = new_body
    msg.is_edited = True
    msg.save(update_fields=["body", "is_edited", "original_body"])

    if msg.channel == "WhatsApp" and msg.external_id:
        from gateways.tasks import edit_whatsapp_message_task
        try:
            edit_whatsapp_message_task.delay(str(msg.id))
        except Exception as e:
            logger.warning("Failed to queue edit_whatsapp_message for msg %s: %s", msg.id, e)

    messages = case.messages.select_related(
        "sender_employee", "sender_staff"
    ).prefetch_related("attachments", "reactions").all()
    return render(request, "partials/chat_thread.html", {
        "case": case, "chat_messages": messages, "unread_msg_ids": [],
    })


@staff_required
@require_POST
def message_react(request, case_id, message_id):
    """Send an emoji reaction to a WhatsApp message."""
    from cases.models import MessageReaction

    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    msg = get_object_or_404(Message, id=message_id, case=case)
    emoji = request.POST.get("emoji", "").strip()

    if emoji:
        # Toggle: if same emoji already exists, remove it (unreact)
        existing = MessageReaction.objects.filter(
            message=msg, reacted_by=request.user, emoji=emoji
        ).first()
        if existing:
            existing.delete()
            # Send empty reaction to WhatsApp to remove it
            if msg.channel == "WhatsApp" and msg.external_id:
                from gateways.tasks import react_whatsapp_message_task
                try:
                    react_whatsapp_message_task.delay(str(msg.id), "")
                except Exception as e:
                    logger.warning("Failed to queue react_whatsapp_message (remove) for msg %s: %s", msg.id, e)
        else:
            MessageReaction.objects.update_or_create(
                message=msg,
                reacted_by=request.user,
                defaults={"emoji": emoji},
            )
            if msg.channel == "WhatsApp" and msg.external_id:
                from gateways.tasks import react_whatsapp_message_task
                try:
                    react_whatsapp_message_task.delay(str(msg.id), emoji)
                except Exception as e:
                    logger.warning("Failed to queue react_whatsapp_message (add %s) for msg %s: %s", emoji, msg.id, e)

    messages = case.messages.select_related(
        "sender_employee", "sender_staff"
    ).prefetch_related("attachments", "reactions").all()
    return render(request, "partials/chat_thread.html", {
        "case": case, "chat_messages": messages, "unread_msg_ids": [],
    })


# =====================================================================
# Kanban Board
# =====================================================================

@staff_required
def case_quick_view(request, case_id):
    """
    HTMX endpoint — returns a quick view modal for a specific case card.
    """
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    return render(request, "partials/quick_view_modal.html", {
        "case": case,
    })

@staff_required
def case_kanban(request):
    """
    Kanban board — cases grouped by status columns.
    Supports same filters as case_list.
    """
    cases = CaseRecord.objects.select_related(
        "requester", "requester__unit", "category", "assigned_to"
    ).all()

    # --- Filtering ---
    folder = request.GET.get("folder", "inbox")
    source_filter = request.GET.get("source")
    category_filter = request.GET.get("category")
    priority_filter = request.GET.get("priority")
    status_filter = request.GET.get("status")
    assigned_to_filter = request.GET.get("assigned_to")
    type_filter = request.GET.get("type")
    tags_filter = request.GET.get("tags", "").strip()
    followers_filter = request.GET.get("followers")
    read_filter = request.GET.get("read_status", "")
    unit_filter = request.GET.get("company_unit")

    search_query = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    # Folder specific base filtering
    if folder == "spam":
        cases = cases.filter(is_spam=True)
    elif folder == "archive":
        cases = cases.filter(is_archived=True)
    else:  # inbox
        cases = cases.filter(is_spam=False, is_archived=False, master_ticket__isnull=True)

    if source_filter:
        cases = cases.filter(source=source_filter)
    if category_filter:
        cases = cases.filter(category__slug=category_filter)
    if priority_filter:
        cases = cases.filter(priority=priority_filter)
    if assigned_to_filter:
        if assigned_to_filter == "unassigned":
            cases = cases.filter(assigned_to__isnull=True)
        else:
            cases = cases.filter(assigned_to_id=assigned_to_filter)
    if type_filter:
        cases = cases.filter(case_type=type_filter)
    if unit_filter:
        cases = cases.filter(requester__unit_id=unit_filter)
    if tags_filter:
        cases = cases.filter(tags__icontains=tags_filter)
    if followers_filter:
        cases = cases.filter(followers__id=followers_filter)
    if read_filter == "unread":
        cases = cases.filter(has_unread_messages=True)
    elif read_filter == "unopened":
        cases = cases.filter(last_viewed_at__isnull=True)
    elif read_filter == "read":
        cases = cases.filter(has_unread_messages=False, last_viewed_at__isnull=False)

    if search_query:
        cases = cases.filter(
            Q(id__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(requester_name__icontains=search_query) |
            Q(requester__full_name__icontains=search_query) |
            Q(requester_email__icontains=search_query) |
            Q(requester__email__icontains=search_query) |
            Q(requester__phone_number__icontains=search_query) |
            Q(requester_unit_name__icontains=search_query) |
            Q(requester__unit__name__icontains=search_query)
        )
    if date_from:
        cases = cases.filter(created_at__date__gte=date_from)
    if date_to:
        cases = cases.filter(created_at__date__lte=date_to)

    # Annotate confidential access for the current user
    cases = _annotate_confidential_access(cases, request.user)

    # Group by status
    status_columns = [
        ("Open", "⏳", "Open", "#6366f1"),
        ("Investigating", "🔍", "Investigating", "#f59e0b"),
        ("PendingInfo", "⏸️", "Pending Info", "#0ea5e9"),
        ("Resolved", "✅", "Resolved", "#10b981"),
        ("Closed", "🔒", "Closed", "#64748b"),
    ]

    # If a status filter is active, only show matching columns
    if status_filter:
        status_list = [s.strip() for s in status_filter.split(',') if s.strip()]
        status_columns = [col for col in status_columns if col[0] in status_list]

    kanban_data = []
    for status_val, icon, label, color in status_columns:
        column_cases = [c for c in cases if c.status == status_val]
        kanban_data.append({
            "status": status_val,
            "icon": icon,
            "label": label,
            "color": color,
            "cases": column_cases,
            "count": len(column_cases),
        })

    # Calculate unread counts
    unread_inbox = CaseRecord.objects.filter(is_spam=False, is_archived=False, master_ticket__isnull=True, has_unread_messages=True).count()
    unread_archive = CaseRecord.objects.filter(is_archived=True, has_unread_messages=True).count()
    unread_spam = CaseRecord.objects.filter(is_spam=True, has_unread_messages=True).count()

    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    assignees = User.objects.filter(is_staff=True, is_active=True).order_by("first_name", "username")
    all_followers = User.objects.filter(is_active=True).order_by("first_name", "username")

    return render(request, "admin/case_kanban.html", {
        "kanban_data": kanban_data,
        "sources": CaseRecord.Source.choices,
        "categories": CaseCategory.objects.all(),
        "priorities": CaseRecord.Priority.choices,
        "types": CaseRecord.Type.choices,
        "assignees": assignees,
        "all_followers": all_followers,
        "current_filters": {
            "folder": folder,
            "status": status_filter or "",
            "source": source_filter or "",
            "category": category_filter or "",
            "priority": priority_filter or "",
            "assigned_to": assigned_to_filter or "",
            "type": type_filter or "",
            "tags": tags_filter or "",
            "followers": followers_filter or "",
            "read_status": read_filter or "",
            "q": search_query,
            "date_from": date_from or "",
            "date_to": date_to or "",
        },
        "unread_inbox": unread_inbox,
        "unread_archive": unread_archive,
        "unread_spam": unread_spam,
    })


@staff_required
def case_update_status(request, case_id):
    """
    HTMX endpoint — update case status (drag-and-drop from Kanban).
    """
    if request.method != "POST":
        return HttpResponse(status=405)

    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    new_status = request.POST.get("status", "")

    valid_statuses = [s[0] for s in CaseRecord.Status.choices]
    if new_status not in valid_statuses:
        return HttpResponse("Invalid status", status=400)

    # Prevent direct closing without SLA details from Kanban
    if new_status == CaseRecord.Status.CLOSED:
        if not case.root_cause_analysis or not case.solving_steps:
            return HttpResponse("SLA details (Root Cause Analysis & Solving Steps) are required before closing. Please click the case to open the detail panel and fill them out.", status=400)
            
        # Generate and return the modal directly
        return render(request, "partials/closure_modal.html", {"case": case})

    case.status = new_status
    if new_status == CaseRecord.Status.INVESTIGATING:
        case.hold_wa_session = True
    elif new_status in [CaseRecord.Status.RESOLVED, CaseRecord.Status.CLOSED]:
        case.hold_wa_session = False
        
    case.updated_by = request.user
    case.save(update_fields=["status", "hold_wa_session", "updated_at", "updated_by"])
    
    if case.sub_tickets.exists():
        case.sub_tickets.update(status=new_status, updated_at=case.updated_at, updated_by=request.user)

    return HttpResponse(status=204)


@staff_required
def case_close_and_notify(request, case_id):
    """
    HTMX endpoint — processes the final SLA closure message from the preview modal,
    closes the ticket, and sends the OUTBOUND message via WA/Email.
    """
    if request.method != "POST":
        return HttpResponse(status=405)
        
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    if case.is_locked:
        return HttpResponseForbidden(
            f"Ticket {case.case_number} is locked — approval workflow is in progress."
        )
    send_notification = request.POST.get("send_notification") == "1"
    closure_msg_body = request.POST.get("closure_message_body", "").strip()

    if send_notification and not closure_msg_body:
        return HttpResponse("Message body cannot be empty.", status=400)

    # Mark the case as closed
    case.status = CaseRecord.Status.CLOSED
    case.hold_wa_session = False
    case.updated_by = request.user
    case.save(update_fields=["status", "hold_wa_session", "updated_at", "updated_by"])

    from core.models import AuditLog
    AuditLog.log(AuditLog.Action.TICKET_CLOSE, request=request, target=case,
                 details={"case_number": str(case.id)})

    if case.sub_tickets.exists():
        case.sub_tickets.update(status=CaseRecord.Status.CLOSED, updated_at=case.updated_at, updated_by=request.user)

    if send_notification:
        # Determine the reply channel based on the source
        reply_channel = Message.Channel.WEB
        if case.source in (
            CaseRecord.Source.WEBFORM,
            CaseRecord.Source.WEBCHAT,
            CaseRecord.Source.EMAIL,
        ):
            reply_channel = Message.Channel.EMAIL
        elif case.source == CaseRecord.Source.EVOLUTION_WA:
            reply_channel = Message.Channel.WHATSAPP

        # Create the Outbound Message containing the closure payload
        msg = Message.objects.create(
            case=case,
            sender_staff=request.user,
            body=closure_msg_body,
            direction=Message.Direction.OUTBOUND,
            channel=reply_channel,
        )

        # Trigger appropriate sending mechanisms
        if reply_channel == Message.Channel.EMAIL:
            from gateways.tasks import send_outbound_email_task
            send_outbound_email_task.delay(str(msg.id))
        elif reply_channel == Message.Channel.WHATSAPP:
            from gateways.tasks import send_outbound_whatsapp_task
            msg.delivery_status = Message.DeliveryStatus.PENDING
            msg.save(update_fields=["delivery_status"])
            send_outbound_whatsapp_task.delay(str(msg.id))
                
    # Return a success script to trigger a reload or update
    return HttpResponse(
        '<script>window.location.reload();</script>',
        content_type="text/html"
    )

@staff_required
def case_request_edit(request, case_id):
    """
    HTMX endpoint — Staff requests permission to edit a closed ticket.
    """
    if request.method != "POST":
        return HttpResponse(status=405)
        
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")
    reason = request.POST.get("reason", "").strip()
    
    if case.status != CaseRecord.Status.CLOSED:
        return HttpResponse("Ticket is not closed.", status=400)
        
    case.edit_permission_status = CaseRecord.EditPermissionStatus.REQUESTED
    case.edit_requested_by = request.user
    case.edit_request_reason = reason
    case.save(update_fields=["edit_permission_status", "edit_requested_by", "edit_request_reason"])
    
    CaseAuditLog.objects.create(
        case=case,
        action=CaseAuditLog.ActionText.UPDATED,
        field_name="edit_permission_status",
        old_value="None",
        new_value="Requested",
        created_by=request.user,
    )
    
    return HttpResponse('<script>window.location.reload();</script>')

@staff_required
def case_approve_edit(request, case_id):
    """
    HTMX endpoint — SuperAdmin/Manager approves an edit request.
    """
    if request.method != "POST":
        return HttpResponse(status=405)
        
    if not (request.user.is_superuser or request.user.role_access == "Manager"):
        return HttpResponse("Unauthorized", status=403)

    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")

    case.edit_permission_status = CaseRecord.EditPermissionStatus.APPROVED
    case.save(update_fields=["edit_permission_status"])
    
    CaseAuditLog.objects.create(
        case=case,
        action=CaseAuditLog.ActionText.UPDATED,
        field_name="edit_permission_status",
        old_value="Requested",
        new_value="Approved",
        created_by=request.user,
    )
    
    return HttpResponse('<script>window.location.reload();</script>')

@staff_required
def case_reject_edit(request, case_id):
    """
    HTMX endpoint — SuperAdmin/Manager rejects an edit request.
    """
    if request.method != "POST":
        return HttpResponse(status=405)
        
    if not (request.user.is_superuser or request.user.role_access == "Manager"):
        return HttpResponse("Unauthorized", status=403)

    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")

    case.edit_permission_status = CaseRecord.EditPermissionStatus.REJECTED
    case.save(update_fields=["edit_permission_status"])
    
    CaseAuditLog.objects.create(
        case=case,
        action=CaseAuditLog.ActionText.UPDATED,
        field_name="edit_permission_status",
        old_value="Requested",
        new_value="Rejected",
        created_by=request.user,
    )
    
    return HttpResponse('<script>window.location.reload();</script>')

    
# =====================================================================
# Calendar View
# =====================================================================

@staff_required
def case_calendar(request):
    """
    Calendar view — shows cases on a monthly calendar.
    """
    import json
    from django.utils.timezone import localtime

    cases = CaseRecord.objects.select_related(
        "requester", "requester__unit", "category", "assigned_to"
    ).all()

    folder = request.GET.get("folder", "inbox")
    source_filter = request.GET.get("source")
    category_filter = request.GET.get("category")
    status_filter = request.GET.get("status")
    assigned_to_filter = request.GET.get("assigned_to")
    type_filter = request.GET.get("type")
    tags_filter = request.GET.get("tags", "").strip()
    followers_filter = request.GET.get("followers")
    read_filter = request.GET.get("read_status", "")
    unit_filter = request.GET.get("company_unit")

    search_query = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    # Folder specific base filtering
    if folder == "spam":
        cases = cases.filter(is_spam=True)
    elif folder == "archive":
        cases = cases.filter(is_archived=True)
    else:  # inbox
        cases = cases.filter(is_spam=False, is_archived=False, master_ticket__isnull=True)

    if source_filter:
        cases = cases.filter(source=source_filter)
    if category_filter:
        cases = cases.filter(category__slug=category_filter)
    if status_filter:
        status_list = [s.strip() for s in status_filter.split(',') if s.strip()]
        cases = cases.filter(status__in=status_list)
    if assigned_to_filter:
        if assigned_to_filter == "unassigned":
            cases = cases.filter(assigned_to__isnull=True)
        else:
            cases = cases.filter(assigned_to_id=assigned_to_filter)
    if type_filter:
        cases = cases.filter(case_type=type_filter)
    if unit_filter:
        cases = cases.filter(requester__unit_id=unit_filter)
    if tags_filter:
        cases = cases.filter(tags__icontains=tags_filter)
    if followers_filter:
        cases = cases.filter(followers__id=followers_filter)
    if read_filter == "unread":
        cases = cases.filter(has_unread_messages=True)
    elif read_filter == "unopened":
        cases = cases.filter(last_viewed_at__isnull=True)
    elif read_filter == "read":
        cases = cases.filter(has_unread_messages=False, last_viewed_at__isnull=False)

    if search_query:
        cases = cases.filter(
            Q(id__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(requester_name__icontains=search_query) |
            Q(requester__full_name__icontains=search_query) |
            Q(requester_email__icontains=search_query) |
            Q(requester__email__icontains=search_query) |
            Q(requester__phone_number__icontains=search_query) |
            Q(requester_unit_name__icontains=search_query) |
            Q(requester__unit__name__icontains=search_query)
        )
    if date_from:
        cases = cases.filter(created_at__date__gte=date_from)
    if date_to:
        cases = cases.filter(created_at__date__lte=date_to)

    # Annotate confidential access for the current user
    cases = _annotate_confidential_access(cases, request.user)

    # --- Dynamic date source for calendar ---
    date_source = request.GET.get("date_source", "created_at")
    DATE_SOURCE_FIELDS = {
        "created_at": "created_at",
        "last_viewed_at": "last_viewed_at",
        "resolution_due_at": "resolution_due_at",
        "response_due_at": "response_due_at",
    }
    date_field = DATE_SOURCE_FIELDS.get(date_source, "created_at")

    # Build JSON events for the calendar
    status_colors = {
        "Open": "#6366f1",
        "Investigating": "#f59e0b",
        "PendingInfo": "#0ea5e9",
        "Resolved": "#10b981",
        "Closed": "#64748b",
    }

    events = []
    for c in cases:
        dt_value = getattr(c, date_field, None)
        if not dt_value:
            continue  # Skip cases without the selected date
        has_access = getattr(c, "has_confidential_access_flag", True)
        events.append({
            "id": str(c.id),
            "title": f"[{c.case_number}] {'Confidential Ticket' if not has_access else c.subject[:40]}",
            "start": localtime(dt_value).strftime("%Y-%m-%d"),
            "url": f"/desk/cases/{c.id}/" if has_access else "",
            "color": "#94a3b8" if not has_access else status_colors.get(c.status, "#6366f1"),
            "status": c.status,
            "requester": c.requester.full_name if c.requester else "—",
        })

    # Calculate unread counts
    unread_inbox = CaseRecord.objects.filter(is_spam=False, is_archived=False, master_ticket__isnull=True, has_unread_messages=True).count()
    unread_archive = CaseRecord.objects.filter(is_archived=True, has_unread_messages=True).count()
    unread_spam = CaseRecord.objects.filter(is_spam=True, has_unread_messages=True).count()

    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    assignees = User.objects.filter(is_staff=True, is_active=True).order_by("first_name", "username")
    all_followers = User.objects.filter(is_active=True).order_by("first_name", "username")

    return render(request, "admin/case_calendar.html", {
        "events_json": json.dumps(events),
        "statuses": CaseRecord.Status.choices,
        "sources": CaseRecord.Source.choices,
        "categories": CaseCategory.objects.all(),
        "types": CaseRecord.Type.choices,
        "assignees": assignees,
        "all_followers": all_followers,
        "current_filters": {
            "folder": folder,
            "status": status_filter or "",
            "source": source_filter or "",
            "category": category_filter or "",
            "assigned_to": assigned_to_filter or "",
            "type": type_filter or "",
            "tags": tags_filter or "",
            "followers": followers_filter or "",
            "read_status": read_filter or "",
            "q": search_query,
            "date_from": date_from or "",
            "date_to": date_to or "",
        },
        "unread_inbox": unread_inbox,
        "unread_archive": unread_archive,
        "unread_spam": unread_spam,
        "date_source": date_source,
    })


# =====================================================================
# Export to Excel
# =====================================================================

@staff_required
@feature_required('audit_export')
def case_export_excel(request):
    """
    Export all cases to Excel (.xlsx) with all fields.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils.timezone import localtime

    cases = CaseRecord.objects.select_related(
        "requester", "requester__unit", "category", "assigned_to",
        "created_by", "updated_by"
    ).prefetch_related("followers").order_by("-created_at")

    # Apply filters if provided
    folder = request.GET.get("folder", "inbox")
    status_filter = request.GET.get("status")
    source_filter = request.GET.get("source")
    category_filter = request.GET.get("category")
    unit_filter = request.GET.get("company_unit")
    search_query = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    # Folder specific base filtering
    if folder == "spam":
        cases = cases.filter(is_spam=True)
    elif folder == "archive":
        cases = cases.filter(is_archived=True)
    else:  # inbox
        cases = cases.filter(is_spam=False, is_archived=False, master_ticket__isnull=True)

    if status_filter:
        cases = cases.filter(status=status_filter)
    if source_filter:
        cases = cases.filter(source=source_filter)
    if category_filter:
        cases = cases.filter(category__slug=category_filter)
    if unit_filter:
        cases = cases.filter(requester__unit_id=unit_filter)
    if search_query:
        cases = cases.filter(subject__icontains=search_query)
    if date_from:
        cases = cases.filter(created_at__date__gte=date_from)
    if date_to:
        cases = cases.filter(created_at__date__lte=date_to)

    # Annotate confidential access for the current user
    cases = _annotate_confidential_access(cases, request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Header style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = [
        "No", "Ticket Number", "Subject", "Status", "Priority", "Type", "Source", "Category",
        "Requester Name", "Requester Email", "Requester Phone",
        "Requester Unit", "Requester Job Role",
        "Problem Description", "Root Cause Analysis", "Solving Steps", "Quick Notes",
        "Tags", "Followers", "Reference Link",
        "Assigned To", "Response SLA", "Resolution SLA",
        "Created At", "Updated At",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    from core.excel_utils import safe_cell
    data_align = Alignment(vertical="top", wrap_text=True)
    MASKED = "[Confidential]"
    for row_idx, case in enumerate(cases, 2):
        has_access = getattr(case, "has_confidential_access_flag", True)
        row_data = [
            row_idx - 1,
            safe_cell(case.case_number),
            safe_cell(case.subject) if has_access else MASKED,
            safe_cell(case.get_status_display()),
            safe_cell(case.get_priority_display()),
            safe_cell(case.get_case_type_display()),
            safe_cell(case.get_source_display()),
            safe_cell(case.category.name if case.category else ""),
            safe_cell(case.requester.full_name if case.requester else case.requester_name),
            safe_cell(case.requester.email if case.requester else case.requester_email),
            safe_cell(case.requester.phone_number if case.requester else ""),
            safe_cell(case.requester.unit.name if case.requester and case.requester.unit else case.requester_unit_name),
            safe_cell(case.requester.job_role if case.requester else case.requester_job_role),
            safe_cell(case.problem_description or "") if has_access else MASKED,
            safe_cell(case.root_cause_analysis or "") if has_access else MASKED,
            safe_cell(case.solving_steps or "") if has_access else MASKED,
            safe_cell(case.quick_notes or "") if has_access else MASKED,
            safe_cell(case.tags or ""),
            safe_cell(", ".join([f.username for f in case.followers.all()])) if has_access else MASKED,
            safe_cell(case.link or "") if has_access else MASKED,
            safe_cell(case.assigned_to.username if case.assigned_to else "Unassigned") if has_access else MASKED,
            localtime(case.response_due_at).strftime("%d/%m/%Y %H:%M") if case.response_due_at else "",
            localtime(case.resolution_due_at).strftime("%d/%m/%Y %H:%M") if case.resolution_due_at else "",
            localtime(case.created_at).strftime("%d/%m/%Y %H:%M") if case.created_at else "",
            localtime(case.updated_at).strftime("%d/%m/%Y %H:%M") if case.updated_at else "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            cell.border = thin_border

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_length + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Response
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    now_str = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"RoC_Desk_Cases_{now_str}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    from core.models import AuditLog
    AuditLog.log(AuditLog.Action.EXPORT, request=request,
                 details={"filename": filename, "type": "case_list"})
    return response


# =====================================================================
# RCA Templates — Create / Delete
# =====================================================================

@manager_or_admin_required
@require_POST
def rca_template_create(request):
    """Create a new RCA Template. Managers/SuperAdmins only."""
    name = request.POST.get("name", "").strip()
    rca_text = request.POST.get("rca_text", "").strip()
    solving_steps_text = request.POST.get("solving_steps_text", "").strip()
    category_id = request.POST.get("category_id", "").strip()
    redirect_case_id = request.POST.get("redirect_case_id", "").strip()

    if name:
        category = None
        if category_id:
            category = CaseCategory.objects.filter(id=category_id).first()
        RCATemplate.objects.create(
            name=name,
            rca_text=rca_text,
            solving_steps_text=solving_steps_text,
            category=category,
            created_by=request.user,
            updated_by=request.user,
        )
        messages.success(request, f"Template '{name}' created.")
    else:
        messages.error(request, "Template name is required.")

    if redirect_case_id:
        return redirect("desk:case_detail", case_id=redirect_case_id)
    return redirect("desk:case_list")


@manager_or_admin_required
@require_POST
def rca_template_delete(request, template_id):
    """Delete an RCA Template. Managers/SuperAdmins only."""
    tpl = get_object_or_404(RCATemplate, id=template_id)
    tpl_name = tpl.name
    tpl.delete()
    messages.success(request, f"Template '{tpl_name}' deleted.")

    redirect_case_id = request.POST.get("redirect_case_id", "").strip()
    if redirect_case_id:
        return redirect("desk:case_detail", case_id=redirect_case_id)
    return redirect("desk:case_list")


@staff_required
def notification_bell(request):
    """
    HTMX endpoint — returns a dropdown of recent cases and unread incoming messages (last 24 hours).
    """
    from datetime import timedelta
    from django.utils import timezone
    
    time_threshold = timezone.now() - timedelta(hours=24)
    
    # Get previously clicked notifications from the session
    read_notifs = request.session.get("read_notifications", [])
    read_case_ids = [n.replace("case_", "") for n in read_notifs if n.startswith("case_")]
    read_msg_ids = [n.replace("msg_", "") for n in read_notifs if n.startswith("msg_")]
    read_mention_ids = [n.replace("mention_", "") for n in read_notifs if n.startswith("mention_")]
    
    # 1. New Tickets in the last 24h (exclude confidential without access)
    _conf_filter = Q()  # no filter needed
    if request.user.role_access != "SuperAdmin" and not request.user.can_handle_confidential:
        _conf_filter = Q(category__is_confidential=False)
    recent_cases = CaseRecord.objects.select_related("requester").filter(
        created_at__gte=time_threshold
    ).filter(_conf_filter).exclude(id__in=read_case_ids).order_by("-created_at")[:10]

    # 2. Unread Incoming Messages in the last 24h (exclude confidential without access)
    _msg_conf_filter = Q()
    if request.user.role_access != "SuperAdmin" and not request.user.can_handle_confidential:
        _msg_conf_filter = Q(case__category__is_confidential=False)
    recent_unread_messages = Message.objects.select_related(
        "case", "sender_employee"
    ).filter(
        direction=Message.Direction.INBOUND,
        is_read=False,
        created_at__gte=time_threshold
    ).filter(_msg_conf_filter).exclude(id__in=read_msg_ids).order_by("-created_at")[:10]
    
    # 3. Recent Mentions in Internal Notes
    recent_mentions = CaseComment.objects.select_related("case", "author").filter(
        mentions=request.user,
        created_at__gte=time_threshold
    ).exclude(id__in=read_mention_ids).order_by("-created_at")[:10]
    
    total_notifications = recent_cases.count() + recent_unread_messages.count() + recent_mentions.count()
    
    return render(request, "partials/notification_bell.html", {
        "recent_cases": recent_cases,
        "recent_unread_messages": recent_unread_messages,
        "recent_mentions": recent_mentions,
        "total_notifications": total_notifications,
    })


@staff_required
def mark_notification_read(request, notif_type, notif_id):
    """
    Marks a specific notification as 'read' in the user's session 
    so it no longer appears in the bell dropdown, then redirects
    to the target case detail URL.
    """
    notif_key = f"{notif_type}_{notif_id}"
    read_notifs = request.session.get("read_notifications", [])
    if notif_key not in read_notifs:
        read_notifs.append(notif_key)
        request.session["read_notifications"] = read_notifs
    
    # Both types ultimately lead to the Ticket Detail page
    target_case_id = notif_id
    if notif_type == "msg":
        try:
            msg = Message.objects.get(id=notif_id)
            target_case_id = msg.case.id
        except Message.DoesNotExist:
            target_case_id = None
    elif notif_type == "mention":
        try:
            comment = CaseComment.objects.get(id=notif_id)
            target_case_id = comment.case.id
        except CaseComment.DoesNotExist:
            target_case_id = None
            
    if target_case_id:
        return redirect("desk:case_detail", case_id=target_case_id)
    return redirect("desk:case_list")


# =====================================================================
# Internal Ticket Comments
# =====================================================================

@staff_required
def case_add_comment(request, case_id):
    """
    HTMX endpoint for saving and retrieving internal staff comments on a case.
    Also parses @username mentions.
    """
    import re
    from core.models import User
    
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")

    if request.method == "POST":
        body = request.POST.get("comment_body", "").strip()
        if body:
            comment = CaseComment.objects.create(
                case=case,
                author=request.user,
                body=body
            )
            
            # Parse @usernames
            usernames = set(re.findall(r"@([a-zA-Z0-9_]+)", body))
            if usernames:
                mentioned_users = User.objects.filter(username__in=usernames)
                if mentioned_users.exists():
                    comment.mentions.add(*mentioned_users)

    # Return the refreshed comments partial
    return render(request, "partials/case_comments.html", {
        "case": case,
        "comments": case.internal_comments.all(),
    })

# =====================================================================
# WhatsApp Integration Dashboard
# =====================================================================

@staff_required
@feature_required('whatsapp')
def whatsapp_disconnect_view(request):
    """Delete and recreate the WA instance so Evolution API shows a fresh QR code."""
    from gateways.services import EvolutionAPIService
    from core.models import SiteConfig
    from django.contrib import messages
    import time

    if request.method != "POST":
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(["POST"])

    site_url = SiteConfig.get_solo().site_url.rstrip("/")
    webhook_url = f"{site_url}/gateways/evolution/webhook/"

    svc = EvolutionAPIService()
    success = svc.disconnect_and_reset(webhook_url)

    if success:
        # After delete+recreate, Evolution API needs a moment to initialise.
        # Wait up to 6 seconds for the state to settle to "connecting" (QR shown).
        for _ in range(6):
            time.sleep(1)
            state_data = svc.get_instance_state()
            state = state_data.get("instance", {}).get("state") if state_data else None
            if state and state not in ["open", "connected"]:
                break

        messages.success(request, "WhatsApp session disconnected. Scan the new QR code to reconnect.")
    else:
        messages.error(request, "Failed to disconnect WhatsApp session. Check the Evolution API logs.")

    from django.shortcuts import redirect
    return redirect("desk:whatsapp_status")


def _compute_wa_warnings(cfg) -> dict:
    """
    Return {field_name: 'warning'|'danger'} for SiteConfig WA fields
    that fall outside recommended safe ranges.
    """
    w = {}

    # Recipient cooldown — lower is riskier
    if cfg.wa_rate_recipient_cooldown < 3:
        w['wa_rate_recipient_cooldown'] = 'danger'
    elif cfg.wa_rate_recipient_cooldown < 5:
        w['wa_rate_recipient_cooldown'] = 'warning'

    # Burst max — higher is riskier
    if cfg.wa_rate_burst_max > 20:
        w['wa_rate_burst_max'] = 'danger'
    elif cfg.wa_rate_burst_max > 12:
        w['wa_rate_burst_max'] = 'warning'

    # Per-minute max — higher is riskier
    if cfg.wa_rate_minute_max > 35:
        w['wa_rate_minute_max'] = 'danger'
    elif cfg.wa_rate_minute_max > 20:
        w['wa_rate_minute_max'] = 'warning'

    # Per-hour max — higher is riskier
    if cfg.wa_rate_hour_max > 600:
        w['wa_rate_hour_max'] = 'danger'
    elif cfg.wa_rate_hour_max > 350:
        w['wa_rate_hour_max'] = 'warning'

    # CB max disconnects — higher = too lenient
    if cfg.wa_cb_max_disconnects > 6:
        w['wa_cb_max_disconnects'] = 'danger'
    elif cfg.wa_cb_max_disconnects > 3:
        w['wa_cb_max_disconnects'] = 'warning'

    # CB block hours — lower = too short recovery
    if cfg.wa_cb_block_hours < 1:
        w['wa_cb_block_hours'] = 'danger'
    elif cfg.wa_cb_block_hours < 2:
        w['wa_cb_block_hours'] = 'warning'

    # Read pause min — lower is riskier
    if cfg.wa_read_pause_min_ms < 200:
        w['wa_read_pause_min_ms'] = 'danger'
    elif cfg.wa_read_pause_min_ms < 500:
        w['wa_read_pause_min_ms'] = 'warning'

    # Read pause max — lower is riskier
    if cfg.wa_read_pause_max_ms < 800:
        w['wa_read_pause_max_ms'] = 'danger'
    elif cfg.wa_read_pause_max_ms < 1500:
        w['wa_read_pause_max_ms'] = 'warning'

    # Typing speed — higher = bot-like
    if cfg.wa_typing_speed_cps > 65:
        w['wa_typing_speed_cps'] = 'danger'
    elif cfg.wa_typing_speed_cps > 40:
        w['wa_typing_speed_cps'] = 'warning'

    # Opt-in TTL — too short or too long both carry risk
    if cfg.wa_opt_in_ttl_days < 3 or cfg.wa_opt_in_ttl_days > 30:
        w['wa_opt_in_ttl_days'] = 'danger'
    elif cfg.wa_opt_in_ttl_days < 5 or cfg.wa_opt_in_ttl_days > 14:
        w['wa_opt_in_ttl_days'] = 'warning'

    return w


@staff_required
@feature_required('whatsapp')
def whatsapp_status_view(request):
    """
    Dashboard view for monitoring Evolution API WhatsApp connection.
    Handles three POST forms: wa_notif, wa_config (gateway/hours), wa_limits (rate/pacing).
    """
    from gateways.services import EvolutionAPIService
    from core.models import NotificationConfig, SiteConfig, User as UserModel
    from django.contrib import messages
    from datetime import datetime

    # Handle notification config save
    if request.method == 'POST' and request.POST.get('form_type') == 'wa_notif':
        notif_cfg = NotificationConfig.get_solo()
        notif_cfg.notify_new_ticket_whatsapp = 'notify_new_ticket_whatsapp' in request.POST
        wa_recipients = request.POST.getlist('whatsapp_recipients')
        notif_cfg.whatsapp_recipients = ', '.join(r for r in wa_recipients if r.strip())
        notif_cfg.updated_by = request.user
        notif_cfg.save()
        messages.success(request, "WhatsApp notification settings updated successfully.")
        return redirect('desk:whatsapp_status')

    # Handle WA gateway config save (instance names + business hours)
    if request.method == 'POST' and request.POST.get('form_type') == 'wa_config':
        site_cfg = SiteConfig.get_solo()
        site_cfg.wa_inbound_enabled = 'wa_inbound_enabled' in request.POST
        site_cfg.wa_outbound_enabled = 'wa_outbound_enabled' in request.POST
        site_cfg.wa_main_instance = request.POST.get('wa_main_instance', '').strip()
        site_cfg.wa_notif_instance = request.POST.get('wa_notif_instance', '').strip()
        try:
            hour_start = int(request.POST.get('wa_business_hour_start', 7))
            hour_end = int(request.POST.get('wa_business_hour_end', 20))
            site_cfg.wa_business_hour_start = max(0, min(23, hour_start))
            site_cfg.wa_business_hour_end = max(0, min(23, hour_end))
        except (ValueError, TypeError):
            pass
        raw_days = request.POST.getlist('wa_business_days')
        valid_days = [d for d in raw_days if d.isdigit() and 0 <= int(d) <= 6]
        site_cfg.wa_business_days = ','.join(valid_days) if valid_days else '0,1,2,3,4'
        site_cfg.updated_by = request.user
        site_cfg.save()
        messages.success(request, "WhatsApp gateway configuration saved successfully.")
        return redirect('desk:whatsapp_status')

    # Handle rate limits & pacing config save
    if request.method == 'POST' and request.POST.get('form_type') == 'wa_limits':
        site_cfg = SiteConfig.get_solo()

        def _int(key, default):
            try:
                return max(1, int(request.POST.get(key, default)))
            except (ValueError, TypeError):
                return default

        site_cfg.wa_rate_recipient_cooldown = _int('wa_rate_recipient_cooldown', 10)
        site_cfg.wa_rate_burst_max = _int('wa_rate_burst_max', 8)
        site_cfg.wa_rate_minute_max = _int('wa_rate_minute_max', 15)
        site_cfg.wa_rate_hour_max = _int('wa_rate_hour_max', 200)
        site_cfg.wa_cb_max_disconnects = _int('wa_cb_max_disconnects', 2)
        site_cfg.wa_cb_block_hours = _int('wa_cb_block_hours', 4)
        site_cfg.wa_read_pause_min_ms = _int('wa_read_pause_min_ms', 1000)
        site_cfg.wa_read_pause_max_ms = _int('wa_read_pause_max_ms', 3500)
        site_cfg.wa_typing_speed_cps = _int('wa_typing_speed_cps', 25)
        site_cfg.wa_opt_in_ttl_days = _int('wa_opt_in_ttl_days', 7)
        site_cfg.updated_by = request.user
        site_cfg.save()
        # Invalidate cached rate config so changes take effect immediately
        from django.core.cache import cache as _djcache
        _djcache.delete("wa_rate_cfg")
        _djcache.delete("wa_pacing_cfg")
        messages.success(request, "Rate limits and pacing configuration saved.")
        return redirect('desk:whatsapp_status')

    svc = EvolutionAPIService()
    state_data = svc.get_instance_state()
    info_data = svc.get_instance_info()

    instance_state = state_data.get("instance", {}).get("state", "UNKNOWN") if state_data else "ERROR"

    last_connected = None
    if info_data and "updatedAt" in info_data:
        dt_str = info_data.get("updatedAt").replace('Z', '+00:00')
        try:
            last_connected = datetime.fromisoformat(dt_str)
        except ValueError:
            pass

    qr_base64 = None
    if instance_state not in ["open", "connected"]:
        qr_data = svc.get_qr_code()
        if qr_data and "base64" in qr_data:
            qr_base64 = qr_data.get("base64")

    # If a disconnect was just requested but the instance still reports
    # connected, keep the status card polling until it actually flips.
    force_poll = False
    if request.session.get('wa_pending_disconnect'):
        if instance_state in ["open", "connected"]:
            force_poll = True
        else:
            del request.session['wa_pending_disconnect']

    if request.headers.get('HX-Request') == 'true':
        return render(request, "partials/whatsapp_status_card.html", {
            "instance_state": instance_state,
            "qr_base64": qr_base64,
            "last_connected": last_connected,
            "force_poll": force_poll,
        })

    # Build notification recipient options from active staff
    notif_cfg = NotificationConfig.get_solo()
    selected_wa = {p.strip() for p in (notif_cfg.whatsapp_recipients or '').split(',') if p.strip()}
    staff_wa_values = set()
    wa_options = []
    for u in UserModel.objects.filter(is_active=True).order_by('username'):
        phone = getattr(u, 'phone_number', None)
        if phone:
            staff_wa_values.add(phone)
            name = u.get_full_name() or u.username
            wa_options.append({'value': phone, 'label': f"{name} ({phone})", 'selected': phone in selected_wa})
    for v in selected_wa:
        if v not in staff_wa_values:
            wa_options.append({'value': v, 'label': v, 'selected': True})

    from gateways.throttle import WARateLimiter, get_circuit_status, _get_business_hours_config
    site_cfg = SiteConfig.get_solo()
    daily_sent = WARateLimiter.get_daily_count()
    daily_limit = site_cfg.get_wa_daily_limit()
    circuit = get_circuit_status()

    # Effective instance names (DB overrides .env)
    from django.conf import settings as djsettings
    effective_main = site_cfg.wa_main_instance.strip() or getattr(djsettings, 'EVOLUTION_INSTANCE_NAME', '')
    effective_notif = site_cfg.wa_notif_instance.strip() or getattr(djsettings, 'EVOLUTION_NOTIF_INSTANCE_NAME', '').strip()

    # Business hours config for display
    bh_start, bh_end, bh_days = _get_business_hours_config()
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    bh_days_display = ', '.join(day_names[d] for d in sorted(bh_days) if d < 7)
    wa_day_options = [
        {'num': str(i), 'label': day_names[i], 'checked': i in bh_days}
        for i in range(7)
    ]

    wa_warnings = _compute_wa_warnings(site_cfg)

    return render(request, "desk/whatsapp_status.html", {
        "instance_state": instance_state,
        "qr_base64": qr_base64,
        "last_connected": last_connected,
        "force_poll": force_poll,
        "notif_cfg": notif_cfg,
        "wa_options": wa_options,
        "daily_sent": daily_sent,
        "daily_limit": daily_limit,
        "wa_instance_activated_at": site_cfg.wa_instance_activated_at,
        "circuit": circuit,
        "notif_instance": effective_notif,
        "site_cfg": site_cfg,
        "effective_main": effective_main,
        "bh_start": bh_start,
        "bh_end": bh_end,
        "bh_days": list(bh_days),
        "bh_days_display": bh_days_display,
        "wa_day_options": wa_day_options,
        "wa_warnings": wa_warnings,
    })


@require_POST
@login_required
def wa_circuit_reset_view(request):
    """Manually reset the WA circuit breaker from the dashboard."""
    from gateways.throttle import reset_circuit
    from django.contrib import messages

    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "Only admins can reset the circuit breaker.")
        return redirect("desk:whatsapp_status")

    reset_circuit()
    messages.success(request, "Circuit breaker reset. Outbound WA sends re-enabled.")
    return redirect("desk:whatsapp_status")


# =====================================================================
# Apps Connection Hub
# =====================================================================

@login_required
def apps_connection_view(request):
    """
    Hub page listing all external app integrations.
    SuperAdmin only.
    """
    if request.user.role_access != 'SuperAdmin':
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    from core.models import TeamsConfig, NotificationConfig
    teams_cfg = TeamsConfig.get_solo()
    notif_cfg = NotificationConfig.get_solo()

    return render(request, 'desk/apps_connection.html', {
        'teams_cfg': teams_cfg,
        'notif_cfg': notif_cfg,
    })


# =====================================================================
# Email Settings Dashboard
# =====================================================================

@staff_required
@feature_required('email_settings')
def email_settings_view(request):
    """
    Dashboard view for monitoring and updating global Email Settings.
    Also handles saving Email internal notification recipients.
    """
    from core.models import EmailConfig, NotificationConfig, User as UserModel
    from core.forms import EmailConfigForm
    from django.contrib import messages

    email_config = EmailConfig.get_solo()

    if request.method == "POST":
        form_type = request.POST.get('form_type', 'email_config')

        if form_type == 'email_notif':
            notif_cfg = NotificationConfig.get_solo()
            notif_cfg.notify_new_ticket_email = 'notify_new_ticket_email' in request.POST
            email_recipients = request.POST.getlist('email_recipients')
            notif_cfg.email_recipients = ', '.join(r for r in email_recipients if r.strip())
            notif_cfg.updated_by = request.user
            notif_cfg.save()
            messages.success(request, "Email notification settings updated successfully.")
            return redirect("desk:email_settings")

        else:
            form = EmailConfigForm(request.POST, instance=email_config)
            if form.is_valid():
                form.save()
                messages.success(request, "Email configuration updated successfully! Background services will use the new credentials on their next run.")
                return redirect("desk:email_settings")
            else:
                messages.error(request, "Please correct the errors below.")
                notif_cfg = NotificationConfig.get_solo()
                return render(request, "desk/email_settings.html", _email_settings_ctx(form, email_config, notif_cfg))
    else:
        form = EmailConfigForm(instance=email_config)

    notif_cfg = NotificationConfig.get_solo()
    return render(request, "desk/email_settings.html", _email_settings_ctx(form, email_config, notif_cfg))


def _email_settings_ctx(form, email_config, notif_cfg):
    from core.models import User as UserModel
    selected_emails = {e.strip() for e in (notif_cfg.email_recipients or '').split(',') if e.strip()}
    staff_email_values = set()
    email_options = []
    for u in UserModel.objects.filter(is_active=True).order_by('username'):
        if u.email:
            staff_email_values.add(u.email)
            name = u.get_full_name() or u.username
            email_options.append({'value': u.email, 'label': f"{name} ({u.email})", 'selected': u.email in selected_emails})
    for v in selected_emails:
        if v not in staff_email_values:
            email_options.append({'value': v, 'label': v, 'selected': True})
    return {
        'form': form,
        'notif_cfg': notif_cfg,
        'email_options': email_options,
    }


# =====================================================================
# Microsoft Teams Settings
# =====================================================================

@login_required
def teams_settings_view(request):
    """
    Dedicated configuration page for Microsoft Teams integration.
    Covers both the free Incoming Webhook (1-way) and the paid Bot Framework (2-way).
    SuperAdmin only.
    """
    if request.user.role_access != 'SuperAdmin':
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    from core.models import TeamsConfig, NotificationConfig
    from core.forms import TeamsConfigForm
    from django.contrib import messages as dj_messages

    teams_cfg = TeamsConfig.get_solo()
    notif_cfg = NotificationConfig.get_solo()
    form = TeamsConfigForm(instance=teams_cfg)
    test_result = None

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'test_webhook':
            import requests as http_requests
            webhook_url = teams_cfg.incoming_webhook_url
            if not webhook_url:
                test_result = {'status': 'error', 'message': 'Webhook URL is not configured yet.'}
            else:
                try:
                    from core.models import SiteConfig
                    site_name = SiteConfig.get_solo().site_name
                    payload = {
                        "type": "message",
                        "attachments": [{
                            "contentType": "application/vnd.microsoft.card.adaptive",
                            "contentUrl": None,
                            "content": {
                                "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                                "type": "AdaptiveCard",
                                "version": "1.4",
                                "body": [{
                                    "type": "TextBlock",
                                    "text": f"✅ Test connection from {site_name} — webhook is working!",
                                    "weight": "Bolder",
                                    "size": "Medium",
                                    "wrap": True,
                                }],
                            },
                        }],
                    }
                    resp = http_requests.post(webhook_url, json=payload, timeout=10)
                    if resp.status_code == 200:
                        test_result = {'status': 'success', 'message': 'Success! A test message has been posted to your Teams channel.'}
                    else:
                        test_result = {'status': 'error', 'message': f'HTTP {resp.status_code}: {resp.text[:300]}'}
                except Exception as exc:
                    test_result = {'status': 'error', 'message': str(exc)[:300]}

        else:
            form = TeamsConfigForm(request.POST, instance=teams_cfg)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.updated_by = request.user
                obj.save()
                notif_cfg.notify_new_ticket_teams = 'notify_new_ticket_teams' in request.POST
                notif_cfg.updated_by = request.user
                notif_cfg.save()
                dj_messages.success(request, "Teams configuration saved successfully.")
                return redirect('desk:teams_settings')
            else:
                dj_messages.error(request, "Please review the form errors below.")

    return render(request, 'desk/teams_settings.html', {
        'form': form,
        'teams_cfg': teams_cfg,
        'notif_cfg': notif_cfg,
        'test_result': test_result,
    })


# =====================================================================
# Internal Staff Contacts API
# =====================================================================

@login_required
def api_staff_contacts(request):
    """
    Returns active staff users with their email and phone for notification selectors.
    """
    from django.http import JsonResponse
    from core.models import User as UserModel

    email_contacts, wa_contacts = [], []
    for u in UserModel.objects.filter(is_active=True).order_by('username'):
        name = u.get_full_name() or u.username
        if u.email:
            email_contacts.append({'value': u.email, 'label': f"{name} ({u.email})"})
        phone = getattr(u, 'phone_number', None)
        if phone:
            wa_contacts.append({'value': phone, 'label': f"{name} ({phone})"})
    return JsonResponse({'email': email_contacts, 'whatsapp': wa_contacts})


# =====================================================================
# Dynamic Form Builder
# =====================================================================

@staff_required
@feature_required('form_builder')
def form_list_view(request):
    """
    List of all Dynamic Forms.
    """
    from core.models import DynamicForm
    from django.db.models import Q
    from django.core.paginator import Paginator

    query = request.GET.get('q', '').strip()
    
    forms_qs = DynamicForm.objects.all().order_by('-created_at')
    
    if query:
        forms_qs = forms_qs.filter(
            Q(title__icontains=query) | 
            Q(slug__icontains=query)
        )

    paginator = Paginator(forms_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "desk/forms/list.html", {
        "forms": page_obj,
        "search_query": query
    })


@staff_required
@feature_required('form_builder')
def form_create_view(request):
    """
    Create a new DynamicForm. Auditors are not allowed.
    """
    if request.user.role_access == User.RoleAccess.AUDITOR:
        return HttpResponseForbidden("Access denied. Auditors cannot create forms.")
    from core.forms import DynamicFormForm
    from django.contrib import messages
    import uuid

    if request.method == "POST":
        form = DynamicFormForm(request.POST, request.FILES)
        if form.is_valid():
            new_form = form.save(commit=False)
            new_form.created_by = request.user
            # Ensure slug is truly unique if not explicitly provided or if duplicate
            if not new_form.slug:
                new_form.slug = str(uuid.uuid4())[:8]
            new_form.save()
            messages.success(request, f"Form '{new_form.title}' created! Now add some fields.")
            return redirect("desk:form_edit", pk=new_form.pk)
    else:
        form = DynamicFormForm()

    return render(request, "desk/forms/create.html", {"form": form})


@staff_required
@feature_required('form_builder')
@require_POST
def form_delete_view(request, pk):
    """Delete a DynamicForm and all related fields/submissions."""
    from core.models import DynamicForm
    from django.contrib import messages

    form = get_object_or_404(DynamicForm, pk=pk)
    title = form.title
    form.delete()
    messages.success(request, f"Form '{title}' has been deleted.")
    return redirect("desk:form_list")


@staff_required
@feature_required('form_builder')
@require_POST
def form_duplicate_view(request, pk):
    """
    Duplicate a DynamicForm and all its fields.
    Does NOT copy background_image and header_image.
    Appends '(copy)' to the title and generates a new slug.
    """
    from core.models import DynamicForm, FormField
    from django.contrib import messages
    import uuid

    original_form = get_object_or_404(DynamicForm, pk=pk)

    # Clone the form instance
    new_form = DynamicForm.objects.get(pk=pk)
    new_form.pk = None
    new_form.id = uuid.uuid4()
    new_form.title = f"{original_form.title} (copy)"
    # Ensure slug is unique by generating a new one
    new_form.slug = str(uuid.uuid4())[:8]
    new_form.created_by = request.user
    
    # Do NOT copy images
    new_form.background_image = None
    new_form.header_image = None
    
    new_form.save()

    # Clone all fields, maintaining order
    original_fields = original_form.fields.all()
    for field in original_fields:
        field.pk = None
        field.id = uuid.uuid4()
        field.form = new_form
        field.save()

    messages.success(request, f"Form duplicated successfully!")
    return redirect("desk:form_edit", pk=new_form.pk)


@staff_required
@feature_required('form_builder')
def form_edit_view(request, pk):
    """
    Drag and drop builder to edit form fields and form settings. Auditors are not allowed.
    """
    if request.user.role_access == User.RoleAccess.AUDITOR:
        return HttpResponseForbidden("Access denied. Auditors cannot edit forms.")
    from core.models import DynamicForm, FormField
    from core.forms import DynamicFormForm
    from django.contrib import messages
    from django.db import models
    import json
    import uuid

    instance = get_object_or_404(DynamicForm, pk=pk)
    
    # Handle AJAX/HTMX Field updates
    if request.headers.get('HX-Request') == 'true':
        action = request.POST.get('action')
        
        if action == 'add_field':
            field_type = request.POST.get('field_type', FormField.FieldTypes.TEXT)
            label = request.POST.get('label', 'New Question')
            help_text = request.POST.get('help_text', '')
            is_required = request.POST.get('is_required') == 'true'
            
            choices = []
            field_settings = {}
            if field_type in [FormField.FieldTypes.DROPDOWN, FormField.FieldTypes.RADIO, FormField.FieldTypes.CHECKBOX, FormField.FieldTypes.SURVEY]:
                choices = request.POST.getlist('choices[]')
                if field_type in [FormField.FieldTypes.DROPDOWN, FormField.FieldTypes.RADIO]:
                    logic_vals = request.POST.getlist('choice_logic[]')
                    logic_dict = {}
                    for ch, val in zip(choices, logic_vals):
                        if ch.strip() and val.strip():
                            logic_dict[ch.strip()] = val.strip()
                    if logic_dict:
                        field_settings['logic'] = logic_dict

            if field_type == FormField.FieldTypes.NUMBER:
                field_settings.update({
                    'number_type': request.POST.get('settings_number_type', 'general'),
                    'currency_type': request.POST.get('settings_currency_type', 'idr'),
                    'phone_code': request.POST.get('settings_phone_code', '+62'),
                })
                
            condition_field_id = request.POST.get('condition_field', '').strip()
            condition_value = request.POST.get('condition_value', '').strip()
            if condition_field_id and condition_value:
                try:
                    cond_obj = FormField.objects.get(id=condition_field_id, form=instance)
                    field_settings['condition'] = {
                        'field_id': condition_field_id,
                        'value': condition_value,
                        'field_label': cond_obj.label,
                    }
                except FormField.DoesNotExist:
                    pass

            # Put at bottom
            last_order = instance.fields.count()

            FormField.objects.create(
                form=instance,
                field_type=field_type,
                label=label,
                help_text=help_text,
                is_required=is_required,
                choices=choices,
                settings=field_settings,
                order=last_order + 1,
                created_by=request.user
            )
            return render(request, "desk/forms/partials/field_list.html", {"fields": instance.fields.all(), "dynamic_form": instance})
            
        elif action == 'duplicate_field':
            field_id = request.POST.get('field_id')
            original_field = get_object_or_404(FormField, id=field_id, form=instance)
            
            # Shift order of subsequent fields down by 1
            FormField.objects.filter(form=instance, order__gt=original_field.order).update(
                order=models.F('order') + 1
            )
            
            # Clone field
            new_field = original_field
            new_field.pk = None
            new_field.id = uuid.uuid4()
            new_field.order = original_field.order + 1
            # Append copy to label so user knows it's the duplicate
            new_field.label = f"{original_field.label} (copy)"
            new_field.save()
            
            return render(request, "desk/forms/partials/field_list.html", {"fields": instance.fields.all(), "dynamic_form": instance})
            
        elif action == 'delete_field':
            field_id = request.POST.get('field_id')
            FormField.objects.filter(id=field_id, form=instance).delete()
            return render(request, "desk/forms/partials/field_list.html", {"fields": instance.fields.all(), "dynamic_form": instance})
            
        elif action == 'reorder':
            order_data = json.loads(request.POST.get('order_data', '[]'))
            for item in order_data:
                FormField.objects.filter(id=item['id'], form=instance).update(order=item['order'])
            return HttpResponse(status=200)

        elif action == 'edit_field':
            field_id = request.POST.get('field_id')
            field = get_object_or_404(FormField, id=field_id, form=instance)
            
            field.label = request.POST.get('label', field.label)
            field.help_text = request.POST.get('help_text', '')
            field.is_required = request.POST.get('is_required') == 'true'
            
            new_type = request.POST.get('field_type')
            if new_type:
                field.field_type = new_type
            
            if not field.settings: field.settings = {}
            if field.field_type in [FormField.FieldTypes.DROPDOWN, FormField.FieldTypes.RADIO, FormField.FieldTypes.CHECKBOX, FormField.FieldTypes.SURVEY]:
                choices = request.POST.getlist('choices[]')
                if choices:
                    field.choices = choices
                
                if field.field_type in [FormField.FieldTypes.DROPDOWN, FormField.FieldTypes.RADIO]:
                    logic_vals = request.POST.getlist('choice_logic[]')
                    logic_dict = {}
                    for ch, val in zip(field.choices, logic_vals):
                        if ch.strip() and val.strip():
                            logic_dict[ch.strip()] = val.strip()
                    if logic_dict:
                        field.settings['logic'] = logic_dict
                    elif 'logic' in field.settings:
                        del field.settings['logic']
            else:
                field.choices = []
                if 'logic' in field.settings:
                    del field.settings['logic']

            if field.field_type == FormField.FieldTypes.NUMBER:
                field.settings['number_type'] = request.POST.get('settings_number_type', 'general')
                field.settings['currency_type'] = request.POST.get('settings_currency_type', 'idr')
                field.settings['phone_code'] = request.POST.get('settings_phone_code', '+62')

            condition_field_id = request.POST.get('condition_field', '').strip()
            condition_value = request.POST.get('condition_value', '').strip()
            if condition_field_id and condition_value:
                try:
                    cond_obj = FormField.objects.get(id=condition_field_id, form=instance)
                    field.settings['condition'] = {
                        'field_id': condition_field_id,
                        'value': condition_value,
                        'field_label': cond_obj.label,
                    }
                except FormField.DoesNotExist:
                    field.settings.pop('condition', None)
            else:
                field.settings.pop('condition', None)

            field.updated_by = request.user
            field.save()
            return render(request, "desk/forms/partials/field_list.html", {"fields": instance.fields.all(), "dynamic_form": instance})

        elif action == 'send_invitation':
            from django.core.mail import send_mail
            from django.conf import settings
            import json as _json

            raw_emails = request.POST.get('emails', '')
            email_list = [e.strip() for e in raw_emails.replace(';', ',').split(',') if e.strip()]
            form_url = request.build_absolute_uri(f'/f/{instance.slug}/')
            
            sent_count = 0
            for email in email_list:
                try:
                    send_mail(
                        subject=f'You are invited to fill out: {instance.title}',
                        message=f'Hello,\n\nYou have been invited to fill out the form "{instance.title}".\n\n{instance.description}\n\nPlease click the link below to access the form:\n{form_url}\n\nBest regards,\n{request.user.get_full_name() or request.user.username}',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[email],
                        fail_silently=True,
                    )
                    sent_count += 1
                except Exception:
                    pass
            
            return JsonResponse({'status': 'ok', 'sent': sent_count})

    # Standard form updating
    if request.method == "POST" and not request.headers.get('HX-Request'):
        form = DynamicFormForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            f = form.save(commit=False)
            f.updated_by = request.user
            f.save()
            messages.success(request, "Form settings updated successfully.")
            return redirect("desk:form_edit", pk=instance.pk)
    else:
        form = DynamicFormForm(instance=instance)

    return render(request, "desk/forms/builder.html", {
        "dynamic_form": instance,
        "form": form,
        "fields": instance.fields.all()
    })


@staff_required
def form_responses_view(request, pk):
    """
    View all submissions for a specific DynamicForm.
    Also calculates aggregated metrics for choice-based fields to power the dashboard.
    """
    from core.models import DynamicForm
    from django.core.paginator import Paginator
    from django.db.models import Q
    import collections
    
    instance = get_object_or_404(DynamicForm, pk=pk)
    
    query = request.GET.get('q', '').strip()
    
    submissions_qs = instance.submissions.all().order_by('-submitted_at')
    
    if query:
        submissions_qs = submissions_qs.filter(
            Q(submitted_by__email__icontains=query) |
            Q(submitted_by__first_name__icontains=query) |
            Q(submitted_by__last_name__icontains=query) |
            Q(submitted_by__username__icontains=query)
        )
        
    paginator = Paginator(submissions_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    fields = instance.fields.all()
    
    # Calculate metrics for chart/progress bar visualization
    metrics = {}
    
    # Pre-filter fields that are choice-based
    choice_fields = [f for f in fields if f.field_type in ['dropdown', 'radio', 'checkbox', 'survey']]
    
    if choice_fields and submissions_qs.exists():
        # Iterate over all submissions recursively for metrics instead of just the page
        for field in choice_fields:
            field_id_str = str(field.id)
            counter = collections.Counter()
            
            for sub in submissions_qs:
                ans = sub.answers.get(field_id_str)
                if ans:
                    if isinstance(ans, list):
                        for a in ans:
                            counter[a] += 1
                    else:
                        counter[ans] += 1
            
            total_responses = sum(counter.values())
            
            if total_responses > 0:
                # Get Top 5 most frequent
                top_5 = counter.most_common(5)
                top_5_count = sum(count for _, count in top_5)
                
                # If there are more than 5 distinct answers, group remaining as "Others"
                others_count = total_responses - top_5_count
                
                chart_data = []
                for label, count in top_5:
                    pct = int(round((count / total_responses) * 100)) if total_responses else 0
                    chart_data.append({'label': label, 'count': count, 'percent': pct})
                
                if others_count > 0:
                    pct = int(round((others_count / total_responses) * 100)) if total_responses else 0
                    chart_data.append({'label': 'Others', 'count': others_count, 'percent': pct, 'is_other': True})
                
                metrics[field.id] = {
                    'total': total_responses,
                    'chart_data': chart_data
                }
    
    return render(request, "desk/forms/responses.html", {
        "dynamic_form": instance,
        "submissions": page_obj,
        "search_query": query,
        "fields": fields,
        "metrics": metrics
    })


@staff_required
def form_responses_export(request, pk):
    """
    Export all submissions for a DynamicForm to an Excel (.xlsx) file.
    """
    from core.models import DynamicForm
    import openpyxl
    from django.http import HttpResponse
    
    instance = get_object_or_404(DynamicForm, pk=pk)
    submissions = instance.submissions.all()
    fields = instance.fields.all()
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{instance.title}_responses_{instance.pk}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"
    
    # Header Row
    header = ['Date Submitted', 'Submitted By']
    field_ids = []
    
    for f in fields:
        # We export all fields except structural ones
        if f.field_type not in ['title_desc', 'page_break']:
            header.append(f.label)
            field_ids.append(str(f.id))
            
    ws.append(header)
    
    # Data Rows
    from core.excel_utils import safe_cell
    for sub in submissions:
        username = "Guest User"
        if sub.submitted_by:
            username = f"{sub.submitted_by.first_name} {sub.submitted_by.last_name}".strip()
            if not username:
                username = sub.submitted_by.username

        row = [
            sub.submitted_at.replace(tzinfo=None) if sub.submitted_at else "",
            safe_cell(username),
        ]

        for fid in field_ids:
            ans = sub.answers.get(fid, "")
            if isinstance(ans, list):
                row.append(safe_cell(" | ".join([str(a) for a in ans])))
            else:
                row.append(safe_cell(ans))

        ws.append(row)
        
    wb.save(response)
    return response

# =====================================================================
# API Endpoints
# =====================================================================

@login_required
@feature_required('api_access')
def api_users_list(request):
    """
    Returns a JSON list of active users to power the @mention autocomplete.
    """
    from django.http import JsonResponse
    from core.models import User
    
    users = User.objects.filter(is_active=True).values("username", "first_name", "last_name")
    
    data = []
    for u in users:
        # Prefer full name if exists, otherwise fallback to username
        name = f"{u['first_name']} {u['last_name']}".strip()
        display = name if name else u['username']
        
        data.append({
            "key": display,
            "value": u['username']
        })
        
    return JsonResponse(data, safe=False)

# =====================================================================
# Change Request Document (Surat Kronologi)
# =====================================================================


def _assert_case_not_locked(case):
    """
    Raises PermissionError if the case is locked due to an active approval workflow.
    Call this at the top of every view that mutates a CaseRecord.
    """
    if case.is_locked:
        raise PermissionError(
            f"Ticket {case.case_number} is locked while approval workflow is running "
            f"(status: {case.status}). No changes are allowed until the workflow completes."
        )


@login_required
@feature_required('whatsapp')
@require_POST
def toggle_wa_session(request, case_id):
    """
    Toggle the hold_wa_session flag for a specific case via HTMX.
    """
    case = get_object_or_404(CaseRecord, id=case_id)
    if not _check_confidential_access(case, request.user):
        return HttpResponseForbidden("You do not have access to this confidential ticket.")

    # Check permissions
    if request.user.role_access not in ['SuperAdmin', 'Manager', 'SupportDesk']:
        return HttpResponseForbidden("You do not have permission to perform this action.")
        
    case.hold_wa_session = not case.hold_wa_session
    case.save(update_fields=['hold_wa_session'])
    
    # Return the updated button HTML
    if case.hold_wa_session:
        return HttpResponse(f"""
        <button type="button" 
                hx-post="{request.path}" 
                hx-swap="outerHTML" 
                class="jk-btn jk-btn-sm border border-amber-200 bg-amber-50 text-amber-700 hover:bg-amber-100 shadow-sm"
                title="WA Session (Currently Held)">
            <svg class="w-3.5 h-3.5 mr-1" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10 9v6m4-6v6m7-3a9 9 0 11-18 0 9 9 0 0118 0z"></path></svg>
            Release Session
        </button>
        """)
    else:
        return HttpResponse(f"""
        <button type="button"
                hx-post="{request.path}"
                hx-swap="outerHTML"
                class="jk-btn jk-btn-sm"
                style="background:rgba(255,255,255,0.12); border:1px solid rgba(255,255,255,0.2); color:#cbd5e1;"
                title="WA Session (Currently Active)">
            <svg class="w-3.5 h-3.5 mr-1" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M14.752 11.168l-3.197-2.132A1 1 0 0010 9.87v4.263a1 1 0 001.555.832l3.197-2.132a1 1 0 000-1.664z"></path><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M21 12a9 9 0 11-18 0 9 9 0 0118 0z"></path></svg>
            Hold Session
        </button>
        """)


# =====================================================================
# Phase 2 — Document Template CRUD (Admin / SuperAdmin only)
# =====================================================================

@staff_required
def _save_template_fields(request, tmpl):
    """Parse fields_json from POST and (re)create all DocumentTemplateField rows."""
    import json as _json
    raw = request.POST.get("fields_json", "[]")
    try:
        fields_data = _json.loads(raw)
    except ValueError:
        fields_data = []

    tmpl.fields.all().delete()
    for idx, fdata in enumerate(fields_data):
        label = (fdata.get("label") or "").strip()
        if not label:
            continue
        DocumentTemplateField.objects.create(
            template=tmpl,
            order=idx + 1,
            label=label,
            placeholder=(fdata.get("placeholder") or "").strip(),
            default_content=(fdata.get("default_content") or "").strip(),
            is_required=bool(fdata.get("is_required", True)),
            field_type=(fdata.get("field_type") or "richtext").strip(),
            assigned_stage=(fdata.get("assigned_stage") or "").strip(),
            variable_name=(fdata.get("variable_name") or "").strip(),
            esign_role=(fdata.get("esign_role") or "").strip(),
            created_by=request.user,
            updated_by=request.user,
        )


def document_template_list(request):
    """List all DocumentTemplates."""
    query = request.GET.get("q", "").strip()
    qs = DocumentTemplate.objects.prefetch_related("categories", "fields").order_by("title")
    if query:
        qs = qs.filter(Q(title__icontains=query) | Q(description__icontains=query))
    return render(request, "desk/document_templates/list.html", {
        "templates": qs,
        "search_query": query,
    })


@staff_required
def document_template_create(request):
    """Create a new DocumentTemplate. SuperAdmin or Manager only."""
    if request.user.role_access not in [User.RoleAccess.SUPERADMIN, User.RoleAccess.MANAGER]:
        return HttpResponseForbidden("Only SuperAdmin and Manager can create document templates.")

    from cases.forms import DocumentTemplateForm
    if request.method == "POST":
        form = DocumentTemplateForm(request.POST)
        if form.is_valid():
            tmpl = form.save(commit=False)
            tmpl.created_by = request.user
            tmpl.updated_by = request.user
            tmpl.save()
            form.save_m2m()
            _save_template_fields(request, tmpl)
            messages.success(request, f"Template '{tmpl.title}' created successfully.")
            return redirect("desk:document_template_list")
    else:
        form = DocumentTemplateForm()

    return render(request, "desk/document_templates/form.html", {
        "form": form,
        "is_edit": False,
        "existing_fields_json": "[]",
    })


@staff_required
def document_template_edit(request, template_id):
    """Edit an existing DocumentTemplate."""
    if request.user.role_access not in [User.RoleAccess.SUPERADMIN, User.RoleAccess.MANAGER]:
        return HttpResponseForbidden("Only SuperAdmin and Manager can edit document templates.")

    import json as _json
    from cases.forms import DocumentTemplateForm
    tmpl = get_object_or_404(DocumentTemplate, id=template_id)

    if request.method == "POST":
        form = DocumentTemplateForm(request.POST, instance=tmpl)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            form.save_m2m()
            _save_template_fields(request, obj)
            messages.success(request, f"Template '{obj.title}' updated.")
            return redirect("desk:document_template_list")
    else:
        form = DocumentTemplateForm(instance=tmpl)

    existing_fields = list(
        tmpl.fields.order_by("order").values(
            "label", "placeholder", "default_content", "is_required", "field_type", "assigned_stage", "variable_name", "esign_role"
        )
    )
    return render(request, "desk/document_templates/form.html", {
        "form": form,
        "tmpl": tmpl,
        "is_edit": True,
        "existing_fields_json": _json.dumps(existing_fields),
    })


@staff_required
@require_POST
def document_template_delete(request, template_id):
    """Delete a DocumentTemplate."""
    if request.user.role_access != User.RoleAccess.SUPERADMIN:
        return HttpResponseForbidden("Only SuperAdmin can delete document templates.")
    tmpl = get_object_or_404(DocumentTemplate, id=template_id)
    title = tmpl.title
    tmpl.delete()
    messages.success(request, f"Template '{title}' deleted.")
    return redirect("desk:document_template_list")



# =====================================================================
# Ticket Category Management (Desk Admin UI)
# =====================================================================

@staff_required
def category_manage_list(request):
    """List all ticket categories grouped as parent → children."""
    if request.user.role_access not in [User.RoleAccess.SUPERADMIN, User.RoleAccess.MANAGER]:
        return HttpResponseForbidden("Access denied.")

    q = request.GET.get("q", "").strip()
    from django.db.models import Prefetch
    root_qs = CaseCategory.objects.filter(parent__isnull=True).prefetch_related(
        Prefetch("children", queryset=CaseCategory.objects.prefetch_related("document_templates")),
        "document_templates",
    ).order_by("name")
    if q:
        matched_ids = set(
            CaseCategory.objects.filter(name__icontains=q).values_list("id", flat=True)
        )
        parent_ids = set(
            CaseCategory.objects.filter(id__in=matched_ids, parent__isnull=False)
            .values_list("parent_id", flat=True)
        )
        root_qs = root_qs.filter(
            Q(id__in=matched_ids) | Q(id__in=parent_ids)
        )

    return render(request, "desk/categories/list.html", {
        "root_categories": root_qs,
        "search_query": q,
    })


@staff_required
def category_manage_create(request):
    """Create a new CaseCategory."""
    if request.user.role_access not in [User.RoleAccess.SUPERADMIN, User.RoleAccess.MANAGER]:
        return HttpResponseForbidden("Access denied.")

    from cases.forms import CaseCategoryForm
    if request.method == "POST":
        form = CaseCategoryForm(request.POST)
        if form.is_valid():
            cat = form.save(commit=False)
            cat.created_by = request.user
            cat.updated_by = request.user
            cat.save()
            messages.success(request, f"Category \"{cat.name}\" created.")
            return redirect("desk:category_list")
    else:
        initial = {}
        parent_id = request.GET.get("parent")
        if parent_id:
            parent_obj = CaseCategory.objects.filter(pk=parent_id, parent__isnull=True).first()
            if parent_obj:
                initial["parent"] = parent_obj
        form = CaseCategoryForm(initial=initial)

    return render(request, "desk/categories/form.html", {
        "form": form,
        "is_edit": False,
    })


@staff_required
def category_manage_edit(request, category_id):
    """Edit an existing CaseCategory."""
    if request.user.role_access not in [User.RoleAccess.SUPERADMIN, User.RoleAccess.MANAGER]:
        return HttpResponseForbidden("Access denied.")

    category = get_object_or_404(CaseCategory, id=category_id)
    from cases.forms import CaseCategoryForm
    if request.method == "POST":
        form = CaseCategoryForm(request.POST, instance=category)
        if form.is_valid():
            cat = form.save(commit=False)
            cat.updated_by = request.user
            cat.save()
            messages.success(request, f"Category \"{cat.name}\" updated.")
            return redirect("desk:category_list")
    else:
        form = CaseCategoryForm(instance=category)

    return render(request, "desk/categories/form.html", {
        "form": form,
        "category": category,
        "is_edit": True,
    })


@staff_required
def category_manage_delete(request, category_id):
    """Show confirmation or perform deletion of a CaseCategory."""
    if request.user.role_access != User.RoleAccess.SUPERADMIN:
        return HttpResponseForbidden("Only SuperAdmin can delete categories.")

    category = get_object_or_404(CaseCategory, id=category_id)
    ticket_count = CaseRecord.objects.filter(category=category).count()
    children_count = category.children.count()

    if request.method == "POST":
        if ticket_count > 0:
            messages.error(
                request,
                f'Cannot delete "{category.name}": {ticket_count} ticket(s) use this category.',
            )
            return redirect("desk:category_list")
        if children_count > 0:
            messages.error(
                request,
                f'Cannot delete "{category.name}": it has {children_count} sub-categor{"y" if children_count == 1 else "ies"}. Delete them first.',
            )
            return redirect("desk:category_list")
        name = category.name
        category.delete()
        messages.success(request, f"Category \"{name}\" deleted.")
        return redirect("desk:category_list")

    return render(request, "desk/categories/confirm_delete.html", {
        "category": category,
        "ticket_count": ticket_count,
        "children_count": children_count,
    })


def document_template_preview_html(request, template_id):
    """
    HTMX/AJAX endpoint. Renders the document body HTML with the submitted
    placeholder values and returns it for inline preview.
    """
    tmpl = get_object_or_404(DocumentTemplate, id=template_id)
    placeholders = tmpl.extract_placeholders()
    filled = {ph: request.POST.get(f"doc_{template_id}_{ph}", f"[{ph}]") for ph in placeholders}

    from cases.utils_pdf import render_document_html
    rendered_html = render_document_html(tmpl.body_html, filled)
    return HttpResponse(rendered_html)

# =====================================================================
# Change Request Document Views (Surat Kronologi)
# =====================================================================

def _notify_change_request_approvers(doc):
    """Post a system message for each pending approver with their magic-link URL."""
    for approval in doc.approvals.filter(
        status=ChangeRequestApproval.Status.PENDING
    ).select_related("approver"):
        approval_url = reverse("cases:change_request_approve", args=[approval.token])
        approver_name = approval.approver.get_full_name() or approval.approver.username
        try:
            Message.objects.create(
                case=doc.case,
                body=(
                    f"[Supporting Letter Approval — Rev.{doc.revision_number}]\n"
                    f"Ticket {doc.case.case_number} requires approval from {approver_name}.\n"
                    f"Review: {approval_url}"
                ),
                direction=Message.Direction.OUTBOUND,
                channel=Message.Channel.WEB,
                is_system=True,
            )
        except Exception:
            pass


def _finalize_approved_change_request(doc):
    """
    Generate the change request PDF, attach it to the ticket as a system
    Message+Attachment, and transition the case to OPEN.
    Returns True on success.
    """
    from cases.utils_pdf import generate_change_request_pdf
    from django.core.files.base import ContentFile

    pdf_bytes = generate_change_request_pdf(doc)
    if pdf_bytes is None:
        return False

    safe_case_number = doc.case.case_number.replace("/", "-").replace(" ", "_")
    doc_title_slug = (
        doc.document_template.title.replace(" ", "_")
        if doc.document_template else "Supporting_Letter"
    )
    filename = f"{doc_title_slug}_{safe_case_number}_Approved.pdf"
    doc.generated_pdf.save(filename, ContentFile(pdf_bytes), save=True)
    doc.status = ChangeRequestDocument.Status.APPROVED
    doc.save(update_fields=["status"])

    case = doc.case
    msg = Message.objects.create(
        case=case,
        body=(
            f"[Supporting Letter Approved — Rev.{doc.revision_number}]\n"
            f"The supporting letter has been fully approved and attached."
        ),
        direction=Message.Direction.OUTBOUND,
        channel=Message.Channel.WEB,
        is_system=True,
    )

    attach = Attachment(
        message=msg,
        original_filename=filename,
        mime_type="application/pdf",
        file_size=len(pdf_bytes),
    )
    attach.file.save(filename, ContentFile(pdf_bytes), save=True)

    was_pending = case.status == CaseRecord.Status.PENDING_APPROVAL
    if was_pending:
        case.status = CaseRecord.Status.OPEN
        case.save(update_fields=["status"])
        # Now that the ticket is active, dispatch the desk notification that
        # was deferred when the ticket was first submitted with pending approvers.
        try:
            from gateways.tasks import _dispatch_new_ticket_notifs
            _dispatch_new_ticket_notifs(str(case.id))
        except Exception as e:
            logger.error("Failed to dispatch new ticket notifications for case %s: %s", case.id, e)

    return True


def portal_change_request_new(request, case_id):
    """
    Create a new ChangeRequestDocument for a ticket.
    GET (no template_id): show template picker if multiple templates linked to category.
    GET (with template_id): show dynamic field form for the selected template.
    POST: create the document and route it to selected approvers (or auto-approve).
    """
    case = get_object_or_404(CaseRecord, id=case_id)
    guest_token = request.GET.get("token", request.POST.get("token", "")).strip()

    is_owner = (
        (request.user.is_authenticated and request.user.email == case.requester_email)
        or (guest_token and case.guest_token == guest_token)
    )
    if not is_owner:
        return HttpResponseForbidden("Access denied.")

    if not case.category.enable_change_request:
        return HttpResponseForbidden("Supporting letter is not enabled for this category.")

    active_doc = case.change_requests.filter(
        status__in=[
            ChangeRequestDocument.Status.PENDING_APPROVAL,
            ChangeRequestDocument.Status.APPROVED,
        ]
    ).first()
    if active_doc:
        messages.info(request, "A supporting letter is already active or approved for this ticket.")
        query = f"?token={guest_token}" if guest_token else ""
        return redirect(f"{reverse('cases:portal_change_request_detail', args=[active_doc.id])}{query}")

    available_approvers = User.objects.filter(is_active=True).order_by("first_name", "last_name")
    available_templates = (
        DocumentTemplate.objects
        .filter(categories=case.category)
        .prefetch_related("fields")
        .order_by("title")
    )

    def _quill_has_content(html):
        text = re.sub(r'<[^>]+>', '', html or '').strip()
        return bool(text) or '<img' in (html or '')

    def _build_approval_chain(doc, approver_ids):
        has_approvers = bool(approver_ids)
        if has_approvers:
            for order, uid in enumerate(approver_ids, start=1):
                approver = User.objects.filter(id=uid).first()
                if approver:
                    ChangeRequestApproval.objects.create(document=doc, approver=approver, order=order)
            case.status = CaseRecord.Status.PENDING_APPROVAL
            case.save(update_fields=["status"])
            _notify_change_request_approvers(doc)
            messages.success(request, "Supporting letter submitted and sent to approvers.")
        else:
            _finalize_approved_change_request(doc)
            messages.success(request, "Supporting letter approved and attached to the ticket.")
        return has_approvers

    if request.method == "POST":
        template_id = request.POST.get("template_id", "").strip()
        approver_ids = request.POST.getlist("approvers")

        if template_id:
            # Dynamic template flow
            selected_template = available_templates.filter(id=template_id).first()
            if not selected_template:
                messages.error(request, "Invalid document template selected.")
                return render(request, "client/change_request_form.html", {
                    "case": case, "available_approvers": available_approvers,
                    "available_templates": available_templates, "guest_token": guest_token,
                })

            field_values = []
            has_any_content = False
            for field in selected_template.ordered_fields():
                val = request.POST.get(f"field_{field.order}", "").strip()
                if field.is_required and not _quill_has_content(val):
                    messages.error(request, f"Field '{field.label}' is required.")
                    return render(request, "client/change_request_form.html", {
                        "case": case, "available_approvers": available_approvers,
                        "available_templates": available_templates,
                        "selected_template": selected_template,
                        "guest_token": guest_token,
                    })
                field_values.append({"label": field.label, "value": val})
                if _quill_has_content(val):
                    has_any_content = True

            if not has_any_content:
                messages.error(request, "Please fill in at least one field.")
                return render(request, "client/change_request_form.html", {
                    "case": case, "available_approvers": available_approvers,
                    "available_templates": available_templates,
                    "selected_template": selected_template,
                    "guest_token": guest_token,
                })

            doc = ChangeRequestDocument.objects.create(
                case=case,
                document_template=selected_template,
                field_values=field_values,
                status=(
                    ChangeRequestDocument.Status.PENDING_APPROVAL
                    if approver_ids else ChangeRequestDocument.Status.APPROVED
                ),
                submitted_at=timezone.now(),
            )
        else:
            # Legacy 2-field flow (fallback when no templates configured for category)
            request_change = request.POST.get("request_change", "").strip()
            chronology = request.POST.get("chronology", "").strip()
            if not _quill_has_content(request_change) or not _quill_has_content(chronology):
                messages.error(request, "Both fields are required.")
                return render(request, "client/change_request_form.html", {
                    "case": case, "available_approvers": available_approvers,
                    "available_templates": available_templates, "guest_token": guest_token,
                })
            doc = ChangeRequestDocument.objects.create(
                case=case,
                request_change=request_change,
                chronology=chronology,
                status=(
                    ChangeRequestDocument.Status.PENDING_APPROVAL
                    if approver_ids else ChangeRequestDocument.Status.APPROVED
                ),
                submitted_at=timezone.now(),
            )

        _build_approval_chain(doc, approver_ids)
        query = f"?token={guest_token}" if guest_token else ""
        return redirect(f"{reverse('cases:portal_change_request_detail', args=[doc.id])}{query}")

    # GET — resolve which template to show
    selected_template = None
    template_id_param = request.GET.get("template_id", "").strip()
    if template_id_param:
        selected_template = available_templates.filter(id=template_id_param).first()
    elif available_templates.count() == 1:
        selected_template = available_templates.first()

    return render(request, "client/change_request_form.html", {
        "case": case,
        "available_approvers": available_approvers,
        "available_templates": available_templates,
        "selected_template": selected_template,
        "guest_token": guest_token,
    })


def portal_change_request_detail(request, doc_id):
    """Show the status and approval timeline for a ChangeRequestDocument."""
    doc = get_object_or_404(ChangeRequestDocument, id=doc_id)
    case = doc.case
    guest_token = request.GET.get("token", "").strip()

    is_owner = (
        (request.user.is_authenticated and request.user.email == case.requester_email)
        or (guest_token and case.guest_token == guest_token)
    )
    staff_access = request.user.is_authenticated and hasattr(request.user, "role_access")
    if not is_owner and not staff_access:
        return HttpResponseForbidden("Access denied.")

    return render(request, "client/change_request_status.html", {
        "doc": doc,
        "case": case,
        "approvals": doc.approvals.order_by("order").select_related("approver"),
        "guest_token": guest_token,
    })


def portal_change_request_revise(request, doc_id):
    """
    Allow the requester to submit a revised change request after rejection.
    Creates a new ChangeRequestDocument with revision_number incremented.
    """
    doc = get_object_or_404(ChangeRequestDocument, id=doc_id)
    case = doc.case
    guest_token = request.GET.get("token", request.POST.get("token", "")).strip()

    is_owner = (
        (request.user.is_authenticated and request.user.email == case.requester_email)
        or (guest_token and case.guest_token == guest_token)
    )
    if not is_owner:
        return HttpResponseForbidden("Access denied.")

    if doc.status != ChangeRequestDocument.Status.REJECTED:
        messages.error(request, "This change request is not in a rejected state.")
        return redirect("cases:portal_change_request_detail", doc_id=doc.id)

    available_approvers = User.objects.filter(is_active=True).order_by("first_name", "last_name")
    selected_template = doc.document_template

    def _quill_has_content(html):
        text = re.sub(r'<[^>]+>', '', html or '').strip()
        return bool(text) or '<img' in (html or '')

    if request.method == "POST":
        approver_ids = request.POST.getlist("approvers")

        if selected_template:
            field_values = []
            for field in selected_template.ordered_fields():
                val = request.POST.get(f"field_{field.order}", "").strip()
                if field.is_required and not _quill_has_content(val):
                    messages.error(request, f"Field '{field.label}' is required.")
                    return render(request, "client/change_request_form.html", {
                        "case": case, "doc": doc, "available_approvers": available_approvers,
                        "selected_template": selected_template, "is_revision": True,
                        "guest_token": guest_token,
                        "prefill_field_values": doc.field_values,
                    })
                field_values.append({"label": field.label, "value": val})

            new_doc = ChangeRequestDocument.objects.create(
                case=case,
                document_template=selected_template,
                field_values=field_values,
                revision_number=doc.revision_number + 1,
                status=(
                    ChangeRequestDocument.Status.PENDING_APPROVAL
                    if approver_ids else ChangeRequestDocument.Status.APPROVED
                ),
                submitted_at=timezone.now(),
            )
        else:
            request_change = request.POST.get("request_change", "").strip()
            chronology = request.POST.get("chronology", "").strip()
            if not _quill_has_content(request_change) or not _quill_has_content(chronology):
                messages.error(request, "Both fields are required.")
                return render(request, "client/change_request_form.html", {
                    "case": case, "doc": doc, "available_approvers": available_approvers,
                    "is_revision": True, "guest_token": guest_token,
                    "prefill_request_change": doc.request_change,
                    "prefill_chronology": doc.chronology,
                })
            new_doc = ChangeRequestDocument.objects.create(
                case=case,
                request_change=request_change,
                chronology=chronology,
                revision_number=doc.revision_number + 1,
                status=(
                    ChangeRequestDocument.Status.PENDING_APPROVAL
                    if approver_ids else ChangeRequestDocument.Status.APPROVED
                ),
                submitted_at=timezone.now(),
            )

        if approver_ids:
            for order, uid in enumerate(approver_ids, start=1):
                approver = User.objects.filter(id=uid).first()
                if approver:
                    ChangeRequestApproval.objects.create(document=new_doc, approver=approver, order=order)
            case.status = CaseRecord.Status.PENDING_APPROVAL
            case.save(update_fields=["status"])
            _notify_change_request_approvers(new_doc)
            messages.success(request, f"Revision #{new_doc.revision_number} submitted.")
        else:
            _finalize_approved_change_request(new_doc)
            messages.success(request, f"Revision #{new_doc.revision_number} approved and attached.")

        query = f"?token={guest_token}" if guest_token else ""
        return redirect(f"{reverse('cases:portal_change_request_detail', args=[new_doc.id])}{query}")

    rejection_note = (
        doc.approvals.filter(status=ChangeRequestApproval.Status.REJECTED)
        .order_by("-acted_at")
        .values_list("notes", flat=True)
        .first()
    )
    return render(request, "client/change_request_form.html", {
        "case": case,
        "doc": doc,
        "available_approvers": available_approvers,
        "selected_template": selected_template,
        "is_revision": True,
        "rejection_note": rejection_note or "",
        "prefill_field_values": doc.field_values,
        "prefill_request_change": doc.request_change,
        "prefill_chronology": doc.chronology,
        "guest_token": guest_token,
    })


def change_request_approve(request, token):
    """
    Magic-link page for approvers to review and act on a ChangeRequestDocument.
    GET: show the document with approve/reject form.
    POST: process the approval or rejection.

    Access rules:
    - Authenticated user: must be the exact approver assigned to this token.
    - Unauthenticated user: the magic-link token in the URL is the credential.
    """
    approval = get_object_or_404(ChangeRequestApproval, token=token)
    doc = approval.document
    case = doc.case

    # Logged-in users who are NOT the assigned approver are denied.
    if request.user.is_authenticated and request.user != approval.approver:
        return render(request, "client/change_request_approval.html", {
            "approval": approval, "doc": doc, "case": case, "access_denied": True,
        })

    if approval.status != ChangeRequestApproval.Status.PENDING:
        return render(request, "client/change_request_approval.html", {
            "approval": approval, "doc": doc, "case": case, "already_acted": True,
        })

    if approval.is_token_expired:
        return render(request, "client/change_request_approval.html", {
            "approval": approval, "doc": doc, "case": case, "token_expired": True,
        })

    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        notes = request.POST.get("notes", "").strip()
        now = timezone.now()

        if action == "approve":
            approval.status = ChangeRequestApproval.Status.APPROVED
            approval.notes = notes
            approval.acted_at = now
            approval.save(update_fields=["status", "notes", "acted_at"])

            if doc.is_fully_approved:
                _finalize_approved_change_request(doc)
            else:
                _notify_change_request_approvers(doc)

            return render(request, "client/change_request_approval.html", {
                "approval": approval, "doc": doc, "case": case,
                "success": True, "action_done": "approved",
            })

        elif action == "reject":
            if not notes:
                messages.error(request, "Please provide a reason for rejection.")
                return render(request, "client/change_request_approval.html", {
                    "approval": approval, "doc": doc, "case": case,
                })

            approval.status = ChangeRequestApproval.Status.REJECTED
            approval.notes = notes
            approval.acted_at = now
            approval.save(update_fields=["status", "notes", "acted_at"])

            doc.status = ChangeRequestDocument.Status.REJECTED
            doc.save(update_fields=["status"])

            if case.status == CaseRecord.Status.PENDING_APPROVAL:
                case.status = CaseRecord.Status.REVISION_REQUIRED
                case.save(update_fields=["status"])

            try:
                approver_name = approval.approver.get_full_name() or approval.approver.username
                Message.objects.create(
                    case=case,
                    body=(
                        f"[Supporting Letter Rejected — Rev.{doc.revision_number}]\n"
                        f"Rejected by {approver_name}.\nReason: {notes}"
                    ),
                    direction=Message.Direction.OUTBOUND,
                    channel=Message.Channel.WEB,
                    is_system=True,
                )
            except Exception:
                pass

            return render(request, "client/change_request_approval.html", {
                "approval": approval, "doc": doc, "case": case,
                "success": True, "action_done": "rejected",
            })

    return render(request, "client/change_request_approval.html", {
        "approval": approval, "doc": doc, "case": case,
    })