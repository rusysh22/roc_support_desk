"""
Core App — User Master Data Views
====================
Provides CRUD operations for User management.
Restricted to SuperAdmin only.
"""
import io
import secrets
import string

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .excel_utils import safe_cell
from .forms import UserAdminForm
from .models import User

# Roles that count against the max_agents license limit
AGENT_ROLES = {User.RoleAccess.SUPPORTDESK, User.RoleAccess.MANAGER}


def _get_agent_limit():
    """
    Return (current_count, max_allowed) for agent-role users.
    Falls back to unlimited (None) if licensing is unavailable.
    """
    try:
        from licensing.models import LicenseRecord
        license_obj = LicenseRecord.get_current()
        max_allowed = license_obj.max_agents
    except Exception:
        max_allowed = None

    current_count = User.objects.filter(role_access__in=AGENT_ROLES, is_active=True).count()
    return current_count, max_allowed


def _generate_password(length=16):
    """Generate a cryptographically random password."""
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def superadmin_required(view_func):
    """Decorator to restrict access to SuperAdmin only."""
    from functools import wraps

    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        if getattr(request.user, "role_access", None) != User.RoleAccess.SUPERADMIN:
            return HttpResponseForbidden("Access denied. SuperAdmin privileges required.")
        return view_func(request, *args, **kwargs)

    return _wrapped


@superadmin_required
def user_list(request):
    """List all users."""
    search = request.GET.get("q", "").strip()
    role_filter = request.GET.get("role", "")
    status_filter = request.GET.get("status", "active") # Default to active

    qs = User.objects.all().order_by("login_username")

    if status_filter == "active":
        qs = qs.filter(is_active=True)
    elif status_filter == "archived":
        qs = qs.filter(is_active=False)

    if search:
        qs = qs.filter(
            Q(username__icontains=search)
            | Q(login_username__icontains=search)
            | Q(email__icontains=search)
            | Q(nik__icontains=search)
        )
    if role_filter:
        qs = qs.filter(role_access=role_filter)

    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get("page", 1))

    current_agents, max_agents = _get_agent_limit()
    return render(request, "desk/users/list.html", {
        "users": page,
        "search_query": search,
        "role_filter": role_filter,
        "status_filter": status_filter,
        "role_choices": User.RoleAccess.choices,
        "current_agents": current_agents,
        "max_agents": max_agents,
        "agent_limit_reached": max_agents is not None and current_agents >= max_agents,
    })


@superadmin_required
def user_create(request):
    """Create a new user."""
    current_agents, max_agents = _get_agent_limit()

    if request.method == "POST":
        form = UserAdminForm(request.POST)
        if form.is_valid():
            new_role = form.cleaned_data.get('role_access')
            if new_role in AGENT_ROLES and max_agents is not None and current_agents >= max_agents:
                messages.error(
                    request,
                    f"Agent limit reached ({current_agents}/{max_agents}). "
                    "Upgrade your license to add more Support Desk or Manager accounts."
                )
                return render(request, "desk/users/form.html", {
                    "form": form,
                    "is_edit": False,
                    "current_agents": current_agents,
                    "max_agents": max_agents,
                })
            user = form.save()
            messages.success(request, f'User "{user.login_username}" created successfully.')
            return redirect("users_desk:user_list")
    else:
        form = UserAdminForm()

    return render(request, "desk/users/form.html", {
        "form": form,
        "is_edit": False,
        "current_agents": current_agents,
        "max_agents": max_agents,
    })


@superadmin_required
def user_edit(request, pk):
    """Edit an existing user."""
    user_obj = get_object_or_404(User, pk=pk)
    current_agents, max_agents = _get_agent_limit()

    if request.method == "POST":
        form = UserAdminForm(request.POST, instance=user_obj)
        if form.is_valid():
            new_role = form.cleaned_data.get('role_access')
            old_role = user_obj.role_access
            # Only check limit when promoting a non-agent to an agent role
            if (new_role in AGENT_ROLES and old_role not in AGENT_ROLES
                    and max_agents is not None and current_agents >= max_agents):
                messages.error(
                    request,
                    f"Agent limit reached ({current_agents}/{max_agents}). "
                    "Upgrade your license to promote this user to a Support Desk or Manager role."
                )
                return render(request, "desk/users/form.html", {
                    "form": form,
                    "is_edit": True,
                    "user_obj": user_obj,
                    "current_agents": current_agents,
                    "max_agents": max_agents,
                })
            form.save()
            messages.success(request, f'User "{user_obj.login_username}" updated successfully.')
            return redirect("users_desk:user_list")
    else:
        form = UserAdminForm(instance=user_obj)

    return render(request, "desk/users/form.html", {
        "form": form,
        "is_edit": True,
        "user_obj": user_obj,
        "current_agents": current_agents,
        "max_agents": max_agents,
    })


@superadmin_required
@require_POST
def user_delete(request, pk):
    """Delete a user."""
    user_obj = get_object_or_404(User, pk=pk)
    
    if user_obj == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect("users_desk:user_list")
        
    username = user_obj.login_username
    user_obj.delete()
    messages.success(request, f'User "{username}" deleted successfully.')
    return redirect("users_desk:user_list")


# ---------------------------------------------------------------------------
# Excel columns used both for export and as the import template
# ---------------------------------------------------------------------------
EXCEL_HEADERS = [
    "login_username",
    "display_name",
    "email",
    "nik",
    "role_access",
    "initials",
    "can_handle_confidential",
]


def _make_xlsx_response(filename):
    """Return an HttpResponse pre-configured for xlsx download."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _style_header_row(sheet):
    """Apply a dark-indigo header style to row 1 of the worksheet."""
    header_fill = PatternFill(start_color="3730A3", end_color="3730A3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 20


@superadmin_required
def user_export(request):
    """Export selected users (by checkbox IDs) to Excel (.xlsx).
    If no IDs supplied, exports ALL users."""
    selected_ids = request.POST.getlist("selected_ids") or request.GET.getlist("ids")

    qs = User.objects.all().order_by("login_username")
    if selected_ids:
        qs = qs.filter(pk__in=selected_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Header row
    ws.append(EXCEL_HEADERS)
    _style_header_row(ws)

    # Auto column widths
    col_widths = {h: len(h) + 4 for h in EXCEL_HEADERS}

    # Data rows
    for u in qs:
        row = [
            safe_cell(u.login_username),
            safe_cell(u.username),
            safe_cell(u.email),
            safe_cell(u.nik or ""),
            safe_cell(u.role_access),
            safe_cell(u.initials),
            "TRUE" if u.can_handle_confidential else "FALSE",
        ]
        ws.append(row)
        for idx, val in enumerate(row):
            col_widths[EXCEL_HEADERS[idx]] = max(col_widths[EXCEL_HEADERS[idx]], len(str(val or "")) + 2)

    for idx, header in enumerate(EXCEL_HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = col_widths[header]

    # Write to response
    response = _make_xlsx_response("users_export.xlsx")
    wb.save(response)
    from .models import AuditLog
    AuditLog.log(AuditLog.Action.EXPORT, request=request,
                 details={"type": "user_list", "filename": "users_export.xlsx"})
    return response


@superadmin_required
def user_import_template(request):
    """Download a blank Excel template with the correct column headers and one sample row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    ws.append(EXCEL_HEADERS)
    _style_header_row(ws)

    # One sample row
    ws.append(["jdoe", "John Doe", "jdoe@example.com", "EMP001", "SupportDesk", "jd", "FALSE"])

    # Add a Notes sheet explaining each column
    notes_ws = wb.create_sheet(title="Notes")
    notes_ws.append(["Column", "Description", "Required", "Valid Values"])
    _style_header_row(notes_ws)
    valid_roles = ", ".join(r[0] for r in User.RoleAccess.choices)
    notes_data = [
        ["login_username", "Unique username used to log in", "Yes", "-"],
        ["display_name", "Full name displayed in the system", "Yes", "-"],
        ["email", "Unique email address", "Yes", "-"],
        ["nik", "Employee ID (Nomor Induk Karyawan)", "No", "-"],
        ["role_access", "Permission level", "Yes", valid_roles],
        ["initials", "Short initials (e.g. jd)", "Yes", "-"],
        ["can_handle_confidential", "Access to confidential tickets", "No", "TRUE / FALSE"],
    ]
    for r in notes_data:
        notes_ws.append(r)
    for col in notes_ws.columns:
        notes_ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4

    # Auto column widths for main sheet
    for idx, header in enumerate(EXCEL_HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(len(header) + 4, 20)

    response = _make_xlsx_response("user_import_template.xlsx")
    wb.save(response)
    return response


@superadmin_required
def user_import(request):
    """Import users from an uploaded Excel (.xlsx) file.
    Duplicates (existing login_username, email, or nik) — entire import fails, nothing saved."""
    if request.method != "POST":
        return redirect("users_desk:user_list")

    xlsx_file = request.FILES.get("csv_file")  # same input name, different format now
    if not xlsx_file:
        messages.error(request, "No file uploaded. Please select an Excel (.xlsx) file.")
        return redirect("users_desk:user_list")

    if not xlsx_file.name.endswith(".xlsx"):
        messages.error(request, "Invalid file type. Please upload a .xlsx Excel file.")
        return redirect("users_desk:user_list")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"Could not read the Excel file: {e}")
        return redirect("users_desk:user_list")

    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 2:
        messages.error(request, "The uploaded file has no data rows (only headers or empty).")
        return redirect("users_desk:user_list")

    # Read header from first row, map to expected fields
    header = [str(h).strip().lower() if h else "" for h in rows_raw[0]]
    expected = [h.lower() for h in EXCEL_HEADERS]
    missing = [h for h in expected if h not in header]
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}. Please use the official template.")
        return redirect("users_desk:user_list")

    def get_col(row, field):
        idx = header.index(field.lower())
        return str(row[idx]).strip() if row[idx] is not None else ""

    data_rows = rows_raw[1:]
    if not data_rows:
        messages.error(request, "The uploaded file is empty.")
        return redirect("users_desk:user_list")

    # --- Validate ALL rows first before touching the database ---
    valid_roles = {r[0] for r in User.RoleAccess.choices}
    errors = []

    for i, row in enumerate(data_rows, start=2):  # start=2 because row 1 is header
        lu = get_col(row, "login_username")
        email = get_col(row, "email")
        nik = get_col(row, "nik")
        role = get_col(row, "role_access")
        initials = get_col(row, "initials")
        display_name = get_col(row, "display_name")

        if not lu:
            errors.append(f"Row {i}: login_username is required.")
        if not email:
            errors.append(f"Row {i}: email is required.")
        if not display_name:
            errors.append(f"Row {i}: display_name is required.")
        if not initials:
            errors.append(f"Row {i}: initials is required.")
        if role and role not in valid_roles:
            errors.append(f"Row {i}: invalid role_access '{role}'. Valid options: {', '.join(valid_roles)}.")

        # Check for duplicates in the existing database
        if lu and User.objects.filter(login_username=lu).exists():
            errors.append(f"Row {i}: login_username '{lu}' already exists.")
        if email and User.objects.filter(email=email).exists():
            errors.append(f"Row {i}: email '{email}' already exists.")
        if nik and User.objects.filter(nik=nik).exists():
            errors.append(f"Row {i}: NIK '{nik}' already exists.")

    if errors:
        for err in errors:
            messages.error(request, err)
        return redirect("users_desk:user_list")

    # --- Check agent limit before creating ---
    current_agents, max_agents = _get_agent_limit()
    if max_agents is not None:
        new_agents_in_file = sum(
            1 for row in data_rows
            if (get_col(row, "role_access") or User.RoleAccess.SUPPORTDESK) in AGENT_ROLES
        )
        if current_agents + new_agents_in_file > max_agents:
            messages.error(
                request,
                f"Import would exceed your agent limit. "
                f"Current: {current_agents}/{max_agents} agents. "
                f"This file contains {new_agents_in_file} agent-role user(s). "
                "Remove non-agent roles or upgrade your license."
            )
            return redirect("users_desk:user_list")

    # --- All OK — create users inside an atomic transaction ---
    created_count = 0
    generated_passwords = []  # [(login_username, display_name, password)]

    try:
        with transaction.atomic():
            for row in data_rows:
                lu = get_col(row, "login_username")
                confidence_raw = get_col(row, "can_handle_confidential").upper()
                can_confidential = confidence_raw in ("TRUE", "1", "YES")
                role = get_col(row, "role_access") or User.RoleAccess.SUPPORTDESK
                pw = _generate_password()

                user_obj = User(
                    login_username=lu,
                    username=get_col(row, "display_name"),
                    email=get_col(row, "email"),
                    nik=get_col(row, "nik") or None,
                    role_access=role,
                    initials=get_col(row, "initials"),
                    can_handle_confidential=can_confidential,
                    must_change_password=True,
                )
                user_obj.set_password(pw)
                user_obj.save()
                generated_passwords.append((lu, get_col(row, "display_name"), pw))
                created_count += 1
    except Exception as e:
        messages.error(request, f"Import failed due to an unexpected error: {e}")
        return redirect("users_desk:user_list")

    # Export generated passwords as a downloadable Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Generated Passwords"
    ws.append(["login_username", "display_name", "generated_password"])
    _style_header_row(ws)
    for lu, dn, pw in generated_passwords:
        ws.append([lu, dn, pw])

    response = _make_xlsx_response(f"imported_users_passwords_{created_count}.xlsx")
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response.write(buf.read())
    from .models import AuditLog
    AuditLog.log(AuditLog.Action.BULK_IMPORT, request=request,
                 details={"created_count": created_count})
    messages.success(request, f"Successfully imported {created_count} user(s). A password sheet has been downloaded — store it securely and delete after distribution.")
    return response


@superadmin_required
@require_POST
def user_bulk_delete(request):
    """Bulk delete selected users. Skips the currently logged-in user."""
    selected_ids = request.POST.getlist("selected_ids")
    if not selected_ids:
        messages.warning(request, "No users selected for deletion.")
        return redirect("users_desk:user_list")

    qs = User.objects.filter(pk__in=selected_ids).exclude(pk=request.user.pk)
    count = qs.count()
    qs.delete()

    skipped = len(selected_ids) - count
    msg = f"Deleted {count} user(s)."
    if skipped:
        msg += f" {skipped} skipped (your own account cannot be deleted)."
    messages.success(request, msg)
    return redirect("users_desk:user_list")


# ---------------------------------------------------------------------------
# Password reset helpers
# ---------------------------------------------------------------------------

PASS_HEADERS = ["login_username", "display_name", "new_password"]
UPDATE_HEADERS = EXCEL_HEADERS + ["new_password"]


@superadmin_required
@require_POST
def user_export_password_template(request):
    """Export selected users as a password-reset Excel template (login_username + new_password)."""
    selected_ids = request.POST.getlist("selected_ids")
    if not selected_ids:
        messages.warning(request, "No users selected.")
        return redirect("users_desk:user_list")

    qs = User.objects.filter(pk__in=selected_ids).order_by("login_username")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Password Reset"
    ws.append(PASS_HEADERS)
    _style_header_row(ws)

    for u in qs:
        ws.append([u.login_username, u.username, ""])

    for idx, w in enumerate([22, 28, 22], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    notes_ws = wb.create_sheet("Notes")
    notes_ws.append(["Column", "Description"])
    _style_header_row(notes_ws)
    notes_ws.append(["login_username", "Identifier — DO NOT change"])
    notes_ws.append(["display_name", "For reference only — not updated"])
    notes_ws.append(["new_password", "New password to set. Leave blank to skip that user."])
    for col in notes_ws.columns:
        notes_ws.column_dimensions[col[0].column_letter].width = (
            max(len(str(c.value or "")) for c in col) + 4
        )

    response = _make_xlsx_response("password_reset_template.xlsx")
    wb.save(response)
    return response


@superadmin_required
@require_POST
def user_import_bulk_password(request):
    """Import Excel with login_username + new_password. Updates passwords only."""
    xlsx_file = request.FILES.get("csv_file")
    if not xlsx_file:
        messages.error(request, "No file uploaded.")
        return redirect("users_desk:user_list")
    if not xlsx_file.name.endswith(".xlsx"):
        messages.error(request, "Please upload a .xlsx file.")
        return redirect("users_desk:user_list")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"Could not read file: {e}")
        return redirect("users_desk:user_list")

    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 2:
        messages.error(request, "File has no data rows.")
        return redirect("users_desk:user_list")

    header = [str(h).strip().lower() if h else "" for h in rows_raw[0]]
    if "login_username" not in header or "new_password" not in header:
        messages.error(request, "File must have 'login_username' and 'new_password' columns. Use the exported template.")
        return redirect("users_desk:user_list")

    lu_idx = header.index("login_username")
    pw_idx = header.index("new_password")

    updated = skipped = 0
    not_found = []

    for i, row in enumerate(rows_raw[1:], start=2):
        lu = str(row[lu_idx]).strip() if row[lu_idx] else ""
        pw = str(row[pw_idx]).strip() if row[pw_idx] else ""

        if not lu or not pw:
            skipped += 1
            continue

        try:
            user_obj = User.objects.get(login_username=lu)
            user_obj.set_password(pw)
            user_obj.save(update_fields=["password"])
            updated += 1
        except User.DoesNotExist:
            not_found.append(f"Row {i}: '{lu}' not found.")

    for err in not_found[:5]:
        messages.warning(request, err)
    if len(not_found) > 5:
        messages.warning(request, f"... and {len(not_found) - 5} more users not found.")

    messages.success(
        request,
        f"Password reset complete: {updated} updated, {skipped} skipped (blank password).",
    )
    return redirect("users_desk:user_list")


@superadmin_required
@require_POST
def user_bulk_reset_password(request):
    """Quick bulk reset: apply one password to all selected users."""
    selected_ids = request.POST.getlist("selected_ids")
    new_password = request.POST.get("new_password", "").strip()

    if not selected_ids:
        messages.warning(request, "No users selected.")
        return redirect("users_desk:user_list")
    if len(new_password) < 6:
        messages.error(request, "Password must be at least 6 characters.")
        return redirect("users_desk:user_list")

    qs = User.objects.filter(pk__in=selected_ids).exclude(pk=request.user.pk)
    count = 0
    for u in qs:
        u.set_password(new_password)
        u.save(update_fields=["password"])
        count += 1

    skipped = len(selected_ids) - count
    msg = f"Password reset for {count} user(s)."
    if skipped:
        msg += f" {skipped} skipped (your own account)."
    messages.success(request, msg)
    return redirect("users_desk:user_list")


# ---------------------------------------------------------------------------
# Update existing users via Excel
# ---------------------------------------------------------------------------

@superadmin_required
def user_export_update_template(request):
    """Export all users (or selected) with all editable fields + optional new_password column."""
    selected_ids = request.POST.getlist("selected_ids") if request.method == "POST" else request.GET.getlist("ids")
    qs = User.objects.all().order_by("login_username")
    if selected_ids:
        qs = qs.filter(pk__in=selected_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Update Users"
    ws.append(UPDATE_HEADERS)
    _style_header_row(ws)

    col_widths = {h: len(h) + 4 for h in UPDATE_HEADERS}
    for u in qs:
        row = [
            safe_cell(u.login_username),
            safe_cell(u.username),
            safe_cell(u.email),
            safe_cell(u.nik or ""),
            safe_cell(u.role_access),
            safe_cell(u.initials),
            "TRUE" if u.can_handle_confidential else "FALSE",
            "",  # new_password — blank = keep current
        ]
        ws.append(row)
        for idx, val in enumerate(row):
            col_widths[UPDATE_HEADERS[idx]] = max(col_widths[UPDATE_HEADERS[idx]], len(str(val or "")) + 2)

    for idx, h in enumerate(UPDATE_HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = col_widths[h]

    notes_ws = wb.create_sheet("Notes")
    notes_ws.append(["Column", "Description", "Notes"])
    _style_header_row(notes_ws)
    valid_roles = ", ".join(r[0] for r in User.RoleAccess.choices)
    for r in [
        ["login_username",           "User identifier",          "DO NOT change — used to match existing user"],
        ["display_name",             "Full display name",         "Will be updated if filled"],
        ["email",                    "Email address",             "Will be updated if filled"],
        ["nik",                      "Employee ID",               "Will be updated (blank = clear)"],
        ["role_access",              "Permission level",          f"Valid: {valid_roles}"],
        ["initials",                 "Short initials",            "Will be updated if filled"],
        ["can_handle_confidential",  "Confidential ticket access","TRUE / FALSE"],
        ["new_password",             "New password (optional)",   "Leave BLANK to keep current password"],
    ]:
        notes_ws.append(r)
    for col in notes_ws.columns:
        notes_ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4

    response = _make_xlsx_response("users_update_template.xlsx")
    wb.save(response)
    return response


@superadmin_required
@require_POST
def user_import_update(request):
    """Update existing users from Excel. Matches by login_username.
    - Known username  → update fields; update password if new_password is filled.
    - Unknown username → skip with warning (does NOT create new users).
    """
    xlsx_file = request.FILES.get("csv_file")
    if not xlsx_file:
        messages.error(request, "No file uploaded.")
        return redirect("users_desk:user_list")
    if not xlsx_file.name.endswith(".xlsx"):
        messages.error(request, "Please upload a .xlsx file.")
        return redirect("users_desk:user_list")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"Could not read file: {e}")
        return redirect("users_desk:user_list")

    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 2:
        messages.error(request, "File has no data rows.")
        return redirect("users_desk:user_list")

    header = [str(h).strip().lower() if h else "" for h in rows_raw[0]]
    if "login_username" not in header:
        messages.error(request, "Missing 'login_username' column. Please use the update template.")
        return redirect("users_desk:user_list")

    def get_col(row, field):
        if field not in header:
            return ""
        idx = header.index(field)
        return str(row[idx]).strip() if row[idx] is not None else ""

    valid_roles = {r[0] for r in User.RoleAccess.choices}
    updated = skipped = 0
    warnings = []

    try:
        with transaction.atomic():
            for i, row in enumerate(rows_raw[1:], start=2):
                lu = get_col(row, "login_username")
                if not lu:
                    skipped += 1
                    continue

                try:
                    user_obj = User.objects.get(login_username=lu)
                except User.DoesNotExist:
                    warnings.append(f"Row {i}: '{lu}' not found — skipped.")
                    skipped += 1
                    continue

                role = get_col(row, "role_access")
                if role and role not in valid_roles:
                    warnings.append(f"Row {i}: invalid role '{role}' for '{lu}' — skipped.")
                    skipped += 1
                    continue

                if "display_name" in header:
                    v = get_col(row, "display_name")
                    if v:
                        user_obj.username = v
                if "email" in header:
                    v = get_col(row, "email")
                    if v:
                        user_obj.email = v
                if "nik" in header:
                    user_obj.nik = get_col(row, "nik") or None
                if "role_access" in header and role:
                    user_obj.role_access = role
                if "initials" in header:
                    v = get_col(row, "initials")
                    if v:
                        user_obj.initials = v
                if "can_handle_confidential" in header:
                    raw = get_col(row, "can_handle_confidential").upper()
                    user_obj.can_handle_confidential = raw in ("TRUE", "1", "YES")

                new_pw = get_col(row, "new_password")
                if new_pw:
                    user_obj.set_password(new_pw)

                user_obj.save()
                updated += 1

    except Exception as e:
        messages.error(request, f"Update failed: {e}")
        return redirect("users_desk:user_list")

    for w in warnings[:5]:
        messages.warning(request, w)
    if len(warnings) > 5:
        messages.warning(request, f"... and {len(warnings) - 5} more rows skipped.")

    messages.success(request, f"Update complete: {updated} user(s) updated, {skipped} skipped.")
    return redirect("users_desk:user_list")


@superadmin_required
@require_POST
def user_bulk_archive(request):
    """Bulk archive selected users."""
    selected_ids = request.POST.getlist("selected_ids")
    if not selected_ids:
        messages.warning(request, "No users selected for archiving.")
        return redirect("users_desk:user_list")

    qs = User.objects.filter(pk__in=selected_ids).exclude(pk=request.user.pk)
    count = qs.count()
    qs.update(is_active=False)

    skipped = len(selected_ids) - count
    msg = f"Archived {count} user(s)."
    if skipped:
        msg += f" {skipped} skipped (you cannot archive yourself)."
    messages.success(request, msg)
    return redirect("users_desk:user_list")


@superadmin_required
@require_POST
def user_bulk_unarchive(request):
    """Bulk unarchive selected users."""
    selected_ids = request.POST.getlist("selected_ids")
    if not selected_ids:
        messages.warning(request, "No users selected for unarchiving.")
        return redirect("users_desk:user_list")

    qs = User.objects.filter(pk__in=selected_ids)
    count = qs.count()
    qs.update(is_active=True)

    messages.success(request, f"Unarchived {count} user(s).")
    return redirect("users_desk:user_list")


@superadmin_required
@require_POST
def user_archive(request, pk):
    """Archive a single user."""
    user_obj = get_object_or_404(User, pk=pk)
    if user_obj == request.user:
        messages.error(request, "You cannot archive your own account.")
    else:
        user_obj.is_active = False
        user_obj.save()
        messages.success(request, f'User "{user_obj.login_username}" archived.')
    return redirect("users_desk:user_list")


@superadmin_required
@require_POST
def user_unarchive(request, pk):
    """Unarchive a single user."""
    user_obj = get_object_or_404(User, pk=pk)
    user_obj.is_active = True
    user_obj.save()
    messages.success(request, f'User "{user_obj.login_username}" unarchived.')
    return redirect("users_desk:user_list")

